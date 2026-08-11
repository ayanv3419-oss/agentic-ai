"""core_system — FastAPI app, every HTTP route, startup hook.

Single-user local-first MVP — NO AUTHENTICATION. The app opens directly.

Run from the project root:
    python backend/main.py
or:
    uvicorn backend.app.core_system:app --port 8000 --reload
"""
from __future__ import annotations

import asyncio
import collections
import logging
import os
import sys
import threading
import time
from contextlib import asynccontextmanager
from pathlib import Path
from typing import Any
from uuid import uuid4

from fastapi import (
    APIRouter, Depends, FastAPI, File, Form, HTTPException, Request, UploadFile,
)
from fastapi.exceptions import RequestValidationError
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, RedirectResponse, StreamingResponse
from pydantic import BaseModel, ConfigDict, Field

from app.agents import DashboardAgent, DataCleanAgent
from app.tenant_context import Principal, require_principal
from app.coordinator import (
    EventEmitter,
    PresentationEmitter as _CoordinatorPresentation,
    TurnState,
    format_sse,
    run_query_turn,
)
from app.coordinator.llm import check_llm_health
from app.dynamic_ingest import (
    drop_all_dynamic_tables,
    ingest_workbook,
    list_dynamic_tables_for_tenant,
    reconcile_registry,
)
from app.database import engine_kind, engine_status
from app.time_engine import (
    invalidate_cache as invalidate_time_cache,
    resolve_dataset_date_tokens,
)
from app.infrastructure import (
    ALLOWED_TABLES,
    SAFE_MESSAGE,
    UploadError,
    archive_upload,
    bump_data_version,
    cache_size,
    count_rows,
    disconnect_upload,
    envelope,
    find_active_upload_by_file_hash,
    get_data_version,
    invalidate_all,
    list_uploads_meta,
    record_upload_meta,
    settings,
    unarchive_upload,
    uploads_dir,
)
from app.conversation_store import (
    delete_conversation,
    get_conversation,
    list_conversations,
)
from app.dedup import DEDUP_MODES, DEFAULT_DEDUP_MODE, compute_file_hash
from app import google_drive
from app.errors import (
    SEVERITIES,
    error_analytics,
    get_error,
    list_errors,
    log_error,
    resolve_error,
)
from app.monitoring import (
    init_sentry,
    instrument_fastapi,
    sentry_status,
    set_request_context,
)

# Single Coordinator orchestration - no v1/v2 dual path. The previous
# legacy analytics_engine and orchestrator_v2 packages have been
# removed; everything routes through app.coordinator.


log = logging.getLogger("agentic_ai.api")


# ===========================================================================
# 1. RATE LIMITER — in-memory token bucket (keyed by client IP)
# ===========================================================================

_RATE_LOCK = threading.Lock()
_RATE_BUCKETS: dict[str, collections.deque] = {}


def _client_ip(request: Request) -> str:
    fwd = request.headers.get("x-forwarded-for", "")
    if fwd:
        return fwd.split(",")[0].strip() or "unknown"
    if request.client and request.client.host:
        return request.client.host
    return "unknown"


def _rate_limit_check(
    user_key: str,
    *,
    bucket_namespace: str = "query",
    limit_per_minute: int | None = None,
) -> dict[str, Any] | None:
    now = time.time()
    window = 60.0
    limit = max(1, int(limit_per_minute or settings.rate_limit_per_minute))
    composite_key = f"{bucket_namespace}:{user_key}"
    with _RATE_LOCK:
        bucket = _RATE_BUCKETS.get(composite_key)
        if bucket is None:
            bucket = collections.deque(maxlen=limit + 1)
            _RATE_BUCKETS[composite_key] = bucket
        while bucket and (now - bucket[0]) > window:
            bucket.popleft()
        if len(bucket) >= limit:
            retry_in = max(1, int(window - (now - bucket[0])))
            return envelope(
                f"Rate limit hit ({limit}/min)",
                detail=f"Too many {bucket_namespace} requests. Retry in ~{retry_in}s.",
                kind="rate_limit",
                extra={
                    "retry_after_seconds": retry_in,
                    "limit_per_minute":    limit,
                    "namespace":           bucket_namespace,
                },
            )
        bucket.append(now)
    return None


# ===========================================================================
# 2. API ROUTES — all public, no auth
# ===========================================================================

api_router = APIRouter()


# ---------------------------------------------------------------------------
# Google Drive integration
#
# Single-user local-first OAuth. The app itself has no startup login — these
# routes power the "Connect Google Drive" card on the Upload page and let the
# user sync data files straight from Drive through the same DataCleanAgent
# ingestion pipeline as manual uploads. All Drive/OAuth logic lives in
# app/google_drive.py; these routes are thin adapters. The whole feature
# stays inert (login returns 501) until GOOGLE_CLIENT_ID / GOOGLE_CLIENT_SECRET
# are set in backend/.env.
# ---------------------------------------------------------------------------

class DriveSyncRequest(BaseModel):
    file_ids: list[str] = Field(default_factory=list)
    target: str = "sales"
    dedup_mode: str = "skip"


def _drive_not_configured() -> JSONResponse:
    return JSONResponse(
        status_code=501,
        content=envelope(
            "Google Drive integration not configured",
            detail="Set GOOGLE_CLIENT_ID and GOOGLE_CLIENT_SECRET in "
                   "backend/.env (OAuth client from Google Cloud Console) "
                   "to enable Drive sync.",
            kind="not_implemented",
        ),
    )


# ---------------------------------------------------------------------------
# App-login gate (matches the frontend's LoginGate component)
# ---------------------------------------------------------------------------
# This is NOT real auth — the platform is single-user local-first. The
# endpoint exists so the frontend's POST /auth/login resolves with a 2xx
# instead of a 404. Credentials are env-overridable; defaults match what
# the frontend ships with so a fresh deploy works without env edits.

_APP_LOGIN_USER = os.environ.get("APP_LOGIN_USER", "").strip()
_APP_LOGIN_PASSWORD = os.environ.get("APP_LOGIN_PASSWORD", "")


class LoginRequest(BaseModel):
    """Permissive login body — accepts a couple of common field names so
    different frontend builds slot in cleanly without a backend redeploy."""

    model_config = ConfigDict(extra="ignore")

    username: str | None = Field(default=None, max_length=120)
    user:     str | None = Field(default=None, max_length=120)
    password: str | None = Field(default=None, max_length=200)
    pass_:    str | None = Field(default=None, alias="pass", max_length=200)


class SignupRequest(BaseModel):
    """Signup body — ``{email, password}``. Extra fields ignored so a
    slightly-divergent frontend build still slots in."""

    model_config = ConfigDict(extra="ignore")

    email:    str | None = Field(default=None, max_length=200)
    password: str | None = Field(default=None, max_length=200)


class ForgotPasswordRequest(BaseModel):
    """Reset-request body — ``{email}``. Always answered with 200 regardless of
    whether the email exists, so it can't be used to enumerate accounts."""

    model_config = ConfigDict(extra="ignore")

    email: str | None = Field(default=None, max_length=200)


class ResetPasswordRequest(BaseModel):
    """Reset body — ``{token, password}`` where ``token`` is the signed,
    time-limited value carried in the emailed link's ``?reset_token=`` param."""

    model_config = ConfigDict(extra="ignore")

    token:    str | None = Field(default=None, max_length=4000)
    password: str | None = Field(default=None, max_length=200)


def _frontend_base_url(request: Request) -> str:
    """Best-effort origin to build the password-reset link against: explicit
    ``FRONTEND_BASE_URL`` env wins, else the request's ``Origin`` header, else
    the first configured CORS origin, else the request base URL. Trailing
    slashes trimmed so ``f"{base}?reset_token=..."`` is well-formed."""
    candidates = [
        os.environ.get("FRONTEND_BASE_URL", "").strip(),
        (request.headers.get("origin") or "").strip(),
    ]
    for origin in _raw_origins:
        if origin and origin != "*":
            candidates.append(origin)
    candidates.append(str(request.base_url))
    for c in candidates:
        if c and c != "*":
            return c.rstrip("/")
    return ""


@api_router.post("/auth/signup")
async def auth_signup(req: SignupRequest, request: Request):
    """Create a real per-user account and return a tenant-bearing token.

    ``{ email, password } -> { token }``. Delegates to ``identity.signup``
    (creates a ``users`` row + an owned ``tenants`` row, bcrypt-hashes the
    password) and mints a token from the resulting Principal. A duplicate
    email is a 409; bad input (empty email / short password) is a 400.

    Rate-limited per client IP, same as login, to blunt automated abuse.
    """
    rate_error = _rate_limit_check(_client_ip(request))
    if rate_error is not None:
        return JSONResponse(status_code=429, content=rate_error)

    from app import identity

    email = (req.email or "").strip()
    password = req.password or ""

    try:
        principal = await identity.signup(email, password)
    except identity.EmailExists:
        return JSONResponse(
            status_code=409,
            content=envelope(
                "Email already registered",
                detail="An account with that email already exists.",
                kind="conflict",
            ),
        )
    except ValueError as e:
        return JSONResponse(
            status_code=400,
            content=envelope(
                "Invalid signup",
                detail=str(e) or "Email and a password of at least 6 "
                                 "characters are required.",
                kind="validation",
            ),
        )

    log.info("auth/signup: created account for user=%r", email[:40])
    return {
        "ok": True,
        "authenticated": True,
        "username": principal.email,
        "token": identity.mint_token(principal),
        "token_ttl_hours": settings.auth_token_ttl_hours,
        "auth_enabled": settings.auth_enabled,
    }


@api_router.post("/auth/login")
async def auth_login(req: LoginRequest, request: Request) -> dict:
    """
    App-level login endpoint. Accepts ``{username, password}`` (also
    accepts ``{user, pass}`` as aliases for tolerance). Validates against
    the ``users`` table via ``identity.authenticate`` (self-serve accounts).

    On success returns ``{ok, username, token}``. The token is a short
    opaque string the frontend may forward as ``Authorization: Bearer``
    if it wants — the backend doesn't currently enforce it (single-user
    local-first), but having a token-shaped response keeps the wire
    contract future-compatible.

    Rate-limited per client IP to mitigate brute force.
    """
    rate_error = _rate_limit_check(_client_ip(request))
    if rate_error is not None:
        return JSONResponse(status_code=429, content=rate_error)

    username = (req.username or req.user or "").strip()
    password = req.password or req.pass_ or ""

    if not username or not password:
        return JSONResponse(
            status_code=400,
            content=envelope(
                "Missing credentials",
                detail="Both `username` and `password` are required.",
                kind="validation",
            ),
        )

    # Self-serve accounts only: authenticate against the `users` table and, on
    # a match, mint a *tenant-bearing* token from the resolved Principal.
    # `authenticate` never raises and returns None on no-match. (The legacy
    # ADMIN_USERNAME/ADMIN_PASSWORD path was removed — it minted a tenant-less
    # token that every data route rejects, so it could never reach any data.)
    from app import identity

    principal = await identity.authenticate(username, password)
    if principal is None:
        log.info("auth/login: rejected attempt for user=%r", username[:40])
        return JSONResponse(
            status_code=401,
            content=envelope(
                "Invalid credentials",
                detail="Username or password is incorrect.",
                kind="auth",
            ),
        )

    log.info("auth/login: success (account) for user=%r", username[:40])
    return {
        "ok": True,
        "authenticated": True,
        "username": principal.email,
        "token": identity.mint_token(principal),
        "token_ttl_hours": settings.auth_token_ttl_hours,
        "auth_enabled": settings.auth_enabled,
    }


@api_router.post("/auth/forgot")
async def auth_forgot(req: ForgotPasswordRequest, request: Request):
    """Begin a password reset. ``{ email } -> { ok: true }`` ALWAYS.

    Neutral by design: whether or not the email maps to an account, the
    response is identical so this endpoint can't be used to enumerate users.
    When a user IS found we mint a signed, 30-min reset token, build the
    frontend reset link, and fire the email (which itself never raises —
    delivery failures are logged, not surfaced).
    """
    rate_error = _rate_limit_check(_client_ip(request))
    if rate_error is not None:
        return JSONResponse(status_code=429, content=rate_error)

    from app import identity
    from app.email import send_password_reset_email
    from app.infrastructure import get_connection

    email = (req.email or "").strip().lower()
    if email:
        try:
            async with get_connection() as db:
                cur = await db.execute(
                    "SELECT id FROM users WHERE email = ?", (email,)
                )
                rows = await cur.fetchall()
                await cur.close()
            if rows:
                user_id = dict(rows[0])["id"]
                token = identity.make_password_reset_token(user_id, email)
                base = _frontend_base_url(request)
                reset_url = f"{base}?reset_token={token}" if base else f"?reset_token={token}"
                await send_password_reset_email(email, reset_url)
                log.info("auth/forgot: reset link issued for user=%r", email[:40])
        except Exception:
            # Never leak failure to the caller — log and still return ok.
            log.warning("auth/forgot: failed to issue reset link", exc_info=True)

    return {"ok": True}


@api_router.post("/auth/reset")
async def auth_reset(req: ResetPasswordRequest, request: Request):
    """Complete a password reset. ``{ token, password } -> { ok: true }``.

    Verifies the signed reset token, then rotates the user's bcrypt hash. A
    bad/expired token or a too-weak password is a 400.
    """
    rate_error = _rate_limit_check(_client_ip(request))
    if rate_error is not None:
        return JSONResponse(status_code=429, content=rate_error)

    from app import identity

    token = (req.token or "").strip()
    password = req.password or ""

    payload = identity.verify_password_reset_token(token)
    if payload is None:
        return JSONResponse(status_code=400, content=envelope(
            "Invalid or expired reset link",
            detail="This password-reset link is invalid or has expired. "
                   "Request a new one.",
            kind="auth",
        ))

    try:
        ok = await identity.set_password(payload["user_id"], password)
    except ValueError as e:
        return JSONResponse(status_code=400, content=envelope(
            "Weak password",
            detail=str(e) or "Password is too short.",
            kind="validation",
        ))

    if not ok:
        return JSONResponse(status_code=400, content=envelope(
            "Invalid or expired reset link",
            detail="The account for this reset link no longer exists.",
            kind="auth",
        ))

    log.info("auth/reset: password rotated for user=%r", payload.get("email", "")[:40])
    return {"ok": True}


@api_router.get("/auth/me")
async def auth_me(request: Request) -> dict:
    """Identity of the caller, plus Google Drive auth status.

    Multi-tenant slice 1: when a valid ``Authorization: Bearer <token>`` is
    present (minted by signup/login), surface the account identity
    ``{user_id, email, tenant_id}`` resolved from it. Absent / invalid
    tokens fall through to the legacy response (Google Drive status for the
    Upload page card), so today's behaviour is unchanged.
    """
    auth_hdr = request.headers.get("Authorization") or ""
    if auth_hdr.lower().startswith("bearer "):
        from app import identity
        principal = identity.verify_token(auth_hdr[7:].strip())
        if principal is not None:
            # Access status (Phase 2) — surfaced so the frontend can show the
            # right blocked screen. `access_enforced` tells it whether the
            # backend is actually gating on it yet (the rollout flag).
            access = await identity.get_access_status(principal.user_id)
            # Drive connection is now PER-TENANT: report status for THIS
            # principal's tenant only.
            try:
                connected = await google_drive.is_connected(principal.tenant_id)
            except Exception:
                log.warning("drive is_connected check failed", exc_info=True)
                connected = False
            return {
                "user_id": principal.user_id,
                "email": principal.email,
                "tenant_id": principal.tenant_id,
                "access_status": access,
                "access_enforced": settings.access_enforcement,
                "authenticated": connected,
                "drive_connected": connected,
                "google_configured": google_drive.is_configured(),
            }

    # No (valid) principal — there is no tenant to scope a token to, so Drive
    # is reported as not connected. (The callback is the only public Drive
    # surface, and it carries its tenant in the signed state.)
    return {
        "authenticated": False,
        "drive_connected": False,
        "google_configured": google_drive.is_configured(),
    }


@api_router.post("/auth/logout")
async def auth_logout(
    principal: Principal = Depends(require_principal),
) -> dict:
    """Disconnect Google Drive — revoke + delete THIS tenant's OAuth token."""
    try:
        await google_drive.revoke_credentials(principal.tenant_id)
    except Exception:
        log.warning("drive revoke failed", exc_info=True)
    return {"ok": True}


@api_router.get("/auth/google/login")
async def google_login(
    principal: Principal = Depends(require_principal),
):
    """Start the Google OAuth flow for the CURRENT tenant.

    Returns JSON ``{auth_url}`` (not a redirect): the frontend kicks this off
    with an authenticated fetch — a plain ``<a href>`` carries no Authorization
    header, so we couldn't know which tenant to bind the token to. The browser
    then navigates to ``auth_url``. The signed ``state`` carries the tenant_id
    so the public callback can recover it.
    """
    if not google_drive.is_configured():
        return _drive_not_configured()
    try:
        url, _state = google_drive.get_authorization_url(principal.tenant_id)
    except Exception as e:
        log.exception("drive: building auth url failed")
        return JSONResponse(status_code=500, content=envelope(
            "Could not start Google sign-in",
            detail=f"{type(e).__name__}: {e}", kind="internal",
        ))
    return {"auth_url": url}


@api_router.get("/auth/google/callback")
async def google_callback(request: Request):
    """OAuth callback (PUBLIC — Google calls it with no bearer token).

    Exchanges the code for credentials, trusting the signed ``state`` to tell
    us which tenant the token belongs to (``exchange_code`` verifies the HMAC
    and extracts the tenant_id), saves it per-tenant, then bounces the browser
    back to the frontend Upload page. On any state/verify/exchange failure we
    redirect with ``?drive=error`` rather than leak a stack trace."""
    if not google_drive.is_configured():
        return _drive_not_configured()
    code = request.query_params.get("code")
    state = request.query_params.get("state")
    error = request.query_params.get("error")
    if error or not code:
        return RedirectResponse(f"{settings.frontend_url}/?drive=error")
    try:
        await google_drive.exchange_code(code, state)
    except Exception:
        # Covers a tampered/invalid signed state (ValueError) and any
        # token-exchange failure — both are a failed connect to the user.
        log.exception("drive: token exchange failed")
        return RedirectResponse(f"{settings.frontend_url}/?drive=error")
    return RedirectResponse(f"{settings.frontend_url}/?drive=connected")


@api_router.get("/drive/status")
async def drive_status(
    principal: Principal = Depends(require_principal),
):
    """List the ingestible Drive files of the CURRENT tenant for the picker."""
    if not google_drive.is_configured():
        return {"connected": False, "configured": False, "files": []}
    try:
        creds = await google_drive.load_credentials(principal.tenant_id)
    except Exception:
        creds = None
    if creds is None:
        return {"connected": False, "configured": True, "files": []}
    try:
        files = await google_drive.list_drive_files(creds)
    except Exception as e:
        log.exception("drive: listing files failed")
        return JSONResponse(status_code=502, content=envelope(
            "Could not list Google Drive files",
            detail=f"{type(e).__name__}: {e}", kind="upstream",
        ))
    return {"connected": True, "configured": True, "files": files}


@api_router.post("/drive/sync")
async def drive_sync(
    request: Request,
    body: DriveSyncRequest,
    principal: Principal = Depends(require_principal),
):
    """Ingest the selected Drive files through the DataCleanAgent pipeline."""
    # DATA-PATH route: ingest_drive_files writes u_* tables (and runs the shared
    # post-ingest refresh) on THIS request task — the asyncio.to_thread calls in
    # it only wrap blocking Google/file I/O, not the DB work — so binding the
    # query tenant here scopes those writes to the tenant's OWN schema
    # (require_principal no longer drives search_path; identity and data
    # isolation are decoupled).
    from app.tenant_context import set_query_tenant
    set_query_tenant(principal.tenant_id)
    rl = _rate_limit_check(_client_ip(request), bucket_namespace="drive", limit_per_minute=5)
    if rl is not None:
        return JSONResponse(status_code=429, content=rl)
    if not google_drive.is_configured():
        return _drive_not_configured()

    target_table = (body.target or "sales").strip().lower()
    if target_table not in ALLOWED_TABLES:
        return JSONResponse(status_code=400, content=envelope(
            "Invalid target",
            detail=f"target must be one of {list(ALLOWED_TABLES)}",
            kind="validation",
        ))
    if body.dedup_mode not in DEDUP_MODES:
        return JSONResponse(status_code=400, content=envelope(
            "Invalid dedup_mode",
            detail=f"dedup_mode must be one of {list(DEDUP_MODES)}",
            kind="validation",
        ))
    if not body.file_ids:
        return JSONResponse(status_code=400, content=envelope(
            "No files selected",
            detail="file_ids must contain at least one Drive file id.",
            kind="validation",
        ))

    creds = await google_drive.load_credentials(principal.tenant_id)
    if creds is None:
        return JSONResponse(status_code=401, content=envelope(
            "Google Drive not connected",
            detail="Connect Google Drive on the Upload page first.",
            kind="auth",
        ))

    try:
        results = await google_drive.ingest_drive_files(
            creds, body.file_ids, target_table, body.dedup_mode,
        )
    except Exception as e:
        log.exception("drive: sync failed")
        return JSONResponse(status_code=500, content=envelope(
            "Drive sync failed",
            detail=f"{type(e).__name__}: {e}", kind="internal",
        ))
    return {
        "target": target_table,
        "dedup_mode": body.dedup_mode,
        "results": results,
        "rows_inserted_total": sum(r.get("rows_inserted", 0) for r in results),
    }



# ---------------------------------------------------------------------------
# Time engine diagnostics
# Returns the current dataset-relative tokens. The AI uses these for every
# relative-time KPI; this route surfaces them for the dashboard / debugging.
# ---------------------------------------------------------------------------

@api_router.get("/time")
async def time_tokens():
    tokens = await resolve_dataset_date_tokens()
    if not tokens:
        return {
            "has_data": False,
            "message": "No uploaded data — relative-time analytics are disabled until first upload.",
        }
    return {
        "has_data": True,
        "dataset_today":     tokens.get("dataset_today"),
        "dataset_yesterday": tokens.get("dataset_yesterday"),
        "dataset_month":     tokens.get("dataset_month"),
        "dataset_year":      tokens.get("dataset_year"),
        "dataset_prev_month": tokens.get("dataset_prev_month"),
        "last_7_window":    [tokens.get("dataset_last_7_start"),    tokens.get("dataset_today")],
        "last_30_window":   [tokens.get("dataset_last_30_start"),   tokens.get("dataset_today")],
        "this_week":        [tokens.get("dataset_week_start"),      tokens.get("dataset_week_end")],
        "this_month":       [tokens.get("dataset_month_start"),     tokens.get("dataset_month_end")],
        "previous_month":   [tokens.get("dataset_prev_month_start"), tokens.get("dataset_prev_month_end")],
        "previous_week":    [tokens.get("dataset_prev_week_start"), tokens.get("dataset_prev_week_end")],
        "tokens": tokens,
    }


# ---------------------------------------------------------------------------
# Error tracking API
# Centralized error log: every uncaught exception, upload failure, AI/KPI
# crash, and explicitly-reported frontend error lands in error_log.
# ---------------------------------------------------------------------------

@api_router.get("/errors")
async def errors_list(
    module: str | None = None,
    severity: str | None = None,
    resolved: bool | None = None,
    limit: int = 200,
    offset: int = 0,
    principal: Principal = Depends(require_principal),
):
    if severity is not None and severity not in SEVERITIES:
        return JSONResponse(status_code=400, content=envelope(
            "Invalid severity",
            detail=f"severity must be one of {list(SEVERITIES)}",
            kind="validation",
        ))
    rows = await list_errors(
        module=module, severity=severity, resolved=resolved,
        limit=limit, offset=offset, tenant_id=principal.tenant_id,
    )
    return {"count": len(rows), "errors": rows}


@api_router.get("/errors/analytics")
async def errors_analytics(principal: Principal = Depends(require_principal)):
    return await error_analytics(tenant_id=principal.tenant_id)


@api_router.get("/errors/{error_id}")
async def errors_get(error_id: str, principal: Principal = Depends(require_principal)):
    row = await get_error(error_id, tenant_id=principal.tenant_id)
    if row is None:
        return JSONResponse(status_code=404, content=envelope(
            "Error not found", detail=f"unknown error_id: {error_id!r}",
            kind="validation",
        ))
    # Parse JSON fields for the consumer.
    import json as _json
    for col in ("request_payload", "context"):
        v = row.get(col)
        if isinstance(v, str) and v:
            try:
                row[col] = _json.loads(v)
            except Exception:
                pass
    return row


@api_router.post("/errors/{error_id}/resolve")
async def errors_resolve(
    error_id: str,
    note: str | None = None,
    principal: Principal = Depends(require_principal),
):
    ok = await resolve_error(error_id, note=note, tenant_id=principal.tenant_id)
    if not ok:
        return JSONResponse(status_code=404, content=envelope(
            "Error not found", detail=f"unknown error_id: {error_id!r}",
            kind="validation",
        ))
    return {"ok": True, "error_id": error_id, "resolved": True}


class FrontendErrorReport(BaseModel):
    message: str = Field(..., min_length=1, max_length=4000)
    module: str = Field(default="frontend", max_length=80)
    severity: str | None = Field(default=None)
    source: str | None = Field(default=None, max_length=400)
    stack: str | None = Field(default=None, max_length=20_000)
    url: str | None = Field(default=None, max_length=2000)
    user_agent: str | None = Field(default=None, max_length=600)
    context: dict | None = Field(default=None)


@api_router.post("/errors/report")
async def errors_report(
    req: FrontendErrorReport,
    principal: Principal = Depends(require_principal),
):
    """Frontend error-boundary reporter. UI crashes / fetch failures /
    chart-render exceptions all POST here. ``require_principal`` binds the
    caller's tenant into the request context so ``log_error`` stamps the
    report with the right tenant (and the /errors dashboard scopes to it)."""
    sev = req.severity if req.severity in SEVERITIES else "medium"
    # Stack arrived as a plain string from the browser; pass it through
    # the `context` field so it shows in the dashboard but doesn't pretend
    # to be a Python traceback.
    ctx = dict(req.context or {})
    if req.stack:
        ctx["js_stack"] = req.stack
    if req.url:
        ctx["url"] = req.url
    if req.user_agent:
        ctx["user_agent"] = req.user_agent
    eid = log_error(
        message=req.message,
        module=req.module or "frontend",
        severity=sev,
        source=req.source or "frontend",
        user_facing=True,
        context=ctx,
    )
    return {"ok": True, "error_id": eid}


async def _health_data_rows() -> int:
    """Return total row count across all dynamic (u_*) tables for the system."""
    from app.db_engine import is_postgres
    if is_postgres():
        tables = await list_dynamic_tables_for_tenant()
        return sum(int(t.get("row_count") or 0) for t in tables)
    return await count_rows("sales") + await count_rows("purchase")


async def _health_db_ping() -> dict:
    """Round-trip ``SELECT 1`` against the live engine.

    ``engine_status()`` only reports what is CONFIGURED; this proves the
    database actually ANSWERS — a paused Supabase project passes every
    config check but fails here (2026-08-03 outage: /health said "ok" for
    hours while the DB was dead). Bounded so a hung database can't stall
    /health indefinitely (Render polls it for liveness): 8s typical worst
    case; tearing down a silently-dead pooled connection can add asyncpg's
    shielded release budget (~15s) on top before the timeout surfaces.
    """
    from app.infrastructure import fetch_one
    try:
        await asyncio.wait_for(fetch_one("SELECT 1"), timeout=8)
        return {"reachable": True}
    except Exception as exc:  # noqa: BLE001 — health must never raise
        return {"reachable": False, "detail": f"{type(exc).__name__}: {str(exc)[:200]}"}


@api_router.get("/health")
async def health() -> dict:
    """Defensive health endpoint.

    Every probe is wrapped in try/except so the load-balancer (Render)
    always gets a 200 with degraded info — never a 500.  This matters
    because /health runs concurrently with the background _startup()
    task, so the DB, KPI registry, etc. may not be ready yet during
    the first few seconds after a deploy.
    """
    def _safe(fn, default=None):
        try:
            return fn()
        except Exception as exc:  # noqa: BLE001 — health must never raise
            return {"error": type(exc).__name__, "detail": str(exc)[:200]} if default is None else default

    async def _safe_async(coro_fn, default=None):
        try:
            return await coro_fn()
        except Exception as exc:  # noqa: BLE001
            return {"error": type(exc).__name__, "detail": str(exc)[:200]} if default is None else default

    llm_block: dict
    try:
        llm_ok, llm_msg = await check_llm_health()
        llm_block = {
            "ok": llm_ok,
            "detail": llm_msg,
            "model": settings.llm_model,
            "base_url": settings.llm_base_url,
        }
    except Exception as exc:  # noqa: BLE001
        llm_block = {
            "ok": False,
            "detail": f"health probe crashed: {type(exc).__name__}: {str(exc)[:200]}",
            "model": settings.llm_model,
            "base_url": settings.llm_base_url,
        }

    db_block = _safe(engine_status, default={"kind": "unknown", "status": "starting"})
    db_probe = await _health_db_ping()
    if isinstance(db_block, dict):
        db_block = {**db_block, **db_probe}

    # Row count only when the ping succeeded: with the DB down it would
    # re-dial the unreachable server under asyncpg's own 60s connect timeout
    # — long enough to trip Render's liveness probe and restart-loop the
    # service. Bounded for the same reason even on the happy path.
    data_rows = 0
    if db_probe.get("reachable"):
        data_rows = await _safe_async(
            lambda: asyncio.wait_for(_health_data_rows(), timeout=10), default=0
        )

    return {
        # "degraded" (still HTTP 200 — Render's liveness check must not
        # restart-loop the service over a paused external DB) signals that
        # the app is up but the database isn't answering.
        "status": "ok" if db_probe.get("reachable") else "degraded",
        "version": app.version,
        "data_version": _safe(get_data_version, default=0),
        "cache": {
            "kind": "json_file",
            "size": _safe(cache_size, default=0),
        },
        "database": db_block,
        "data_rows": data_rows,
        "sentry": _safe(sentry_status, default={"enabled": False}),
        "llm": llm_block,
        "auth_enabled": settings.auth_enabled,
    }


# ---------------------------------------------------------------------------
# Shared post-ingest refresh — runs after ANY ingestion path (file upload or
# Google Drive sync) so derived state stays consistent: data version bump,
# cache invalidation, relationship graph refresh. Returns the new data version.
# ---------------------------------------------------------------------------

async def _post_ingest_refresh() -> int:
    refresh_log = logging.getLogger("agentic_ai.api.ingest")
    new_version = bump_data_version()
    invalidate_all()
    invalidate_time_cache()

    from app.db_engine import is_postgres
    if is_postgres():
        try:
            from app.schema_mapping.relationships import refresh_relationships_pg
            rel_count = await refresh_relationships_pg()
            refresh_log.info("relationships refreshed: count=%d", rel_count)
        except Exception:
            refresh_log.warning("relationship refresh failed (continuing)", exc_info=True)
    return new_version


# ---------------------------------------------------------------------------
# /upload  (DataCleanAgent)
# ---------------------------------------------------------------------------

@api_router.post("/upload")
async def upload(
    request: Request,
    file: UploadFile = File(...),
    target: str = Form("sales"),
    dedup_mode: str = Form(DEFAULT_DEDUP_MODE),
    principal: Principal = Depends(require_principal),
):
    upload_log = logging.getLogger("agentic_ai.api.upload")
    rl = _rate_limit_check(_client_ip(request), bucket_namespace="upload", limit_per_minute=5)
    if rl is not None:
        return JSONResponse(status_code=429, content=rl)

    target_table = (target or "sales").strip().lower()
    filename = file.filename or "upload"
    batch_id = str(uuid4())

    async def _record_error(reason: str, target_for_record: str) -> None:
        try:
            await record_upload_meta(
                batch_id=batch_id,
                filename=filename,
                target=target_for_record if target_for_record in ALLOWED_TABLES else "sales",
                rows_inserted=0,
                rows_failed=0,
                tenant_id=principal.tenant_id,
                source="upload",
                status="error",
                error_message=reason[:500],
            )
        except Exception:
            upload_log.warning("failed to record error meta for batch=%s", batch_id, exc_info=True)

    if target_table not in ALLOWED_TABLES:
        await _record_error(f"invalid target: {target_table!r}", "sales")
        return JSONResponse(status_code=400, content=envelope(
            "Invalid target",
            detail=f"target must be one of {list(ALLOWED_TABLES)}",
            kind="validation",
        ))

    if dedup_mode not in DEDUP_MODES:
        return JSONResponse(status_code=400, content=envelope(
            "Invalid dedup_mode",
            detail=f"dedup_mode must be one of {list(DEDUP_MODES)}",
            kind="validation",
        ))

    lower = filename.lower()
    if lower.endswith(".csv"):
        suffix = ".csv"
    elif lower.endswith(".xlsx"):
        suffix = ".xlsx"
    else:
        await _record_error(f"unsupported file type: {filename!r}", target_table)
        return JSONResponse(status_code=400, content=envelope(
            "Unsupported file type",
            detail="Only .csv and .xlsx are accepted.",
            kind="upload",
        ))

    # Persist the source file under data/uploads/{batch_id}{suffix} — kept
    # forever until the user explicitly disconnects this dataset.
    persistent_path = uploads_dir() / f"{batch_id}{suffix}"
    bytes_written = 0
    crashed = False
    try:
        try:
            with persistent_path.open("wb") as out:
                while True:
                    chunk = await file.read(settings.upload_chunk_bytes)
                    if not chunk:
                        break
                    bytes_written += len(chunk)
                    if bytes_written > settings.max_upload_bytes:
                        await _record_error(
                            f"file too large ({bytes_written} > {settings.max_upload_bytes})",
                            target_table,
                        )
                        crashed = True
                        return JSONResponse(status_code=413, content=envelope(
                            "File too large",
                            detail=f"Max {settings.max_upload_bytes // (1024*1024)} MB per upload.",
                            kind="upload",
                        ))
                    out.write(chunk)
        except Exception as e:
            upload_log.exception("spool: failed")
            await _record_error(f"spool failed: {type(e).__name__}: {e}", target_table)
            crashed = True
            return JSONResponse(status_code=400, content=envelope(
                "Could not read upload",
                detail=f"{type(e).__name__}: {e}",
                kind="upload",
            ))

        if bytes_written == 0:
            await _record_error("empty file", target_table)
            crashed = True
            return JSONResponse(status_code=400, content=envelope(
                "Empty file", kind="upload",
            ))

        # File-level duplicate detection: SHA256 the spooled bytes and look
        # up any existing active upload with the same hash. On collision +
        # dedup_mode='block', refuse with a 409 so the user explicitly
        # chooses how to proceed.
        file_hash, file_bytes_total = compute_file_hash(persistent_path)
        existing = await find_active_upload_by_file_hash(file_hash, tenant_id=principal.tenant_id)
        if existing is not None and dedup_mode == "block":
            crashed = True   # delete the just-spooled redundant copy
            return JSONResponse(status_code=409, content=envelope(
                "Duplicate file already on record",
                detail=(
                    f"This exact file was already uploaded as batch "
                    f"{existing.get('batch_id')!r} on {existing.get('uploaded_at')}. "
                    f"Re-upload with dedup_mode=skip / replace / append to override."
                ),
                kind="duplicate",
                extra={
                    "existing_batch_id":  existing.get("batch_id"),
                    "existing_filename":  existing.get("filename"),
                    "existing_target":    existing.get("target"),
                    "existing_uploaded_at": existing.get("uploaded_at"),
                    "existing_rows_inserted": existing.get("rows_inserted"),
                    "file_hash":          file_hash,
                    "options": ["skip", "replace", "append"],
                },
            ))

        # On Postgres the legacy sales/purchase insert path is a no-op, so a
        # CSV here would silently ingest ZERO rows and then fail. Route CSV
        # through the dynamic u_* pipeline (the same machinery /upload_workbook
        # uses) so it actually creates a table and inserts rows. SQLite keeps
        # the legacy DataCleanAgent path unchanged (that's what the suite
        # exercises and it works on SQLite).
        from app.db_engine import is_postgres as _is_pg_upload
        if _is_pg_upload() and suffix == ".csv":
            from app.dynamic_ingest import ingest_csv, DynamicIngestError
            try:
                summary = await ingest_csv(
                    file_bytes=persistent_path.read_bytes(),
                    source_file_name=filename,
                    batch_id=batch_id,
                )
            except DynamicIngestError as e:
                await _record_error(f"bad upload: {e}", target_table)
                log_error(
                    exc=e, module="upload", severity="high",
                    endpoint="/upload", method="POST", user_facing=True,
                    source="ingest_csv",
                    context={"batch_id": batch_id, "filename": filename},
                )
                crashed = True
                return JSONResponse(status_code=400, content=envelope(
                    "Bad upload", detail=str(e), kind="upload",
                ))
            except Exception as e:
                upload_log.exception("ingest_csv: crashed")
                await _record_error(f"{type(e).__name__}: {e}", target_table)
                log_error(
                    exc=e, module="upload", severity="critical",
                    endpoint="/upload", method="POST", user_facing=False,
                    source="ingest_csv",
                    context={"batch_id": batch_id, "filename": filename},
                )
                crashed = True
                return JSONResponse(status_code=500, content=envelope(
                    "Ingest failed",
                    detail=f"{type(e).__name__}: {e}",
                    kind="internal",
                ))

            ingested = summary.get("ingested") or []
            if not ingested:
                reason = "; ".join(
                    s.get("reason", "no header detected")
                    for s in (summary.get("skipped") or [])
                ) or "Could not detect a header row in the CSV."
                await _record_error(f"bad upload: {reason}", target_table)
                crashed = True
                return JSONResponse(status_code=400, content=envelope(
                    "Bad upload", detail=reason, kind="upload",
                ))

            rows_inserted = sum(int(s.get("rows_inserted") or 0) for s in ingested)
            rows_failed = sum(int(s.get("skipped_rows") or 0) for s in ingested)
            table_name = ingested[0].get("table") or target_table

            # ingest_csv (via the dynamic path) already wrote the uploads row;
            # patch the audit fields so disconnect + dedup can find it later.
            try:
                from app.infrastructure import get_connection
                async with get_connection() as db:
                    await db.execute(
                        "UPDATE uploads SET file_path = ?, file_hash = ?, "
                        "file_bytes = ? WHERE batch_id = ?",
                        (str(persistent_path), file_hash, file_bytes_total, batch_id),
                    )
                    await db.commit()
            except Exception:
                upload_log.warning("could not stamp file_path on uploads row", exc_info=True)

            new_version = await _post_ingest_refresh()
            upload_log.info(
                "upload(csv->dynamic) ok rows=%d table=%s batch=%s data_version=%d",
                rows_inserted, table_name, batch_id, new_version,
            )
            return {
                "batch_id":          batch_id,
                "filename":          filename,
                "target":            table_name,
                "rows_inserted":     rows_inserted,
                "rows_failed":       rows_failed,
                "rows_skipped_duplicate": 0,
                "rows_replaced":     0,
                "dedup_mode":        dedup_mode,
                "dedup":             None,
                "errors":            [],
                "summary":           {"total_sales": 0, "min_date": "", "max_date": ""},
                "unmatched_headers": [],
                "sheet_name":        ingested[0].get("source_sheet"),
                "header_row_used":   [c.get("name") for c in (ingested[0].get("columns") or [])],
                "validation":        None,
                "bytes_received":    bytes_written,
                "table_total":       rows_inserted,
                "file_path":         str(persistent_path),
                "file_hash":         file_hash,
                "data_version":      new_version,
            }

        agent = DataCleanAgent()
        try:
            result = await agent.run(
                tmp_path=persistent_path,
                filename=filename,
                target=target_table,
                batch_id=batch_id,
                dedup_mode=dedup_mode,
            )
        except UploadError as e:
            await _record_error(f"bad upload: {e}", target_table)
            log_error(
                exc=e, module="upload", severity="high",
                endpoint="/upload", method="POST", user_facing=True,
                source="DataCleanAgent",
                context={"target": target_table, "batch_id": batch_id, "filename": filename},
            )
            crashed = True
            return JSONResponse(status_code=400, content=envelope(
                "Bad upload", detail=str(e), kind="upload",
            ))
        except Exception as e:
            upload_log.exception("dataclean: crashed")
            await _record_error(f"{type(e).__name__}: {e}", target_table)
            log_error(
                exc=e, module="upload", severity="critical",
                endpoint="/upload", method="POST", user_facing=False,
                source="DataCleanAgent",
                context={"target": target_table, "batch_id": batch_id, "filename": filename},
            )
            crashed = True
            return JSONResponse(status_code=500, content=envelope(
                "Ingest failed",
                detail=f"{type(e).__name__}: {e}",
                kind="internal",
            ))

        # Stamp persisted path + file hash onto the uploads metadata row.
        # DataCleanAgent already inserted the success record; we patch
        # these audit fields on top so disconnect_upload + the duplicate
        # detector can find them later.
        try:
            from app.infrastructure import get_connection
            async with get_connection() as db:
                await db.execute(
                    "UPDATE uploads SET file_path = ?, file_hash = ?, "
                    "file_bytes = ? WHERE batch_id = ?",
                    (str(persistent_path), file_hash, file_bytes_total, batch_id),
                )
                await db.commit()
        except Exception:
            upload_log.warning("could not stamp file_path on uploads row", exc_info=True)

        new_version = await _post_ingest_refresh()
        upload_log.info(
            "upload ok rows=%d batch=%s file=%s data_version=%d",
            result["rows_inserted"], batch_id, persistent_path.name, new_version,
        )

        return {
            "batch_id":          result["batch_id"],
            "filename":          result["filename"],
            "target":            result["target"],
            "rows_inserted":     result["rows_inserted"],
            "rows_failed":       result["rows_failed"],
            "rows_skipped_duplicate": result.get("rows_skipped_duplicate", 0),
            "rows_replaced":     result.get("rows_replaced", 0),
            "dedup_mode":        result.get("dedup_mode", dedup_mode),
            "dedup":             result.get("dedup"),
            "errors":            result["errors"],
            "summary":           result["summary"],
            "unmatched_headers": result["unmatched_headers"],
            "sheet_name":        result.get("sheet_name"),
            "header_row_used":   result.get("header_row_used"),
            "validation":        result.get("validation"),
            "bytes_received":    bytes_written,
            "table_total":       await count_rows(target_table),
            "file_path":         str(persistent_path),
            "file_hash":         file_hash,
            # Stamp the post-ingest data_version so the frontend can
            # mark its cached Dashboard / Uploads list as stale and
            # auto-refetch without polling.
            "data_version":      new_version,
        }
    except Exception as e:
        upload_log.exception("upload: unhandled crash")
        crashed = True
        return JSONResponse(status_code=500, content=envelope(
            "Upload failed",
            detail=f"{type(e).__name__}: {e}",
            kind="internal",
        ))
    finally:
        # On any error path, clean up the partial file from data/uploads/ so
        # the directory doesn't accumulate orphans. Successful uploads keep
        # the file — that's the whole point of persistence.
        if crashed:
            try:
                if persistent_path.exists():
                    persistent_path.unlink()
            except Exception:
                log.warning("cleanup of failed upload file failed: %s", persistent_path, exc_info=True)


# ---------------------------------------------------------------------------
# /upload_workbook — multi-sheet dynamic ingestion (ADR-0005)
# ---------------------------------------------------------------------------
# Drops the whole .xlsx, each sheet becomes its own `u_<sheet>` table.
# Bypasses the strict sales/purchase SCHEMA_SPEC — column types are inferred
# from the data, no alias mapping required. Re-uploading replaces the per-
# sheet tables. Used by the "drop a workbook" flow on the Upload page when
# the user has multi-domain data (sales + inventory + product catalog +
# hierarchy) that doesn't fit the legacy two-table model.

# In-memory registry of async workbook-ingest jobs, keyed by job_id (== batch_id).
# Each entry is the shape returned by GET /upload_status/{job_id}:
#   {
#     "status":      "running" | "done" | "error",
#     "done_sheets": int,
#     "total_sheets": int,
#     "message":     str,
#     "summary":     dict | None,   # the ingest summary (+ totals) once done
#     "error":       str | None,    # error text when status == "error"
#     "tenant_id":   str,           # owner; used to scope status reads
#     "filename":    str,
#   }
# Lives only for the process lifetime (single-process uvicorn). A restart loses
# in-flight jobs, which is acceptable for an upload-progress affordance.
_UPLOAD_JOBS: dict[str, dict] = {}


async def _run_workbook_ingest(
    *,
    job_id: str,
    tenant_id: str,
    persistent_path: Path,
    filename: str,
    bytes_written: int,
) -> None:
    """Background coroutine that runs the actual multi-sheet ingestion.

    Spawned with ``asyncio.create_task`` from the /upload_workbook handler so
    the HTTP response can return immediately (202). Updates ``_UPLOAD_JOBS``
    as it progresses and on completion / failure.

    TENANT BINDING: a freshly-created asyncio task does NOT reliably inherit
    the request's contextvars, so we MUST re-bind the query tenant explicitly
    at the very top — before any DB-touching call (ingest_workbook,
    record_upload_meta, _post_ingest_refresh). Without this the upload could
    land in the wrong (default) schema and break tenant isolation.
    """
    upload_log = logging.getLogger("agentic_ai.api.upload_workbook")
    from app.tenant_context import set_query_tenant
    set_query_tenant(tenant_id)

    # Defensive: signup-time provisioning is best-effort (identity.py), so a
    # tenant whose schema was never created would crash on the first CREATE
    # TABLE under the strict per-tenant search_path. Idempotent + non-fatal.
    try:
        from app.db_engine import ensure_tenant_schema
        await ensure_tenant_schema(tenant_id)
    except Exception:
        upload_log.exception(
            "ensure_tenant_schema failed for tenant=%r (upload continues)",
            tenant_id,
        )

    job = _UPLOAD_JOBS.get(job_id)

    def _progress(done: int, total: int, msg: str) -> None:
        j = _UPLOAD_JOBS.get(job_id)
        if j is None:
            return
        j["done_sheets"] = done
        j["total_sheets"] = total
        j["message"] = msg

    crashed = False
    try:
        try:
            summary = await ingest_workbook(
                wb_path=persistent_path,
                source_file_name=filename,
                batch_id=job_id,
                progress_cb=_progress,
            )
        except Exception as e:
            upload_log.exception("dynamic_ingest: crashed (job=%s)", job_id)
            crashed = True
            if job is not None:
                job["status"] = "error"
                job["error"] = f"{type(e).__name__}: {e}"
                job["message"] = "Ingest failed"
            return

        # Audit row + bump data_version so caches invalidate. Both run under
        # the tenant binding set above: record_upload_meta writes the shared
        # public.uploads table, while _post_ingest_refresh introspects THIS
        # tenant's u_* schema (relationship detection), so the binding matters.
        total_rows = sum(s["rows_inserted"] for s in summary["ingested"])
        try:
            await record_upload_meta(
                batch_id=job_id,
                filename=filename,
                target="(workbook)",
                rows_inserted=total_rows,
                rows_failed=0,
                tenant_id=tenant_id,
                source="upload",
                status="active",
            )
        except Exception:
            upload_log.warning("could not record uploads meta", exc_info=True)
        new_version = await _post_ingest_refresh()

        upload_log.info(
            "upload_workbook ok sheets=%d rows=%d batch=%s file=%s data_version=%d",
            len(summary["ingested"]), total_rows, job_id, persistent_path.name,
            new_version,
        )

        if job is not None:
            job["status"] = "done"
            job["message"] = "Upload complete"
            job["done_sheets"] = summary["sheet_count"]
            job["total_sheets"] = summary["sheet_count"]
            job["summary"] = {
                "batch_id":       job_id,
                "filename":       filename,
                "sheet_count":    summary["sheet_count"],
                "tables":         summary["tables"],
                "ingested":       summary["ingested"],
                "skipped":        summary["skipped"],
                "total_rows":     total_rows,
                "rows_inserted":  total_rows,   # matches UploadResponse contract
                "data_version":   new_version,
                "bytes_received": bytes_written,
                "file_path":      str(persistent_path),
            }
    except Exception as e:
        upload_log.exception("upload_workbook: unhandled crash (job=%s)", job_id)
        crashed = True
        if job is not None:
            job["status"] = "error"
            job["error"] = f"{type(e).__name__}: {e}"
            job["message"] = "Upload failed"
    finally:
        # Preserve the original partial-file cleanup behavior: a failed ingest
        # removes the spooled workbook so it isn't left dangling on disk.
        if crashed:
            try:
                if persistent_path.exists():
                    persistent_path.unlink()
            except Exception:
                log.warning("cleanup of failed upload file failed: %s", persistent_path, exc_info=True)


@api_router.post("/upload_workbook", status_code=202)
async def upload_workbook(
    request: Request,
    file: UploadFile = File(...),
    principal: Principal = Depends(require_principal),
):
    upload_log = logging.getLogger("agentic_ai.api.upload_workbook")
    rl = _rate_limit_check(_client_ip(request), bucket_namespace="upload", limit_per_minute=5)
    if rl is not None:
        return JSONResponse(status_code=429, content=rl)

    filename = file.filename or "upload"
    batch_id = str(uuid4())

    lower = filename.lower()
    if not lower.endswith(".xlsx"):
        return JSONResponse(status_code=400, content=envelope(
            "Unsupported file type",
            detail="/upload_workbook only accepts .xlsx files (multi-sheet). "
                   "Use /upload for single-sheet xlsx or csv.",
            kind="upload",
        ))

    # Capture the tenant on the request thread — the background task can't rely
    # on inheriting the request's contextvars, so we pass this explicitly and
    # re-bind it inside _run_workbook_ingest before any DB work.
    tenant_id = principal.tenant_id

    # Spool to disk under data/uploads/{batch_id}.xlsx — kept until disconnect.
    persistent_path = uploads_dir() / f"{batch_id}.xlsx"
    bytes_written = 0
    spool_failed = False
    try:
        try:
            with persistent_path.open("wb") as out:
                while True:
                    chunk = await file.read(settings.upload_chunk_bytes)
                    if not chunk:
                        break
                    bytes_written += len(chunk)
                    if bytes_written > settings.max_upload_bytes:
                        spool_failed = True
                        return JSONResponse(status_code=413, content=envelope(
                            "File too large",
                            detail=f"Max {settings.max_upload_bytes // (1024*1024)} MB per upload.",
                            kind="upload",
                        ))
                    out.write(chunk)
        except Exception as e:
            upload_log.exception("spool: failed")
            spool_failed = True
            return JSONResponse(status_code=400, content=envelope(
                "Could not read upload",
                detail=f"{type(e).__name__}: {e}",
                kind="upload",
            ))

        if bytes_written == 0:
            spool_failed = True
            return JSONResponse(status_code=400, content=envelope(
                "Empty file", kind="upload",
            ))
    finally:
        # Only clean up here for spooling failures handled inline. Once the
        # background task owns the file, IT is responsible for cleanup on error.
        if spool_failed:
            try:
                if persistent_path.exists():
                    persistent_path.unlink()
            except Exception:
                log.warning("cleanup of failed upload file failed: %s", persistent_path, exc_info=True)

    # Read total sheet count up front (cheap) so the very first status poll has
    # a meaningful total_sheets. Best-effort — ingestion still reports it too.
    total_sheets = 0
    try:
        from openpyxl import load_workbook as _load_wb
        _wb = _load_wb(filename=str(persistent_path), read_only=True, data_only=True)
        try:
            total_sheets = len(_wb.sheetnames)
        finally:
            try:
                _wb.close()
            except Exception:
                pass
    except Exception:
        upload_log.warning("could not pre-read sheet count for %s", persistent_path, exc_info=True)

    # Register the job, then kick off ingestion in the background and return 202
    # immediately so the client never blocks on the ~minute-long ingest.
    _UPLOAD_JOBS[batch_id] = {
        "status":       "running",
        "done_sheets":  0,
        "total_sheets": total_sheets,
        "message":      "Queued",
        "summary":      None,
        "error":        None,
        "tenant_id":    tenant_id,
        "filename":     filename,
    }

    asyncio.create_task(_run_workbook_ingest(
        job_id=batch_id,
        tenant_id=tenant_id,
        persistent_path=persistent_path,
        filename=filename,
        bytes_written=bytes_written,
    ))

    return {"job_id": batch_id, "status": "running"}


@api_router.get("/upload_status/{job_id}")
async def upload_status(
    job_id: str,
    principal: Principal = Depends(require_principal),
):
    """Poll the progress of an async /upload_workbook ingestion job.

    Scoped to the calling tenant: a job started by one tenant is invisible to
    another (returns 404), so the in-memory registry can't leak cross-tenant
    upload state.
    """
    job = _UPLOAD_JOBS.get(job_id)
    if job is None or job.get("tenant_id") != principal.tenant_id:
        return JSONResponse(status_code=404, content=envelope(
            "Unknown upload job",
            detail=f"No upload job with id {job_id!r}.",
            kind="upload",
        ))
    return {
        "status":       job["status"],
        "done_sheets":  job["done_sheets"],
        "total_sheets": job["total_sheets"],
        "message":      job["message"],
        "summary":      job["summary"],
        "error":        job["error"],
    }


# ---------------------------------------------------------------------------
# /tables/dynamic — list all u_* tables currently ingested (debug/UI helper)
# ---------------------------------------------------------------------------

@api_router.get("/tables/dynamic")
async def list_dynamic_tables_route(
    principal: Principal = Depends(require_principal),
):
    # DATA-PATH route: lists u_* tables. Bind the query tenant so the listing
    # resolves against the tenant's OWN schema rather than the default.
    # NOTE: use the tenant-scoped introspecting variant (NOT the global
    # registry) so this never leaks other tenants' table metadata.
    from app.tenant_context import set_query_tenant
    set_query_tenant(principal.tenant_id)
    return {"tables": await list_dynamic_tables_for_tenant()}


@api_router.post("/tables/dynamic/disconnect_all")
async def disconnect_all_dynamic_tables(principal: Principal = Depends(require_principal)):
    dropped = await drop_all_dynamic_tables()
    new_version = await _post_ingest_refresh()
    return {"dropped": dropped, "data_version": new_version}


# ---------------------------------------------------------------------------
# /schema/* — tenant schema-concept mapping (propose / confirm / current).
# Lets a tenant override the deterministic resolver with an LLM-proposed,
# human-confirmed mapping of canonical concepts -> their real columns. All
# three are DATA-PATH routes (they introspect / read the tenant's own u_*
# schema), so each binds the query tenant first like /dashboard does.
# ---------------------------------------------------------------------------

class SchemaConfirmRequest(BaseModel):
    """Body for POST /schema/confirm — the mapping the user is confirming.

    Extra fields ignored so a slightly-divergent frontend build still slots
    in. ``mapping`` mirrors the store/llm_mapper dict format:
    ``{"concepts": {...}, "source": ..., "updated_at": ...}``.
    """

    model_config = ConfigDict(extra="ignore")

    mapping: dict[str, Any] = Field(default_factory=dict)


@api_router.post("/schema/propose")
async def schema_propose(
    principal: Principal = Depends(require_principal),
):
    # DATA-PATH route: introspects the tenant's u_* schema. Bind the query
    # tenant so the proposal + validation resolve against the tenant's OWN
    # schema rather than the default search_path.
    from app.tenant_context import set_query_tenant
    set_query_tenant(principal.tenant_id)
    from app.schema_mapping.llm_mapper import propose_mapping, validate_mapping

    mapping = await propose_mapping(principal.tenant_id)
    tables = await list_dynamic_tables_for_tenant()
    validation = validate_mapping(mapping, tables)
    return {"mapping": mapping, "validation": validation}


@api_router.post("/schema/confirm")
async def schema_confirm(
    req: SchemaConfirmRequest,
    principal: Principal = Depends(require_principal),
):
    # DATA-PATH route: validates against the tenant's u_* schema before
    # persisting. Bind the query tenant first.
    from app.tenant_context import set_query_tenant
    set_query_tenant(principal.tenant_id)
    from app.schema_mapping.llm_mapper import validate_mapping
    from app.schema_mapping.store import save_tenant_mapping

    mapping = dict(req.mapping or {})
    tables = await list_dynamic_tables_for_tenant()
    validation = validate_mapping(mapping, tables)
    if not validation.get("ok"):
        return JSONResponse(status_code=400, content=envelope(
            "Invalid schema mapping",
            detail="The mapping failed validation; see issues.",
            kind="validation",
            extra={"issues": validation.get("issues", [])},
        ))
    # Force source="user" — a confirmed mapping is human-authored regardless
    # of what the body claims.
    mapping["source"] = "user"
    await save_tenant_mapping(principal.tenant_id, mapping)
    return {"ok": True}


@api_router.get("/schema/current")
async def schema_current(
    principal: Principal = Depends(require_principal),
):
    # DATA-PATH route: reads the tenant's stored mapping. Bind the query
    # tenant first for consistency with the other /schema/* routes.
    from app.tenant_context import set_query_tenant
    set_query_tenant(principal.tenant_id)
    from app.schema_mapping.store import load_tenant_mapping

    return {"mapping": await load_tenant_mapping(principal.tenant_id)}


# ---------------------------------------------------------------------------
# /upload/preview — dry-run a file: parse, validate, classify, but DO NOT insert.
# Returns the duplicate classification so the user can pick a dedup_mode
# before committing.
# ---------------------------------------------------------------------------

@api_router.post("/upload/preview")
async def upload_preview(
    request: Request,
    file: UploadFile = File(...),
    target: str = Form("sales"),
    principal: "Principal" = Depends(require_principal),
):
    upload_log = logging.getLogger("agentic_ai.api.upload.preview")
    target_table = (target or "sales").strip().lower()
    if target_table not in ALLOWED_TABLES:
        return JSONResponse(status_code=400, content=envelope(
            "Invalid target",
            detail=f"target must be one of {list(ALLOWED_TABLES)}",
            kind="validation",
        ))

    filename = file.filename or "preview"
    lower = filename.lower()
    if lower.endswith(".csv"):
        suffix = ".csv"
    elif lower.endswith(".xlsx"):
        suffix = ".xlsx"
    else:
        return JSONResponse(status_code=400, content=envelope(
            "Unsupported file type",
            detail="Only .csv and .xlsx are accepted.",
            kind="upload",
        ))

    # Spool to a TEMP path (preview doesn't keep the file).
    import tempfile
    fd, tmp_path = tempfile.mkstemp(prefix="agentic_preview_", suffix=suffix)
    bytes_written = 0
    try:
        try:
            with os.fdopen(fd, "wb") as out:
                while True:
                    chunk = await file.read(settings.upload_chunk_bytes)
                    if not chunk:
                        break
                    bytes_written += len(chunk)
                    if bytes_written > settings.max_upload_bytes:
                        return JSONResponse(status_code=413, content=envelope(
                            "File too large",
                            detail=f"Max {settings.max_upload_bytes // (1024*1024)} MB.",
                            kind="upload",
                        ))
                    out.write(chunk)
        except Exception as e:
            return JSONResponse(status_code=400, content=envelope(
                "Could not read upload",
                detail=f"{type(e).__name__}: {e}",
                kind="upload",
            ))

        if bytes_written == 0:
            return JSONResponse(status_code=400, content=envelope(
                "Empty file", kind="upload",
            ))

        # File-hash check against existing active uploads (scoped to this tenant)
        file_hash, _ = compute_file_hash(tmp_path)
        existing = await find_active_upload_by_file_hash(file_hash, tenant_id=principal.tenant_id)
        existing_summary = None
        if existing is not None:
            existing_summary = {
                "batch_id":      existing.get("batch_id"),
                "filename":      existing.get("filename"),
                "target":        existing.get("target"),
                "uploaded_at":   existing.get("uploaded_at"),
                "rows_inserted": existing.get("rows_inserted"),
            }

        # Run the agent in preview-only mode for row-level classification
        agent = DataCleanAgent()
        try:
            result = await agent.run(
                tmp_path=Path(tmp_path),
                filename=filename,
                target=target_table,
                preview_only=True,
                dedup_mode="skip",   # any non-block mode skips the early raise
            )
        except UploadError as e:
            return JSONResponse(status_code=400, content=envelope(
                "Bad upload", detail=str(e), kind="upload",
            ))
        except Exception as e:
            upload_log.exception("preview crashed")
            return JSONResponse(status_code=500, content=envelope(
                "Preview failed",
                detail=f"{type(e).__name__}: {e}",
                kind="internal",
            ))

        return {
            "preview":               True,
            "filename":              filename,
            "target":                target_table,
            "file_hash":             file_hash,
            "bytes_received":        bytes_written,
            "existing_file_upload":  existing_summary,
            "rows_failed":           result.get("rows_failed", 0),
            "errors":                result.get("errors", []),
            "header_row_used":       result.get("header_row_used"),
            "unmatched_headers":     result.get("unmatched_headers"),
            "sheet_name":            result.get("sheet_name"),
            "dedup":                 result.get("dedup"),
            "recommended_mode": (
                "skip" if existing_summary or (
                    result.get("dedup", {}).get("rows_duplicate", 0) > 0
                ) else "block"
            ),
        }
    finally:
        try:
            os.unlink(tmp_path)
        except FileNotFoundError:
            pass
        except Exception:
            log.warning("preview temp cleanup failed: %s", tmp_path, exc_info=True)


# ---------------------------------------------------------------------------
# /dashboard
# ---------------------------------------------------------------------------

@api_router.get("/dashboard")
async def dashboard(
    month: str | None = None,
    principal: Principal = Depends(require_principal),
):
    # DATA-PATH route: reads u_* tables. Explicitly bind the query tenant so
    # those reads resolve against the tenant's OWN schema (require_principal no
    # longer drives search_path — identity and data isolation are decoupled).
    from app.tenant_context import set_query_tenant
    set_query_tenant(principal.tenant_id)
    if month is not None and (len(month) != 7 or month[4] != "-"):
        return JSONResponse(status_code=400, content=envelope(
            "Invalid month",
            detail="Expected format YYYY-MM",
            kind="validation",
        ))
    try:
        result = await DashboardAgent().run(month=month)
    except ValueError as e:
        return JSONResponse(status_code=400, content=envelope(
            "Invalid dashboard query", detail=str(e), kind="validation",
        ))
    except Exception as e:
        log.exception("dashboard failed")
        return JSONResponse(status_code=500, content=envelope(
            "Dashboard failed",
            detail=f"{type(e).__name__}: {e}",
            kind="internal",
        ))
    # Stamp data_version so the frontend can detect server-side changes
    # since its last fetch and short-circuit redundant reloads.
    if isinstance(result, dict) and "data_version" not in result:
        result = {**result, "data_version": get_data_version()}
    return result


# ---------------------------------------------------------------------------
# /uploads — metadata listing + dataset removal
# ---------------------------------------------------------------------------

@api_router.get("/uploads")
async def uploads(
    principal: Principal = Depends(require_principal),
):
    # First-party isolation: the uploads registry lives in ``public`` and is
    # scoped by ``tenant_id`` so each tenant only sees their own datasets.
    return {
        "uploads": await list_uploads_meta(tenant_id=principal.tenant_id),
        "total_rows": {
            "sales":    await count_rows("sales"),
            "purchase": await count_rows("purchase"),
        },
    }


@api_router.post("/uploads/{batch_id}/archive")
async def upload_archive(batch_id: str, principal: Principal = Depends(require_principal)):
    """Soft-deactivate a dataset. AI immediately stops using it. Source
    file is preserved on disk; rows are moved to the archive table. Use
    `/uploads/{batch_id}/unarchive` to reactivate."""
    if not batch_id or len(batch_id) > 64:
        return JSONResponse(status_code=400, content=envelope(
            "Invalid batch_id", detail="batch_id is empty or too long.",
            kind="validation",
        ))
    try:
        result = await archive_upload(batch_id, tenant_id=principal.tenant_id)
    except ValueError as e:
        return JSONResponse(status_code=400, content=envelope(
            "Archive failed", detail=str(e), kind="validation",
        ))
    except Exception as e:
        log.exception("archive_upload failed")
        return JSONResponse(status_code=500, content=envelope(
            "Archive failed",
            detail=f"{type(e).__name__}: {e}",
            kind="internal",
        ))
    bump_data_version()
    invalidate_all()
    invalidate_time_cache()
    return result


@api_router.post("/uploads/{batch_id}/unarchive")
async def upload_unarchive(batch_id: str, principal: Principal = Depends(require_principal)):
    """Restore an archived dataset back to active. Rows move from the
    archive table to the live table; AI starts using them again."""
    if not batch_id or len(batch_id) > 64:
        return JSONResponse(status_code=400, content=envelope(
            "Invalid batch_id", detail="batch_id is empty or too long.",
            kind="validation",
        ))
    try:
        result = await unarchive_upload(batch_id, tenant_id=principal.tenant_id)
    except ValueError as e:
        return JSONResponse(status_code=400, content=envelope(
            "Unarchive failed", detail=str(e), kind="validation",
        ))
    except Exception as e:
        log.exception("unarchive_upload failed")
        return JSONResponse(status_code=500, content=envelope(
            "Unarchive failed",
            detail=f"{type(e).__name__}: {e}",
            kind="internal",
        ))
    bump_data_version()
    invalidate_all()
    invalidate_time_cache()
    return result


@api_router.get("/datasets")
async def datasets_view(
    principal: Principal = Depends(require_principal),
):
    """Single-call dataset management view. Returns the full upload
    inventory bucketed by status, plus live + archive row counts. The
    frontend dataset manager renders this directly."""
    # First-party isolation: scope the uploads registry to the caller's tenant.
    all_uploads = await list_uploads_meta(limit=1000, tenant_id=principal.tenant_id)
    by_status: dict[str, list] = {"active": [], "archived": [], "error": [], "removed": []}
    for u in all_uploads:
        by_status.setdefault(u.get("status") or "active", []).append(u)

    return {
        "data_version": get_data_version(),
        "totals": {
            "sales_live":         await count_rows("sales"),
            "purchase_live":      await count_rows("purchase"),
            "sales_archived":     await count_rows("sales_archive"),
            "purchase_archived":  await count_rows("purchase_archive"),
        },
        "active":   by_status["active"],
        "archived": by_status["archived"],
        "errored":  by_status["error"],
        "removed":  by_status["removed"],
        "counts": {
            "active":   len(by_status["active"]),
            "archived": len(by_status["archived"]),
            "errored":  len(by_status["error"]),
            "removed":  len(by_status["removed"]),
            "total":    len(all_uploads),
        },
    }


@api_router.post("/uploads/{batch_id}/disconnect")
async def upload_disconnect(batch_id: str, principal: Principal = Depends(require_principal)):
    if not batch_id or len(batch_id) > 64:
        return JSONResponse(status_code=400, content=envelope(
            "Invalid batch_id", detail="batch_id is empty or too long.",
            kind="validation",
        ))
    try:
        result = await disconnect_upload(batch_id, tenant_id=principal.tenant_id)
    except ValueError as e:
        return JSONResponse(status_code=404, content=envelope(
            "Dataset not found", detail=str(e), kind="validation",
        ))
    except Exception as e:
        log.exception("disconnect_upload failed")
        return JSONResponse(status_code=500, content=envelope(
            "Disconnect failed",
            detail=f"{type(e).__name__}: {e}",
            kind="internal",
        ))
    bump_data_version()
    invalidate_all()
    return result


# ---------------------------------------------------------------------------
# /cache/clear
# ---------------------------------------------------------------------------

@api_router.post("/cache/clear")
async def cache_clear(principal: Principal = Depends(require_principal)):
    n = invalidate_all()
    new_version = bump_data_version()
    return {"cleared": n, "data_version": new_version}


# ---------------------------------------------------------------------------
# /conversations — read-only chat history (list / open / delete)
# ---------------------------------------------------------------------------

@api_router.get("/conversations")
async def conversations_list(
    principal: Principal = Depends(require_principal),
):
    """List recent chat conversations (most-recently-updated first).

    First-party isolation: conversations live in ``public`` and are scoped by
    ``tenant_id`` (require_principal no longer drives search_path), so each
    tenant only sees their own history.
    """
    return {"conversations": await list_conversations(tenant_id=principal.tenant_id)}


@api_router.get("/conversations/{conversation_id}")
async def conversation_detail(
    conversation_id: str,
    principal: Principal = Depends(require_principal),
):
    """Return one conversation's metadata + its full message transcript."""
    if not conversation_id or len(conversation_id) > 64:
        return JSONResponse(status_code=400, content=envelope(
            "Invalid conversation_id",
            detail="conversation_id is empty or too long.",
            kind="validation",
        ))
    convo = await get_conversation(conversation_id, tenant_id=principal.tenant_id)
    if convo is None:
        return JSONResponse(status_code=404, content=envelope(
            "Conversation not found",
            detail=f"No conversation with id {conversation_id}.",
            kind="validation",
        ))
    return convo


@api_router.delete("/conversations/{conversation_id}")
async def conversation_delete(
    conversation_id: str,
    principal: Principal = Depends(require_principal),
):
    """Delete a conversation and all of its messages."""
    if not conversation_id or len(conversation_id) > 64:
        return JSONResponse(status_code=400, content=envelope(
            "Invalid conversation_id",
            detail="conversation_id is empty or too long.",
            kind="validation",
        ))
    deleted = await delete_conversation(conversation_id, tenant_id=principal.tenant_id)
    if not deleted:
        return JSONResponse(status_code=404, content=envelope(
            "Conversation not found",
            detail=f"No conversation with id {conversation_id}.",
            kind="validation",
        ))
    return {"deleted": True}


# ---------------------------------------------------------------------------
# /query_stream (SSE)
# ---------------------------------------------------------------------------

class QueryRequest(BaseModel):
    question: str = Field(..., min_length=1, max_length=4000)
    conversation_id: str | None = Field(default=None, max_length=128)
    # Language the assistant writes its answer in: "en" (default) | "hi".
    # "hi" = ROMANISED Hindi (Hindi in Latin letters, "Hinglish"), not Devanagari.
    # Anything unrecognised is coerced to "en" at the request boundary.
    language: str = Field(default="en", max_length=8)


# Languages the assistant can answer in. Anything outside this set falls back
# to English so a bad/stale client value can never wedge a turn.
ANSWER_LANGUAGES = {"en", "hi"}


HEARTBEAT_SECONDS = 15.0


def _validate_pre_stream(req: QueryRequest) -> dict[str, Any] | None:
    if not req.question.strip():
        return envelope("Empty question",
                        detail="Question must contain non-whitespace.", kind="validation")
    return None


def _safe_create_task(coro):
    try:
        return asyncio.create_task(coro)
    except Exception:
        try:
            coro.close()
        except Exception:
            pass
        return None


# PresentationEmitter (event scrubbing + hidden-event filtering) now
# lives in app.coordinator.emitter and is wrapped around every emitter
# automatically by the runner.


async def _runner(initial: TurnState, emitter: EventEmitter) -> None:
    """Single-Coordinator turn runner. PresentationEmitter sanitises
    every event before it hits the wire."""
    user_emitter = _CoordinatorPresentation(emitter)
    try:
        await run_query_turn(initial, user_emitter.emit)
    except Exception:
        log.exception("coordinator crashed for turn %s", initial.turn_id)
        # User-safe error: never leak the exception class or message.
        await user_emitter.emit("agent.result", envelope(
            "We hit an issue answering your question.",
            detail="The analytics pipeline encountered an error. Try again or rephrase.",
            kind="internal",
        ))
        await user_emitter.emit("turn.end", {
            "turn_id": initial.turn_id,
            "errors": ["pipeline_error"],
            "final_answer": None,
        })
    finally:
        await emitter.close()


async def _heartbeat(emitter: EventEmitter) -> None:
    try:
        while True:
            await asyncio.sleep(HEARTBEAT_SECONDS)
            try:
                await emitter.comment("ping")
            except Exception:
                return
    except asyncio.CancelledError:
        pass


async def _emit_pre_stream_error(emitter: EventEmitter, payload: dict) -> None:
    fid = str(uuid4())
    try:
        await emitter.emit("turn.start", {"turn_id": fid, "question": "", "mode": "ERROR"})
        await emitter.emit("agent.result", payload)
        await emitter.emit("turn.end", {
            "turn_id": fid,
            "errors": [payload.get("error", "unknown")],
            "final_answer": None,
        })
    finally:
        await emitter.close()


def _stream_response(
    emitter: EventEmitter,
    *tasks: asyncio.Task | None,
) -> StreamingResponse:
    async def gen():
        try:
            async for chunk in emitter.stream():
                yield chunk
        except asyncio.CancelledError:
            raise
        except Exception as e:
            yield format_sse("agent.result", envelope(
                f"{type(e).__name__}: {e}", kind="internal", message=SAFE_MESSAGE,
            ))
            yield format_sse("turn.end", {"turn_id": "fatal", "errors": [str(e)]})
        finally:
            for t in tasks:
                if t is not None and not t.done():
                    try:
                        t.cancel()
                    except Exception:
                        pass
    return StreamingResponse(
        gen(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",
            "Connection": "keep-alive",
        },
    )


@api_router.post("/query_stream")
async def query_stream(req: QueryRequest, request: Request):
    emitter = EventEmitter()
    pre_error = _validate_pre_stream(req)
    if pre_error is not None:
        _safe_create_task(_emit_pre_stream_error(emitter, pre_error))
        return _stream_response(emitter)

    rate_error = _rate_limit_check(_client_ip(request))
    if rate_error is not None:
        _safe_create_task(_emit_pre_stream_error(emitter, rate_error))
        return _stream_response(emitter)

    set_request_context(
        conversation_id=req.conversation_id,
        question_chars=len(req.question),
    )

    # Resolve the calling Principal so the turn carries its tenant. With AUTH
    # off this is the implicit single-tenant ("public") principal, so behaviour
    # is unchanged; with AUTH on it enforces a valid Bearer token (401 here
    # before any streaming begins).
    from app.tenant_context import require_principal
    principal = await require_principal(request)

    answer_language = (req.language or "en").strip().lower()
    if answer_language not in ANSWER_LANGUAGES:
        answer_language = "en"

    initial = TurnState(
        question=req.question,
        conversation_id=req.conversation_id,
        tenant_id=principal.tenant_id,
        answer_language=answer_language,
    )
    runner_task = _safe_create_task(_runner(initial, emitter))
    heartbeat_task = _safe_create_task(_heartbeat(emitter))
    return _stream_response(emitter, runner_task, heartbeat_task)


# ===========================================================================
# 3. FASTAPI APP — middleware, exception handlers, startup
# ===========================================================================

def configure_logging(level: int = logging.INFO) -> None:
    # Force UTF-8 on the underlying streams so unicode in log messages
    # (e.g. the `→` arrow used by dynamic_ingest) renders cleanly on
    # Windows consoles instead of leaking as escaped `→`.
    for stream in (sys.stdout, sys.stderr):
        reconfigure = getattr(stream, "reconfigure", None)
        if callable(reconfigure):
            try:
                reconfigure(encoding="utf-8", errors="replace")
            except Exception:
                pass
    logging.basicConfig(
        level=level,
        format="%(asctime)s %(levelname)s %(name)s %(message)s",
    )
    for noisy in ("httpx", "httpcore", "uvicorn.access"):
        logging.getLogger(noisy).setLevel(logging.WARNING)


configure_logging()
_app_log = logging.getLogger("agentic_ai")

init_sentry()


@asynccontextmanager
async def lifespan(_app: FastAPI):
    # Two-phase startup:
    #   1. Critical-path (blocks before `yield`): init the DB so no request
    #      can race with an un-opened SQLite file.  init_database() is fast.
    #   2. Heavy stuff (KPI rebuild, hierarchy sync, enrichment, LLM probe,
    #      vector vocab) runs in a background asyncio task so the HTTP
    #      server binds immediately and Render's health-check grace period
    #      isn't consumed.  Every step inside _startup() is already wrapped
    #      in try/except, so errors are logged but the server stays up.
    # Postgres-required guard. MetricAi no longer supports SQLite as a
    # runtime engine; refuse to start without a Postgres DATABASE_URL.
    # This lives in the lifespan startup path ONLY (not at import/settings
    # time) so direct unit tests that call functions in isolation are
    # unaffected.
    from app.db_engine import is_postgres
    if not is_postgres():
        raise RuntimeError(
            "MetricAi now requires PostgreSQL. Set DATABASE_URL to your Supabase/Postgres "
            "connection string. (SQLite is no longer a supported runtime engine.)"
        )

    try:
        from app.infrastructure import init_database
        await init_database()
        _app_log.info("MetricAi %s: database initialised (critical path done)", _app.version)
    except Exception:
        _app_log.exception("init_database() failed — DB-backed endpoints will return errors")

    _startup_task = asyncio.create_task(_startup())
    _app_log.info(
        "MetricAi %s: HTTP server accepting connections (background startup running)",
        _app.version,
    )
    try:
        yield
    finally:
        # Give the background task a chance to finish before we tear down
        # the DB connections it may still be using.
        if not _startup_task.done():
            _app_log.info("waiting for background startup to complete before shutdown…")
            try:
                await asyncio.wait_for(_startup_task, timeout=120.0)
            except (asyncio.TimeoutError, asyncio.CancelledError, Exception):
                _startup_task.cancel()
        await _shutdown()


app = FastAPI(
    title="MetricAi",
    description=(
        "Local-first single-user analytics powered by a Coordinator that "
        "talks to a local Qwen 3 model through Ollama."
    ),
    version="5.2.0-coordinator",
    lifespan=lifespan,
)

# Public routes that must NEVER require auth — load-balancer health probe,
# the login endpoint itself, the session probe, and the FastAPI/MCP
# discovery surface. Everything else is gated when AUTH_ENABLED=true.
# ---------------------------------------------------------------------------
# Admin portal API (owner-only). A SEPARATE static admin site flips customer
# access via these routes. Guarded by the ADMIN_PORTAL_TOKEN secret in the
# `X-Admin-Token` header — NOT the customer Bearer auth (so "/admin" is in
# _AUTH_PUBLIC_PREFIXES below). Reads/writes `public.users` across all tenants:
# the one legitimate cross-tenant surface, reachable only by the token holder.
# Fails CLOSED when the token is unset.
# ---------------------------------------------------------------------------
_VALID_ACCESS_STATUS = frozenset({"pending", "allowed", "denied"})


def _require_admin(request: Request) -> None:
    """Reject unless the request carries the exact admin portal token.

    Constant-time compare; fails closed when ADMIN_PORTAL_TOKEN is unset so a
    misconfigured deploy never exposes access-control to the public.
    """
    import hmac as _hmac

    configured = (settings.admin_portal_token or "").strip()
    supplied = (request.headers.get("X-Admin-Token") or "").strip()
    if not configured or not supplied or not _hmac.compare_digest(configured, supplied):
        raise HTTPException(status_code=403, detail="Forbidden")


class AdminStatusBody(BaseModel):
    model_config = ConfigDict(extra="ignore")
    status: str = Field(default="", max_length=16)


class AdminNotesBody(BaseModel):
    model_config = ConfigDict(extra="ignore")
    admin_notes: str | None = Field(default=None, max_length=2000)


@api_router.get("/admin/users")
async def admin_list_users(request: Request):
    """List every account for the admin portal (owner-only)."""
    _require_admin(request)
    from app.infrastructure import get_connection

    async with get_connection() as db:
        cur = await db.execute(
            "SELECT id, email, access_status, admin_notes, created_at "
            "FROM public.users ORDER BY created_at DESC"
        )
        rows = await cur.fetchall()
        await cur.close()
    return {"users": [dict(r) for r in rows]}


@api_router.post("/admin/users/{user_id}/status")
async def admin_set_status(user_id: str, body: AdminStatusBody, request: Request):
    """Set a customer's access_status to pending / allowed / denied."""
    _require_admin(request)
    new_status = (body.status or "").strip().lower()
    if new_status not in _VALID_ACCESS_STATUS:
        raise HTTPException(status_code=400, detail="invalid status")
    from app.infrastructure import get_connection

    async with get_connection() as db:
        await db.execute(
            "UPDATE public.users SET access_status = ? WHERE id = ?",
            (new_status, user_id),
        )
        await db.commit()
    # Bust the enforcement cache so the change takes effect on the next request.
    from app import identity
    identity.invalidate_access_status(user_id)
    return {"ok": True, "id": user_id, "access_status": new_status}


@api_router.post("/admin/users/{user_id}/notes")
async def admin_set_notes(user_id: str, body: AdminNotesBody, request: Request):
    """Update the owner's free-text notes for a customer."""
    _require_admin(request)
    from app.infrastructure import get_connection

    async with get_connection() as db:
        await db.execute(
            "UPDATE public.users SET admin_notes = ? WHERE id = ?",
            (body.admin_notes, user_id),
        )
        await db.commit()
    return {"ok": True}


_AUTH_PUBLIC_PREFIXES = (
    "/health",
    "/admin",
    "/auth/login",
    "/auth/signup",
    "/auth/me",
    "/auth/forgot",
    "/auth/reset",
    "/auth/google/login",
    "/auth/google/callback",
    "/docs",
    "/openapi.json",
    "/redoc",
)


@app.middleware("http")
async def _auth_middleware(request: Request, call_next):
    """Single touch-point for Bearer-token enforcement. Disabled when
    AUTH_ENABLED=false (the historical single-user MVP mode) so a
    missing env var on Render does not brick the deploy.
    """
    if not settings.auth_enabled:
        return await call_next(request)
    # CORS preflight must always pass — the browser fires OPTIONS before
    # any Authorization header is reachable.
    if request.method == "OPTIONS":
        return await call_next(request)
    path = request.url.path or "/"
    if any(path == p or path.startswith(p + "/") or path == p for p in _AUTH_PUBLIC_PREFIXES):
        return await call_next(request)
    # Defer the actual check to the helper so the logic stays in one place.
    from app.identity import verify_token, get_access_status
    hdr = request.headers.get("Authorization") or ""
    token = hdr[7:].strip() if hdr.lower().startswith("bearer ") else ""
    principal = verify_token(token) if token else None
    if principal is None:
        return JSONResponse(
            status_code=401,
            content=envelope(
                "Unauthorized",
                detail="Missing or invalid Bearer token",
                kind="auth",
            ),
            headers={"WWW-Authenticate": 'Bearer realm="agentic-ai"'},
        )
    # Access enforcement (Phase 2). OFF by default; when ACCESS_ENFORCEMENT is
    # on, block any customer who isn't 'allowed' from the data routes. The
    # status lookup is cached (identity.get_access_status), so this adds at
    # most one DB read per user every few minutes. /auth/me stays reachable
    # (it's a public prefix) so the frontend can still learn the status and
    # show the right blocked screen.
    if settings.access_enforcement:
        access = await get_access_status(principal.user_id)
        if access != "allowed":
            body = envelope(
                "Access blocked",
                detail=(
                    "Your account is pending approval."
                    if access == "pending"
                    else "Your MetricAI subscription is inactive."
                ),
                kind="access",
            )
            body["access_status"] = access
            return JSONResponse(status_code=403, content=body)
    return await call_next(request)

app.include_router(api_router)

# Mount the FastMCP server at /mcp so external MCP clients (Claude Desktop,
# the MCP inspector, other agents) can list + ingest Google Drive files.
# Guarded: if `fastmcp` is not installed the rest of the app still boots.
# The mounted ASGI app has its own lifespan (the streamable-HTTP session
# manager) which Starlette does NOT auto-run for mounted sub-apps — the
# _startup / _shutdown hooks below drive it manually.
_mcp_sub_app = None
_mcp_lifespan_cm = None
try:
    from app.mcp_server import mcp_app

    _mcp_sub_app = mcp_app()
    app.mount("/mcp", _mcp_sub_app)
    log.info("FastMCP server mounted at /mcp")
except Exception:
    log.warning("FastMCP server not mounted (fastmcp unavailable?)", exc_info=True)

instrument_fastapi(app)

# CORS is registered LAST on purpose. Starlette runs the most recently
# added middleware OUTERMOST, and CORS must wrap everything: an early-return
# 401 from the auth middleware above would otherwise leave without an
# Access-Control-Allow-Origin header, and browsers mask such responses as
# "TypeError: Failed to fetch" — the frontend can't tell "session expired"
# from "backend down". Resulting stack, outermost → innermost:
# CORS → request-context/security-headers → auth → routes.
_raw_origins = [o.strip() for o in settings.allowed_origins.split(",") if o.strip()]
_allow_all = _raw_origins == ["*"]

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"] if _allow_all else _raw_origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["*"],
)


def _cors_headers_for(request: Request) -> dict[str, str]:
    """CORS headers for responses that BYPASS CORSMiddleware.

    Starlette hoists the ``@app.exception_handler(Exception)`` handler into
    ServerErrorMiddleware, which is built OUTSIDE every user middleware —
    including CORS, no matter the registration order. Without these headers
    an unhandled-500 envelope reaches the browser CORS-blocked and gets
    masked as "TypeError: Failed to fetch" (the 2026-08-03 fake-network-error
    symptom, on the 500 path). Mirrors CORSMiddleware's origin policy.
    """
    origin = (request.headers.get("origin") or "").strip()
    if not origin or not (_allow_all or origin in _raw_origins):
        return {}
    return {
        "Access-Control-Allow-Origin": origin,
        "Access-Control-Allow-Credentials": "true",
        "Vary": "Origin",
    }


def _sanitize_validation_errors(errors: list) -> list:
    safe: list = []
    for err in errors:
        e = dict(err) if isinstance(err, dict) else {"msg": str(err)}
        inp = e.get("input")
        if isinstance(inp, (bytes, bytearray)):
            try:
                e["input"] = inp.decode("utf-8", errors="replace")
            except Exception:
                e["input"] = repr(inp)
        elif inp is not None and not isinstance(inp, (str, int, float, bool, list, dict, type(None))):
            e["input"] = repr(inp)
        safe.append(e)
    return safe


@app.exception_handler(RequestValidationError)
async def validation_exception_handler(_: Request, exc: RequestValidationError):
    try:
        detail = _sanitize_validation_errors(exc.errors())
    except Exception:
        detail = [{"msg": "Validation failed"}]
    return JSONResponse(
        status_code=400,
        content=envelope("Invalid request body", detail=str(detail), kind="validation",
                         extra={"validation_errors": detail}),
    )


@app.exception_handler(Exception)
async def unhandled_exception_handler(request: Request, exc: Exception):
    _app_log.exception("Unhandled exception in %s %s", request.method, request.url.path)
    # Record into the central error log so the /errors dashboard sees it.
    try:
        error_id = log_error(
            exc=exc,
            module=_module_for_path(request.url.path),
            endpoint=request.url.path,
            method=request.method,
            user_facing=False,
            source="core_system.unhandled_exception_handler",
        )
    except Exception:
        error_id = None
    return JSONResponse(
        status_code=500,
        content=envelope(
            str(exc) or type(exc).__name__,
            detail=type(exc).__name__,
            kind="internal",
            extra={"error_id": error_id} if error_id else None,
        ),
        # This handler renders OUTSIDE CORSMiddleware (see _cors_headers_for)
        # so it must attach the CORS headers itself or browsers hide the 500.
        headers=_cors_headers_for(request),
    )


def _module_for_path(path: str) -> str:
    """Best-effort path → module mapping for the error_log table."""
    p = (path or "").lower()
    if p.startswith("/upload"):     return "upload"
    if p.startswith("/uploads"):    return "upload"
    if p.startswith("/dashboard"):  return "dashboard"
    if p.startswith("/query"):      return "ai"
    if p.startswith("/auth"):       return "auth"
    if p.startswith("/drive"):      return "auth"
    if p.startswith("/datasets"):   return "upload"
    if p.startswith("/cache"):      return "system"
    if p.startswith("/time"):       return "time_engine"
    if p.startswith("/health"):     return "system"
    if p.startswith("/errors"):     return "system"
    return "system"


async def _startup() -> None:
    from app.infrastructure import init_database, list_uploads_meta, load_synonyms
    from app.vector import register_vocabulary

    await init_database()
    _app_log.info("database engine: %s", engine_kind())

    # Registry reconcile — rebuild dynamic_tables.json from SQLite so AI
    # queries work immediately after restart even if the JSON was lost.
    try:
        rec = reconcile_registry()
        if rec["added"] or rec["removed"]:
            _app_log.info("dynamic table registry reconciled: %s", rec)
        else:
            _app_log.info("dynamic table registry: up to date")
    except Exception:
        _app_log.exception("dynamic table registry reconcile failed (continuing)")

    # Time engine — invalidate any stale cached tokens, then probe once so
    # the operator sees the current dataset date in the boot log.
    try:
        invalidate_time_cache()
        tokens = await resolve_dataset_date_tokens()
        if tokens:
            _app_log.info(
                "time engine: dataset_today=%s month=%s prev_month=%s",
                tokens.get("dataset_today"), tokens.get("dataset_month"),
                tokens.get("dataset_prev_month"),
            )
        else:
            _app_log.info("time engine: no uploaded data yet — relative-time KPIs will fail until first upload")
    except Exception:
        _app_log.exception("time engine probe failed (continuing)")

    # Coordinator registry — Phase 2: 4 capabilities only.
    # Sub-agents (sqlWriter, rcaReasoner, insightFmt) are called internally
    # by the capabilities and must NOT appear in the public registry — the
    # LLM must not be able to call them directly.
    try:
        from app.coordinator.tools import build_default_registry
        registry = build_default_registry()
        _app_log.info(
            "coordinator registry: %d capabilities registered: %s",
            len(registry.names()), registry.names(),
        )
        _app_log.info("financial DB ready at %s", settings.financial_db_path)
    except Exception:
        _app_log.exception("coordinator registry bootstrap failed (continuing — /query_stream will surface the error)")

    # Local LLM health check - fail loudly on startup if Ollama is down
    # or the configured model isn't pulled. No cloud fallback.
    try:
        ok, msg = await check_llm_health()
        if ok:
            _app_log.info("LLM health: %s", msg)
        else:
            _app_log.error("LLM health FAILED: %s", msg)
    except Exception:
        _app_log.exception("LLM health check crashed (continuing - /query_stream will surface the error)")

    # Report persistent dataset state at boot so the operator can confirm
    # data survived the restart. SQLite tables hold the parsed rows;
    # data/uploads/ holds the original source files.
    try:
        sales_rows = await count_rows("sales")
        purchase_rows = await count_rows("purchase")
        uploads_meta = await list_uploads_meta(limit=1000)
        active_uploads = [u for u in uploads_meta if u.get("status") == "active"]
        _app_log.info(
            "persistent data: sales=%d rows, purchase=%d rows, "
            "active datasets=%d (total uploads=%d)",
            sales_rows, purchase_rows, len(active_uploads), len(uploads_meta),
        )
        if active_uploads:
            for u in active_uploads[:5]:
                _app_log.info(
                    "  dataset batch=%s file=%s target=%s rows=%d uploaded_at=%s",
                    u["batch_id"][:8],
                    u.get("filename"),
                    u.get("target"),
                    u.get("rows_inserted"),
                    u.get("uploaded_at"),
                )
    except Exception:
        _app_log.exception("could not summarize persistent data at startup")

    try:
        syns = load_synonyms()
        if syns:
            n = register_vocabulary("entity", syns)
            _app_log.info(
                "vector vocabulary bootstrapped: %d canonicals from synonyms.json", n,
            )
    except Exception:
        _app_log.exception("vector vocabulary bootstrap failed (continuing without)")

    _app_log.info("sentry: %s", sentry_status())
    _app_log.info(
        "auth: %s",
        "ENABLED (login required)" if settings.auth_enabled else "DISABLED — all routes public",
    )

    # Drive the mounted FastMCP sub-app's lifespan — Starlette does not run
    # lifespans of mounted ASGI apps, so the streamable-HTTP session manager
    # would otherwise never start. Best-effort: a failure here only disables
    # /mcp, the rest of the app stays up.
    global _mcp_lifespan_cm
    if _mcp_sub_app is not None:
        try:
            _mcp_lifespan_cm = _mcp_sub_app.router.lifespan_context(_mcp_sub_app)
            await _mcp_lifespan_cm.__aenter__()
            _app_log.info("FastMCP /mcp lifespan started")
        except Exception:
            _mcp_lifespan_cm = None
            _app_log.exception("FastMCP /mcp lifespan failed to start (continuing)")


async def _shutdown() -> None:
    global _mcp_lifespan_cm
    if _mcp_lifespan_cm is not None:
        try:
            await _mcp_lifespan_cm.__aexit__(None, None, None)
        except Exception:
            _app_log.warning("FastMCP /mcp lifespan shutdown error", exc_info=True)
        finally:
            _mcp_lifespan_cm = None
