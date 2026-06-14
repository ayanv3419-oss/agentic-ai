"""Dynamic multi-sheet xlsx ingestion (ADR-0005).

Each sheet in an uploaded workbook becomes its own SQLite table named
``u_<snake_case_sheet_name>``. Columns are derived from the sheet's header
row; types are inferred from the first ~200 data rows.

The set of dynamic tables is recorded in ``data/dynamic_tables.json`` so the
LLM prompt and SQL validator know what's available between server restarts.

Hard separation from the legacy sales/purchase pipeline:
    • Dynamic tables ALWAYS carry the ``u_`` prefix — never collides with
      static system tables (sales, purchase, product_hierarchy, etc.).
    • Re-uploading a sheet DROPs and recreates its table (matches the
      "Replace" semantics the user picked during planning).
    • Header detection here is LIBERAL (any row with ≥3 string-looking
      cells, where the next row has at least one non-string value or
      different shape) — distinct from the strict alias-driven detection
      used by sales/purchase ingestion.
"""
from __future__ import annotations

import asyncio
import csv
import io
import json
import logging
import os
import re
import sqlite3
import tempfile
from datetime import date as _date_cls, datetime
from pathlib import Path
from threading import Lock
from typing import Any, Callable, Iterator

from openpyxl import Workbook, load_workbook

from app.infrastructure import db_path, get_connection, quoted, settings

log = logging.getLogger("agentic_ai.dynamic_ingest")

# Optional per-sheet progress hook: callable(done, total, msg). Used by the
# async /upload_workbook background task to surface ingestion progress.
ProgressCb = Callable[[int, int, str], None]


def _emit_progress(cb: "ProgressCb | None", done: int, total: int, msg: str) -> None:
    """Invoke a progress callback, swallowing any error. Progress reporting
    must never abort or corrupt an ingest."""
    if cb is None:
        return
    try:
        cb(done, total, msg)
    except Exception:  # pragma: no cover - defensive
        log.warning("progress_cb raised; ignoring", exc_info=True)

# Tables here NEVER collide with these static system tables.
_RESERVED_TABLE_NAMES: frozenset[str] = frozenset({
    "sales", "purchase", "sales_archive", "purchase_archive",
    "uploads", "users", "memberships", "workspaces",
    "conversations", "v2_conversation_turns", "v2_execution_log",
    "v2_shadow_log", "kpi_registry", "analytics_cache",
    "audit_logs", "feedback", "error_log",
    "product_hierarchy", "product_hierarchy_v2", "product_master",
    "product_sku_master", "product_cost_master",
    "location_hierarchy", "branch_master",
    "sku_inventory", "sku_forecast",
    "sqlite_sequence",
})

# Header detection thresholds.
_MIN_STRING_CELLS_FOR_HEADER = 3
_MAX_HEADER_SCAN_ROWS = 20
_TYPE_INFER_SAMPLE = 2000

_PREFIX = "u_"
_REGISTRY_LOCK = Lock()


# ---------------------------------------------------------------------------
# Identifier sanitization
# ---------------------------------------------------------------------------

_SNAKE_RE = re.compile(r"[^a-z0-9]+")


def sanitize_identifier(raw: str) -> str:
    """Lowercase, snake_case, alphanumeric+underscore only. Leading digits
    get prefixed with `_`."""
    s = (raw or "").strip().lower()
    s = _SNAKE_RE.sub("_", s)
    s = s.strip("_")
    if not s:
        s = "col"
    if s[0].isdigit():
        s = "_" + s
    return s


def sheet_to_table_name(sheet: str) -> str:
    """`Sales_Transactions` → `u_sales_transactions`.

    The ``u_`` prefix is sufficient on its own to avoid collisions with
    every system table — no further mangling needed.
    """
    return f"{_PREFIX}{sanitize_identifier(sheet)}"


def _dedupe_columns(cols: list[str]) -> list[str]:
    """Sanitize + ensure uniqueness. `Quantity` appearing twice becomes
    `quantity` and `quantity_2`."""
    seen: dict[str, int] = {}
    out: list[str] = []
    for raw in cols:
        c = sanitize_identifier(raw)
        if c in seen:
            seen[c] += 1
            c = f"{c}_{seen[c]}"
        else:
            seen[c] = 1
        out.append(c)
    return out


# ---------------------------------------------------------------------------
# Header detection — liberal
# ---------------------------------------------------------------------------

def _is_string_like(v: Any) -> bool:
    if v is None:
        return False
    if isinstance(v, (int, float, bool)):
        return False
    if isinstance(v, datetime):
        return False
    s = str(v).strip()
    if not s:
        return False
    # Pure numbers parse as floats but openpyxl already returned int/float;
    # surviving str values here are header-like.
    return True


def _row_has_any_value(row: list[Any] | None) -> bool:
    if not row:
        return False
    return any(v is not None and str(v).strip() != "" for v in row)


def _next_non_empty(rows: list[list[Any]], i: int) -> list[Any] | None:
    """Skip blank rows after index i and return the next row that has any
    value, or None."""
    for j in range(i + 1, len(rows)):
        if _row_has_any_value(rows[j]):
            return rows[j]
    return None


def _find_header_row(rows: list[list[Any]]) -> int | None:
    """Return the 0-based index of the first row that looks like a header.

    A header is a row that:
      • has ≥ _MIN_STRING_CELLS_FOR_HEADER non-empty STRING cells, AND
      • is followed (possibly across blank rows) by a row that contains at
        least one VALUE — preferably a non-string (number/date), which is
        strong evidence the surrounding region is data.

    Picks the candidate whose NEXT non-empty row has the highest "data-like"
    score (numeric cells > string cells). Falls back to the first
    sufficiently-string-heavy row if no candidate has a clear data follower.
    """
    scored: list[tuple[float, int]] = []
    for i, row in enumerate(rows):
        if i >= _MAX_HEADER_SCAN_ROWS:
            break
        if not row:
            continue
        string_count = sum(1 for c in row if _is_string_like(c))
        if string_count < _MIN_STRING_CELLS_FOR_HEADER:
            continue
        # Header rows are mostly-strings (no numbers, no dates).
        non_string_in_row = sum(
            1 for c in row
            if c is not None and not _is_string_like(c)
        )
        if non_string_in_row > string_count:
            # More numbers than strings — looks like a data row, not header.
            continue
        next_row = _next_non_empty(rows, i)
        if next_row is None:
            # Header with nothing after it — keep as last-resort candidate.
            scored.append((string_count * 0.1, i))
            continue
        next_numeric = sum(
            1 for v in next_row
            if v is not None and not _is_string_like(v)
        )
        next_values = sum(
            1 for v in next_row
            if v is not None and str(v).strip()
        )
        # Strong signal: next row has any numeric value (= data).
        # Weak signal: next row has values at all.
        score = next_numeric * 10 + next_values + string_count
        scored.append((score, i))
    if not scored:
        return None
    scored.sort(key=lambda t: (-t[0], t[1]))
    return scored[0][1]


# ---------------------------------------------------------------------------
# Type inference
# ---------------------------------------------------------------------------

_DATE_PATTERNS: tuple[re.Pattern, ...] = (
    re.compile(r"^\d{4}-\d{2}-\d{2}$"),
    re.compile(r"^\d{2}/\d{2}/\d{4}$"),
    re.compile(r"^\d{2}-\d{2}-\d{4}$"),
)


def _parse_date_str(s: str, prefer_mdy: bool = False) -> str | None:
    """Return ISO-8601 YYYY-MM-DD or None.

    For ambiguous slash-separated dates (both DD/MM and MM/DD would parse),
    ``prefer_mdy=True`` tries MM/DD/YYYY first (US style); otherwise DD/MM/YYYY
    is tried first (ISO world default). Pass ``prefer_mdy`` based on a
    per-column sample from ``_infer_date_order``.
    """
    s = s.strip()
    if not s:
        return None
    # ISO and hyphen forms are unambiguous — try first.
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    # Slash form (DD/MM or MM/DD) — order depends on column heuristic.
    slash_fmts = (
        ("%m/%d/%Y", "%d/%m/%Y") if prefer_mdy else ("%d/%m/%Y", "%m/%d/%Y")
    )
    for fmt in slash_fmts:
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def _infer_date_order(samples: list[Any]) -> bool:
    """Return True if MM/DD/YYYY is more likely than DD/MM/YYYY for this column.

    Looks for values where the first token is >12 (unambiguously a day,
    ruling out MM/DD) vs. the second token being >12 (unambiguously a month
    overflow, ruling out DD/MM). If more samples suggest MM/DD, return True.
    """
    mdy_votes = 0
    dmy_votes = 0
    for v in samples:
        if v is None:
            continue
        sv = str(v).strip()
        parts = sv.split("/")
        if len(parts) != 3:
            continue
        try:
            a, b = int(parts[0]), int(parts[1])
        except ValueError:
            continue
        if a > 12:
            dmy_votes += 1  # first token can only be a day → DD/MM
        elif b > 12:
            mdy_votes += 1  # second token can only be a day → MM/DD
    return mdy_votes > dmy_votes


def _to_number_str(value: Any) -> Any:
    """Strip thousands separators + surrounding whitespace from a string
    value so it parses as a number; non-strings pass through unchanged.
    Mirrors the comma-stripping ``_infer_column_type`` does — so a
    "1,200" column inferred REAL also COERCES instead of becoming NULL."""
    if isinstance(value, str):
        return value.strip().replace(",", "")
    return value


def _coerce(value: Any, sql_type: str, prefer_mdy: bool = False) -> Any:
    """Convert a python value into something SQLite-friendly for the
    inferred column type."""
    if value is None:
        return None
    if isinstance(value, datetime):
        # Strip the time component on TEXT columns to normalize dates.
        if sql_type == "TEXT":
            return value.strftime("%Y-%m-%d")
        return value.isoformat()
    if sql_type == "TEXT":
        if isinstance(value, str):
            s = value.strip()
            iso = _parse_date_str(s, prefer_mdy=prefer_mdy)
            return iso if iso is not None else s
        return str(value)
    if sql_type == "INTEGER":
        try:
            return int(float(_to_number_str(value)))
        except (TypeError, ValueError):
            return None
    if sql_type == "REAL":
        try:
            return float(_to_number_str(value))
        except (TypeError, ValueError):
            return None
    return value


def _infer_column_type(samples: list[Any]) -> str:
    """Decide a SQLite type from a sample of values. REAL > INTEGER > TEXT.

    Ignores Nones. Anything that doesn't parse cleanly as a number forces TEXT.
    """
    saw_real = False
    saw_int = False
    saw_any = False
    for v in samples:
        if v is None:
            continue
        if isinstance(v, bool):
            return "INTEGER"
        if isinstance(v, datetime):
            return "TEXT"
        saw_any = True
        if isinstance(v, int):
            saw_int = True
            continue
        if isinstance(v, float):
            saw_real = True
            continue
        # str — try to parse as number; if it parses, treat as numeric
        s = str(v).strip()
        if not s:
            continue
        # Date-looking strings -> TEXT
        if any(p.match(s) for p in _DATE_PATTERNS):
            return "TEXT"
        try:
            int(s)
            saw_int = True
        except ValueError:
            try:
                float(s.replace(",", ""))
                saw_real = True
            except ValueError:
                return "TEXT"
    if not saw_any:
        return "TEXT"
    if saw_real:
        return "REAL"
    if saw_int:
        return "INTEGER"
    return "TEXT"


# ---------------------------------------------------------------------------
# Registry — persisted to data/dynamic_tables.json
# ---------------------------------------------------------------------------

def _registry_path() -> Path:
    return Path(settings.financial_db_path).parent / "dynamic_tables.json"


def load_registry() -> dict[str, dict[str, Any]]:
    """Return ``{table_name: {columns: [...], source_sheet, source_file, row_count, uploaded_at}}``.
    Missing file → empty dict."""
    p = _registry_path()
    if not p.exists():
        return {}
    try:
        with p.open("r", encoding="utf-8") as f:
            data = json.load(f)
        return data if isinstance(data, dict) else {}
    except Exception:
        log.warning("dynamic_tables.json load failed; treating as empty", exc_info=True)
        return {}


def _save_registry(reg: dict[str, dict[str, Any]]) -> None:
    p = _registry_path()
    p.parent.mkdir(parents=True, exist_ok=True)
    tmp = p.with_suffix(".json.tmp")
    with tmp.open("w", encoding="utf-8") as f:
        json.dump(reg, f, indent=2, ensure_ascii=False, default=str)
    tmp.replace(p)


def register_table(
    table: str,
    *,
    columns: list[dict[str, str]],
    source_sheet: str,
    source_file: str,
    row_count: int,
) -> None:
    """Upsert a dynamic table entry. Thread-safe."""
    with _REGISTRY_LOCK:
        reg = load_registry()
        reg[table] = {
            "columns":      columns,
            "source_sheet": source_sheet,
            "source_file":  source_file,
            "row_count":    row_count,
            "uploaded_at":  datetime.now().astimezone().isoformat(timespec="seconds"),
        }
        _save_registry(reg)


def unregister_table(table: str) -> None:
    with _REGISTRY_LOCK:
        reg = load_registry()
        reg.pop(table, None)
        _save_registry(reg)


def reconcile_registry() -> dict[str, int]:
    """Reconcile data/dynamic_tables.json against the live SQLite DB.

    On every startup:
    - Adds entries for u_* tables that exist in the DB but not the registry
      (covers the case where dynamic_tables.json was deleted or corrupted).
    - Removes entries whose table no longer exists in the DB.

    Column metadata is recovered from PRAGMA table_info; source_file and
    uploaded_at are cross-referenced from the uploads table via _batch_id.

    Returns {'added': N, 'removed': M}.

    On Postgres this is currently a no-op — the schema tool queries
    information_schema live for every turn so a stale JSON registry is
    harmless for the agentic-loop path. Re-port over the weekend if the
    JSON registry needs to stay accurate for other consumers.
    """
    from app.db_engine import is_postgres
    if is_postgres():
        log.info("reconcile_registry: skipped on Postgres (no-op)")
        return {"added": 0, "removed": 0}
    with _REGISTRY_LOCK:
        reg = load_registry()

        conn = sqlite3.connect(str(db_path()))
        try:
            rows = conn.execute(
                "SELECT name FROM sqlite_master WHERE type='table' AND name LIKE 'u_%'"
            ).fetchall()
            live_tables: set[str] = {r[0] for r in rows}
            reg_tables: set[str] = set(reg.keys())

            missing = live_tables - reg_tables    # in DB, not in JSON
            orphaned = reg_tables - live_tables   # in JSON, not in DB

            if not missing and not orphaned:
                return {"added": 0, "removed": 0}

            for t in orphaned:
                reg.pop(t, None)

            try:
                upload_rows = conn.execute(
                    "SELECT batch_id, filename, uploaded_at FROM uploads"
                ).fetchall()
                uploads_by_batch: dict[str, tuple[str, str]] = {
                    r[0]: (r[1] or "(recovered)", r[2] or "")
                    for r in upload_rows
                }
            except Exception:
                log.warning(
                    "reconcile_registry: uploads table read failed; "
                    "recovered tables will lack source-file metadata",
                    exc_info=True,
                )
                uploads_by_batch = {}

            for table in sorted(missing):
                try:
                    col_rows = conn.execute(
                        f'PRAGMA table_info("{table}")'
                    ).fetchall()
                    columns = [
                        {"name": r[1], "type": r[2] or "TEXT"}
                        for r in col_rows
                        if not r[1].startswith("_")
                    ]

                    try:
                        row_count: int = conn.execute(
                            f'SELECT COUNT(*) FROM "{table}"'
                        ).fetchone()[0]
                    except Exception:
                        log.warning(
                            "reconcile_registry: COUNT failed for %r; "
                            "registering with row_count=0", table, exc_info=True,
                        )
                        row_count = 0

                    source_file = "(recovered)"
                    uploaded_at = datetime.now().astimezone().isoformat(timespec="seconds")
                    try:
                        batch_row = conn.execute(
                            f'SELECT DISTINCT _batch_id FROM "{table}" LIMIT 1'
                        ).fetchone()
                        if batch_row and batch_row[0] in uploads_by_batch:
                            fname, u_at = uploads_by_batch[batch_row[0]]
                            source_file = fname
                            if u_at:
                                uploaded_at = u_at
                    except Exception:
                        log.debug(
                            "reconcile_registry: _batch_id lookup failed for %r; "
                            "leaving source_file as '(recovered)'",
                            table, exc_info=True,
                        )

                    sheet_name = (
                        table[2:].replace("_", " ").title()
                        if table.startswith("u_") else table
                    )

                    reg[table] = {
                        "columns":      columns,
                        "source_sheet": sheet_name,
                        "source_file":  source_file,
                        "row_count":    row_count,
                        "uploaded_at":  uploaded_at,
                    }
                except Exception:
                    log.warning(
                        "reconcile_registry: could not rebuild entry for %r (skipping)",
                        table, exc_info=True,
                    )
        finally:
            conn.close()

        _save_registry(reg)

    added = len(missing)
    removed = len(orphaned)
    if added or removed:
        log.info(
            "dynamic table registry reconciled: added=%d removed=%d live=%d",
            added, removed, len(live_tables),
        )
    return {"added": added, "removed": removed}


# ---------------------------------------------------------------------------
# Per-sheet ingestion
# ---------------------------------------------------------------------------

class DynamicIngestError(ValueError):
    """Sheet-level failure: no detectable header / no usable rows."""


def _iter_sheet_rows(ws) -> Iterator[list[Any]]:
    for row in ws.iter_rows(values_only=True):
        yield list(row) if row is not None else []


def _build_create_sql(table: str, cols: list[tuple[str, str]]) -> str:
    col_defs = ",\n  ".join(f'"{name}" {sqltype}' for name, sqltype in cols)
    return (
        f'CREATE TABLE "{table}" (\n'
        f'  _id INTEGER PRIMARY KEY AUTOINCREMENT,\n'
        f'  _batch_id TEXT NOT NULL,\n'
        f'  _source_sheet TEXT NOT NULL,\n'
        f'  _inserted_at TEXT NOT NULL DEFAULT (datetime(\'now\')),\n'
        f'  {col_defs}\n'
        f')'
    )


# SQLite TEXT/INTEGER/REAL → Postgres mapping. Used as a fallback only;
# `_refine_pg_type` produces richer types (DATE, TIMESTAMPTZ, NUMERIC) when
# the column's name and samples warrant it.
_PG_TYPE_MAP: dict[str, str] = {
    "TEXT": "TEXT",
    "INTEGER": "BIGINT",
    "REAL": "DOUBLE PRECISION",
}

# Column-name hints used to upgrade a SQLite-inferred type to a richer
# Postgres type. NUMERIC(18,2) is preferred over DOUBLE PRECISION for any
# REAL column whose name reads as money — IEEE-754 floats introduce drift
# on millions of rows and compare unreliably for equality.
_MONEY_HINT = re.compile(
    r"(amount|price|value|revenue|cost|sales|net|gross|total|"
    r"profit|margin|turnover|expense|fee|tax|discount|charge|"
    r"sale|paid|due|balance)",
    re.IGNORECASE,
)
# Column names that lean timestamp-of-day over plain date.
_TIMESTAMP_HINT = re.compile(
    r"(timestamp|datetime|created|updated|modified|inserted|_at\b|_ts\b)",
    re.IGNORECASE,
)


def _looks_like_timestamp_str(s: str) -> bool:
    """Quick check for a string that carries a time component (HH:MM)."""
    return ("T" in s and ":" in s) or (" " in s and ":" in s and "-" in s)


def _refine_pg_type(name: str, sql_type: str, samples: list[Any]) -> str:
    """Pick a richer Postgres column type than the static _PG_TYPE_MAP.

    Rules:
      - INTEGER → BIGINT (unchanged)
      - REAL, money-named → NUMERIC(18,2)
      - REAL, otherwise → DOUBLE PRECISION
      - TEXT, ≥90% of non-null samples parse as ISO date → DATE
        (TIMESTAMPTZ if name hints time or samples carry HH:MM)
      - TEXT, otherwise → TEXT
    """
    if sql_type == "INTEGER":
        return "BIGINT"
    if sql_type == "REAL":
        if _MONEY_HINT.search(name):
            return "NUMERIC(18,2)"
        return "DOUBLE PRECISION"
    # TEXT — try to upgrade to DATE / TIMESTAMPTZ when the values look it.
    non_null = 0
    date_count = 0
    ts_count = 0
    for v in samples[:50]:
        if v is None:
            continue
        if isinstance(v, datetime):
            non_null += 1
            date_count += 1
            ts_count += 1
            continue
        if isinstance(v, _date_cls):
            non_null += 1
            date_count += 1
            continue
        s = str(v).strip()
        if not s:
            continue
        non_null += 1
        if _parse_date_str(s) is not None:
            date_count += 1
            if _looks_like_timestamp_str(s):
                ts_count += 1
    if non_null >= 3 and date_count / non_null >= 0.9:
        if _TIMESTAMP_HINT.search(name) or (ts_count >= non_null / 2):
            return "TIMESTAMPTZ"
        return "DATE"
    return "TEXT"


def _coerce_pg(value: Any, pg_type: str, prefer_mdy: bool = False) -> Any:
    """Convert a raw Excel cell into something asyncpg accepts for ``pg_type``.

    DATE → datetime.date or None.
    TIMESTAMPTZ → datetime.datetime or None.
    NUMERIC(18,2) / DOUBLE PRECISION → float or None.
    BIGINT → int or None.
    TEXT → str or None.
    """
    if value is None:
        return None
    if pg_type == "TEXT":
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d")
        if isinstance(value, _date_cls):
            return value.strftime("%Y-%m-%d")
        s = str(value).strip()
        return s if s else None
    if pg_type == "BIGINT":
        try:
            return int(float(_to_number_str(value)))
        except (TypeError, ValueError):
            return None
    if pg_type in ("DOUBLE PRECISION", "NUMERIC(18,2)"):
        try:
            return float(_to_number_str(value))
        except (TypeError, ValueError):
            return None
    if pg_type == "DATE":
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, _date_cls):
            return value
        s = str(value).strip()
        if not s:
            return None
        iso = _parse_date_str(s, prefer_mdy=prefer_mdy)
        if iso is None:
            return None
        return _date_cls.fromisoformat(iso)
    if pg_type == "TIMESTAMPTZ":
        if isinstance(value, datetime):
            return value
        if isinstance(value, _date_cls):
            return datetime.combine(value, datetime.min.time())
        s = str(value).strip()
        if not s:
            return None
        for fmt in (
            "%Y-%m-%dT%H:%M:%S",
            "%Y-%m-%d %H:%M:%S",
            "%Y-%m-%dT%H:%M:%S.%f",
            "%Y-%m-%d %H:%M:%S.%f",
        ):
            try:
                return datetime.strptime(s, fmt)
            except ValueError:
                continue
        iso = _parse_date_str(s, prefer_mdy=prefer_mdy)
        if iso is not None:
            return datetime.strptime(iso, "%Y-%m-%d")
        return None
    return value


def _build_create_sql_pg(table: str, cols: list[tuple[str, str]]) -> str:
    """Build CREATE TABLE DDL for Postgres.

    ``cols`` is a list of (column_name, pg_type) pairs where pg_type is the
    REFINED Postgres type (one of TEXT, BIGINT, DOUBLE PRECISION,
    NUMERIC(18,2), DATE, TIMESTAMPTZ). Callers must call _refine_pg_type
    BEFORE building DDL so the richer types are honoured.
    """
    col_defs = ",\n  ".join(
        f'"{name}" {pgtype}' for name, pgtype in cols
    )
    return (
        f'CREATE TABLE "{table}" (\n'
        f'  _id BIGSERIAL PRIMARY KEY,\n'
        f'  _batch_id TEXT NOT NULL,\n'
        f'  _source_sheet TEXT NOT NULL,\n'
        f'  _inserted_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),\n'
        f'  {col_defs}\n'
        f')'
    )


def ingest_sheet(
    *,
    wb_path: Path,
    sheet_name: str,
    source_file_name: str,
    batch_id: str,
    table_name: str | None = None,
) -> dict[str, Any]:
    """Drop + create + insert one sheet into ``u_<sheet_name>``.

    Synchronous SQLite. Called from inside the async upload route via
    ``run_in_executor`` if needed; for the multi-sheet loop a synchronous
    call inside the request is fine (single-user MVP).

    Returns
    -------
    summary : dict
        {table, columns, header_row, rows_inserted, dropped_existing, skipped_rows}
    """
    wb = load_workbook(filename=str(wb_path), read_only=True, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise DynamicIngestError(f"sheet not found: {sheet_name!r}")
        ws = wb[sheet_name]

        # Collect first ~20 rows for header detection
        buffered: list[list[Any]] = []
        row_iter = _iter_sheet_rows(ws)
        for _ in range(_MAX_HEADER_SCAN_ROWS):
            try:
                buffered.append(next(row_iter))
            except StopIteration:
                break
        if not buffered:
            raise DynamicIngestError(f"sheet {sheet_name!r} is empty")

        hdr_idx = _find_header_row(buffered)
        if hdr_idx is None:
            raise DynamicIngestError(
                f"could not detect a header row in first {_MAX_HEADER_SCAN_ROWS} "
                f"rows of sheet {sheet_name!r}"
            )
        header_raw = [str(c).strip() if c is not None else "" for c in buffered[hdr_idx]]
        # Drop trailing all-empty header cells, keep first non-empty span.
        while header_raw and not header_raw[-1]:
            header_raw.pop()
        if not header_raw:
            raise DynamicIngestError(f"sheet {sheet_name!r} has empty header row")
        col_names = _dedupe_columns(header_raw)
        n_cols = len(col_names)

        # Collect data rows (post-header) for type inference + insert
        data_rows: list[list[Any]] = []
        for r in buffered[hdr_idx + 1:]:
            if r is None:
                continue
            trimmed = r[:n_cols] + [None] * max(0, n_cols - len(r))
            if any(v is not None and str(v).strip() != "" for v in trimmed):
                data_rows.append(trimmed)
        # Continue draining the iterator
        for r in row_iter:
            trimmed = r[:n_cols] + [None] * max(0, n_cols - len(r))
            if any(v is not None and str(v).strip() != "" for v in trimmed):
                data_rows.append(trimmed)

        if not data_rows:
            raise DynamicIngestError(f"sheet {sheet_name!r} has no data rows after header")

        # Infer types from first N rows per column
        col_types: list[str] = []
        mdy_flags: list[bool] = []
        sample = data_rows[:_TYPE_INFER_SAMPLE]
        for i in range(n_cols):
            samples = [row[i] for row in sample]
            col_types.append(_infer_column_type(samples))
            mdy_flags.append(_infer_date_order(samples))

        table = table_name or sheet_to_table_name(sheet_name)

        # DROP + CREATE in raw sqlite3 (sync; simpler than aiosqlite here)
        con = sqlite3.connect(str(db_path()))
        try:
            cur = con.cursor()
            dropped = False
            try:
                row = cur.execute(
                    "SELECT name FROM sqlite_master WHERE type='table' AND name=?",
                    (table,),
                ).fetchone()
                if row is not None:
                    cur.execute(f'DROP TABLE "{table}"')
                    dropped = True
            except sqlite3.Error:
                pass
            cur.execute(_build_create_sql(table, list(zip(col_names, col_types))))

            # Bulk insert
            placeholders = ", ".join("?" * (n_cols + 2))  # +2 for _batch_id, _source_sheet
            quoted_cols = ", ".join(f'"{c}"' for c in col_names)
            insert_sql = (
                f'INSERT INTO "{table}" (_batch_id, _source_sheet, {quoted_cols}) '
                f'VALUES ({placeholders})'
            )
            payload = []
            skipped = 0
            for row in data_rows:
                try:
                    coerced = [_coerce(row[i], col_types[i], mdy_flags[i]) for i in range(n_cols)]
                    payload.append((batch_id, sheet_name, *coerced))
                except Exception:
                    skipped += 1
            cur.executemany(insert_sql, payload)
            con.commit()
            rows_inserted = len(payload)
        finally:
            con.close()

        # Persist registry entry. The table + data are already committed
        # above, so a registry write failure must NOT lose the data: swallow
        # it with a logged warning rather than letting it propagate (which
        # would leave the table on disk but invisible to the registry until a
        # restart with no error trail). reconcile_registry() re-adds any
        # committed u_* table missing from the JSON registry on next startup,
        # so the data stays discoverable either way.
        columns_meta = [
            {"name": name, "type": sqltype}
            for name, sqltype in zip(col_names, col_types)
        ]
        try:
            register_table(
                table,
                columns=columns_meta,
                source_sheet=sheet_name,
                source_file=source_file_name,
                row_count=rows_inserted,
            )
        except Exception:
            log.warning(
                "dynamic_ingest: register_table failed for table=%r; data is "
                "committed and reconcile_registry will recover it on next "
                "startup",
                table, exc_info=True,
            )

        log.info(
            "dynamic_ingest: sheet=%r → table=%r rows=%d cols=%d dropped=%s skipped=%d",
            sheet_name, table, rows_inserted, n_cols, dropped, skipped,
        )

        return {
            "table":             table,
            "source_sheet":      sheet_name,
            "header_row":        hdr_idx + 1,
            "columns":           columns_meta,
            "rows_inserted":     rows_inserted,
            "dropped_existing":  dropped,
            "skipped_rows":      skipped,
        }
    finally:
        try:
            wb.close()
        except Exception:
            pass


def ingest_workbook_sync(
    *,
    wb_path: Path,
    source_file_name: str,
    batch_id: str,
    progress_cb: "ProgressCb | None" = None,
) -> dict[str, Any]:
    """Loop every sheet in the workbook, ingest each as a dynamic table.

    Sheets where header detection fails are recorded under ``skipped`` with
    the reason, never aborting the whole upload.

    This is the SQLite path. For Postgres, see ``ingest_workbook_pg``.
    The public async ``ingest_workbook`` dispatches between them based on
    the active engine.
    """
    wb = load_workbook(filename=str(wb_path), read_only=True, data_only=True)
    try:
        sheet_names = list(wb.sheetnames)
    finally:
        try:
            wb.close()
        except Exception:
            pass

    total = len(sheet_names)
    _emit_progress(progress_cb, 0, total, "Reading workbook")

    # Pre-compute de-collided table names so two sheets whose names map to the
    # same u_<slug> (e.g. "Sales 2024" and "Sales-2024" both → u_sales_2024)
    # get unique names (u_sales_2024, u_sales_2024_2) instead of the second
    # sheet silently DROP+overwriting the first.
    seen_tables: dict[str, int] = {}
    sheet_table_map: dict[str, str] = {}
    for sn in sheet_names:
        base = sheet_to_table_name(sn)
        if base in seen_tables:
            seen_tables[base] += 1
            sheet_table_map[sn] = f"{base}_{seen_tables[base]}"
        else:
            seen_tables[base] = 1
            sheet_table_map[sn] = base

    ingested: list[dict[str, Any]] = []
    skipped: list[dict[str, str]] = []
    for done, sheet in enumerate(sheet_names, start=1):
        try:
            summary = ingest_sheet(
                wb_path=wb_path,
                sheet_name=sheet,
                source_file_name=source_file_name,
                batch_id=batch_id,
                table_name=sheet_table_map[sheet],
            )
            ingested.append(summary)
        except DynamicIngestError as e:
            log.warning("dynamic_ingest: skip sheet=%r reason=%s", sheet, e)
            skipped.append({"sheet": sheet, "reason": str(e)})
        except Exception as e:
            log.exception("dynamic_ingest: sheet=%r crashed", sheet)
            skipped.append({"sheet": sheet, "reason": f"{type(e).__name__}: {e}"})
        _emit_progress(progress_cb, done, total, f"Processed sheet {done}/{total}: {sheet}")

    return {
        "batch_id":      batch_id,
        "source_file":   source_file_name,
        "sheet_count":   len(sheet_names),
        "ingested":      ingested,
        "skipped":       skipped,
        "tables":        [s["table"] for s in ingested],
    }


async def ingest_sheet_pg(
    *,
    wb_path: Path,
    sheet_name: str,
    source_file_name: str,
    batch_id: str,
    table_name: str | None = None,
) -> dict[str, Any]:
    """Postgres counterpart of ``ingest_sheet``. Same XLSX parsing + type
    inference; differs only in the storage path (asyncpg via the shim).
    """
    from app.db_engine import pg_connection

    wb = await asyncio.to_thread(load_workbook, filename=str(wb_path), read_only=True, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise DynamicIngestError(f"sheet not found: {sheet_name!r}")
        ws = wb[sheet_name]

        buffered: list[list[Any]] = []
        row_iter = _iter_sheet_rows(ws)
        for _ in range(_MAX_HEADER_SCAN_ROWS):
            try:
                buffered.append(next(row_iter))
            except StopIteration:
                break
        if not buffered:
            raise DynamicIngestError(f"sheet {sheet_name!r} is empty")

        hdr_idx = _find_header_row(buffered)
        if hdr_idx is None:
            raise DynamicIngestError(
                f"could not detect a header row in first {_MAX_HEADER_SCAN_ROWS} "
                f"rows of sheet {sheet_name!r}"
            )
        header_raw = [str(c).strip() if c is not None else "" for c in buffered[hdr_idx]]
        while header_raw and not header_raw[-1]:
            header_raw.pop()
        if not header_raw:
            raise DynamicIngestError(f"sheet {sheet_name!r} has empty header row")
        col_names = _dedupe_columns(header_raw)
        n_cols = len(col_names)

        data_rows: list[list[Any]] = []
        for r in buffered[hdr_idx + 1:]:
            if r is None:
                continue
            trimmed = r[:n_cols] + [None] * max(0, n_cols - len(r))
            if any(v is not None and str(v).strip() != "" for v in trimmed):
                data_rows.append(trimmed)
        for r in row_iter:
            trimmed = r[:n_cols] + [None] * max(0, n_cols - len(r))
            if any(v is not None and str(v).strip() != "" for v in trimmed):
                data_rows.append(trimmed)

        if not data_rows:
            raise DynamicIngestError(f"sheet {sheet_name!r} has no data rows after header")

        # Two-stage typing: first the SQLite-flavoured inference (kept for
        # registry metadata + downstream callers), then a Postgres-flavoured
        # refinement that upgrades dates and money to native DATE / NUMERIC.
        col_types: list[str] = []
        pg_types: list[str] = []
        mdy_flags: list[bool] = []
        sample = data_rows[:_TYPE_INFER_SAMPLE]
        for i in range(n_cols):
            samples = [row[i] for row in sample]
            sql_t = _infer_column_type(samples)
            col_types.append(sql_t)
            pg_types.append(_refine_pg_type(col_names[i], sql_t, samples))
            mdy_flags.append(_infer_date_order(samples))

        table = table_name or sheet_to_table_name(sheet_name)

        skipped = 0
        payload: list[tuple[Any, ...]] = []
        for row in data_rows:
            try:
                coerced = [_coerce_pg(row[i], pg_types[i], mdy_flags[i]) for i in range(n_cols)]
                payload.append((batch_id, sheet_name, *coerced))
            except Exception:
                skipped += 1

        # DROP + CREATE + bulk INSERT — wrapped in one transaction so a crash
        # mid-INSERT ROLLBACKs and leaves the previous table contents intact
        # (the alternative is an empty CREATEd table replacing real data).
        async with pg_connection() as db:
            async with db.transaction():
                # CASCADE so re-upload doesn't crash when a FOREIGN KEY in
                # another u_* table references this one. Any dependent FKs
                # are dropped silently — `_post_ingest_refresh` reruns the
                # relationship detector after every ingest so the next
                # Schema-tool call still surfaces the join keys.
                await db.execute(f'DROP TABLE IF EXISTS "{table}" CASCADE')
                dropped = True  # IF EXISTS is idempotent; we don't track prior state
                create_sql = _build_create_sql_pg(table, list(zip(col_names, pg_types)))
                await db.execute(create_sql)

                if payload:
                    placeholders = ", ".join("?" * (n_cols + 2))
                    quoted_cols = ", ".join(f'"{c}"' for c in col_names)
                    insert_sql = (
                        f'INSERT INTO "{table}" (_batch_id, _source_sheet, {quoted_cols}) '
                        f'VALUES ({placeholders})'
                    )
                    await db.executemany(insert_sql, payload)

        rows_inserted = len(payload)

        columns_meta = [
            {"name": name, "type": sqltype}
            for name, sqltype in zip(col_names, col_types)
        ]
        # The table + data are already committed above; a registry write
        # failure must not lose them. Swallow it with a logged warning rather
        # than propagating (which would leave the table in Postgres but absent
        # from the JSON registry with no error trail). On Postgres the Schema
        # tool queries information_schema live every turn, so the data stays
        # discoverable to the agentic loop even when the registry entry is
        # missing.
        try:
            register_table(
                table,
                columns=columns_meta,
                source_sheet=sheet_name,
                source_file=source_file_name,
                row_count=rows_inserted,
            )
        except Exception:
            log.warning(
                "dynamic_ingest_pg: register_table failed for table=%r; data "
                "is committed and the Schema tool reads information_schema "
                "live, so it stays discoverable",
                table, exc_info=True,
            )

        # Profile every column so the resolver can prefer cleaner columns
        # when concept synonyms tie, and the Schema tool can surface data
        # quality (% null, % numeric, distinct count) to the LLM. Failures
        # never break the upload — profiling is best-effort.
        try:
            from app.schema_mapping.profiler import (
                profile_table_pg, save_profiles_pg,
            )
            profiles = await profile_table_pg(table, list(col_names))
            saved = await save_profiles_pg(profiles)
            log.info(
                "profile_pg: table=%r cols_profiled=%d saved=%d",
                table, len(profiles), saved,
            )
        except Exception:
            log.warning(
                "profile_pg: table=%r profiling failed (continuing)",
                table, exc_info=True,
            )

        log.info(
            "dynamic_ingest_pg: sheet=%r → table=%r rows=%d cols=%d skipped=%d",
            sheet_name, table, rows_inserted, n_cols, skipped,
        )

        return {
            "table":             table,
            "source_sheet":      sheet_name,
            "header_row":        hdr_idx + 1,
            "columns":           columns_meta,
            "rows_inserted":     rows_inserted,
            "dropped_existing":  dropped,
            "skipped_rows":      skipped,
        }
    finally:
        try:
            wb.close()
        except Exception:
            pass


async def ingest_workbook_pg(
    *,
    wb_path: Path,
    source_file_name: str,
    batch_id: str,
    progress_cb: "ProgressCb | None" = None,
) -> dict[str, Any]:
    """Async Postgres workbook ingest. Mirrors ingest_workbook_sync."""
    wb = await asyncio.to_thread(load_workbook, filename=str(wb_path), read_only=True, data_only=True)
    try:
        sheet_names = list(wb.sheetnames)
    finally:
        try:
            wb.close()
        except Exception:
            pass

    total = len(sheet_names)
    _emit_progress(progress_cb, 0, total, "Reading workbook")

    seen_tables: dict[str, int] = {}
    sheet_table_map: dict[str, str] = {}
    for sn in sheet_names:
        base = sheet_to_table_name(sn)
        if base in seen_tables:
            seen_tables[base] += 1
            sheet_table_map[sn] = f"{base}_{seen_tables[base]}"
        else:
            seen_tables[base] = 1
            sheet_table_map[sn] = base

    ingested: list[dict[str, Any]] = []
    skipped: list[dict[str, str]] = []
    for done, sheet in enumerate(sheet_names, start=1):
        try:
            summary = await ingest_sheet_pg(
                wb_path=wb_path,
                sheet_name=sheet,
                source_file_name=source_file_name,
                batch_id=batch_id,
                table_name=sheet_table_map[sheet],
            )
            ingested.append(summary)
        except DynamicIngestError as e:
            log.warning("dynamic_ingest_pg: skip sheet=%r reason=%s", sheet, e)
            skipped.append({"sheet": sheet, "reason": str(e)})
        except Exception as e:
            log.exception("dynamic_ingest_pg: sheet=%r crashed", sheet)
            skipped.append({"sheet": sheet, "reason": f"{type(e).__name__}: {e}"})
        _emit_progress(progress_cb, done, total, f"Processed sheet {done}/{total}: {sheet}")

    return {
        "batch_id":      batch_id,
        "source_file":   source_file_name,
        "sheet_count":   len(sheet_names),
        "ingested":      ingested,
        "skipped":       skipped,
        "tables":        [s["table"] for s in ingested],
    }


async def ingest_workbook(
    *,
    wb_path: Path,
    source_file_name: str,
    batch_id: str,
    progress_cb: "ProgressCb | None" = None,
) -> dict[str, Any]:
    """Async dispatcher — Postgres or SQLite based on DATABASE_URL.

    Was previously a sync function. The async signature is a small breaking
    change at the one call site (core_system.py upload route) which already
    runs inside an async handler — just `await` the call.

    ``progress_cb`` is an OPTIONAL ``callable(done: int, total: int, msg: str)``
    invoked once before the loop (``done=0``) and once after each sheet is
    ingested/skipped. It must be cheap and non-blocking (it runs on the event
    loop for the PG path, and in the worker thread for the SQLite path). Any
    exception it raises is swallowed so progress reporting can never break an
    upload.
    """
    import asyncio as _asyncio
    from app.db_engine import is_postgres
    if is_postgres():
        return await ingest_workbook_pg(
            wb_path=wb_path,
            source_file_name=source_file_name,
            batch_id=batch_id,
            progress_cb=progress_cb,
        )
    # SQLite path stays synchronous + uses raw sqlite3 — push it to a thread
    # so it doesn't block the event loop.
    return await _asyncio.to_thread(
        ingest_workbook_sync,
        wb_path=wb_path,
        source_file_name=source_file_name,
        batch_id=batch_id,
        progress_cb=progress_cb,
    )


# ---------------------------------------------------------------------------
# Single-table / CSV ingest — routed through the SAME dynamic u_* machinery
# ---------------------------------------------------------------------------

def _csv_grid_to_workbook_path(
    rows: list[list[Any]],
    *,
    sheet_title: str,
) -> Path:
    """Materialize a raw 2-D cell grid as a single-sheet ``.xlsx`` on disk and
    return its path. The caller owns the file and MUST unlink it.

    The grid is written verbatim — NO header detection, NO type coercion here.
    That keeps the synthetic workbook byte-identical (cell-for-cell) to what the
    user would have uploaded as an XLSX, so the downstream ``ingest_workbook``
    pipeline runs its own liberal header detection + type inference unchanged.

    The sheet title is sanitized to openpyxl's 31-char / no-special-chars limit.
    The downstream pipeline derives the persisted ``u_*`` table name from this
    worksheet title via ``sheet_to_table_name``, so a short clean stem like
    ``sales`` becomes ``u_sales``; very long or special-char stems are clamped
    to the 31-char title first (same name they would get as an XLSX sheet).
    """
    wb = Workbook(write_only=False)
    ws = wb.active
    # openpyxl rejects sheet titles >31 chars or containing []:*?/\\ — clamp it.
    safe_title = re.sub(r"[\[\]:*?/\\]", "_", (sheet_title or "Sheet1"))[:31] or "Sheet1"
    ws.title = safe_title
    for row in rows:
        ws.append(list(row) if row is not None else [])
    fd, tmp_name = tempfile.mkstemp(suffix=".xlsx", prefix="csv_ingest_")
    os.close(fd)
    tmp_path = Path(tmp_name)
    try:
        wb.save(str(tmp_path))
    finally:
        try:
            wb.close()
        except Exception:
            pass
    return tmp_path


async def ingest_csv(
    *,
    file_bytes: bytes,
    source_file_name: str,
    batch_id: str,
    progress_cb: "ProgressCb | None" = None,
) -> dict[str, Any]:
    """Ingest a SINGLE-TABLE upload (a CSV file) through the SAME dynamic
    ``u_*`` machinery the multi-sheet XLSX path uses — engine-agnostic.

    The CSV is treated as a synthetic single-sheet workbook: its raw cell grid
    is written to a throwaway ``.xlsx`` file, which is then handed to the public
    :func:`ingest_workbook` dispatcher. That means CSV ingestion reuses, with
    ZERO duplication, the exact header detection (:func:`_find_header_row`), type
    inference (:func:`_infer_column_type` / :func:`_refine_pg_type`), table
    creation, and transactional insert that XLSX uploads already use — and it
    branches SQLite vs Postgres for free, because ``ingest_workbook`` does
    (``ingest_workbook_pg`` on Postgres, ``ingest_workbook_sync`` on SQLite).

    The single table is named ``u_<snake_case(file_stem)>`` (e.g. ``sales.csv``
    → ``u_sales``), mirroring how a sheet name maps to a table name.

    Parameters
    ----------
    file_bytes : bytes
        Raw CSV file content. Decoded as UTF-8 (BOM-tolerant), latin-1 fallback.
    source_file_name : str
        Original filename (e.g. ``"sales.csv"``). Used for the registry
        ``source_file`` metadata and to derive the table / sheet name.
    batch_id : str
        Upload batch id, written to every row's ``_batch_id`` column.
    progress_cb : callable(done, total, msg) | None
        Optional progress hook, forwarded to ``ingest_workbook`` unchanged. For a
        single CSV the totals are (0, 1) then (1, 1).

    Returns
    -------
    summary : dict
        IDENTICAL shape to :func:`ingest_workbook` (the XLSX path), so the HTTP
        route can respond identically::

            {
              "batch_id":    <batch_id>,
              "source_file": <source_file_name>,
              "sheet_count": 1,
              "ingested": [ {                       # one entry for the CSV
                  "table":            "u_<stem>",
                  "source_sheet":     "<stem>",
                  "header_row":       <1-based int>,
                  "columns":          [{"name","type"}, ...],
                  "rows_inserted":    <int>,
                  "dropped_existing": <bool>,
                  "skipped_rows":     <int>,
              } ],
              "skipped":  [ {"sheet","reason"}, ... ],   # empty on success
              "tables":   ["u_<stem>"],
            }

    Raises
    ------
    DynamicIngestError
        If the CSV is empty / undecodable. A CSV with content but no detectable
        header does NOT raise here — like the XLSX path, the sheet lands in the
        returned ``skipped`` list with a reason, and ``ingested`` is empty.
    """
    # 1. Decode + parse the raw CSV into a 2-D grid (header detection happens
    #    LATER, inside the shared pipeline — we only preserve the raw cells).
    if not file_bytes:
        raise DynamicIngestError("CSV file is empty")
    try:
        text = file_bytes.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = file_bytes.decode("latin-1", errors="replace")
    grid: list[list[Any]] = list(csv.reader(io.StringIO(text)))
    # Drop fully-empty trailing rows the csv module sometimes yields.
    while grid and not any(str(c).strip() for c in grid[-1]):
        grid.pop()
    if not grid:
        raise DynamicIngestError("CSV has no rows")

    # 2. Derive the source sheet / table name from the filename stem so the
    #    persisted table is u_<stem> (sales.csv -> u_sales).
    stem = Path(source_file_name or "table").stem or "table"

    # 3. Materialize the grid as a synthetic single-sheet workbook and route it
    #    through the SAME engine-aware dispatcher the XLSX upload uses.
    tmp_path = await asyncio.to_thread(
        _csv_grid_to_workbook_path, grid, sheet_title=stem
    )
    try:
        summary = await ingest_workbook(
            wb_path=tmp_path,
            source_file_name=source_file_name,
            batch_id=batch_id,
            progress_cb=progress_cb,
        )
    finally:
        try:
            tmp_path.unlink()
        except Exception:
            log.warning(
                "ingest_csv: could not delete temp workbook %s", tmp_path,
                exc_info=True,
            )
    return summary


# ---------------------------------------------------------------------------
# Public helpers used by analytics_engine + core_system
# ---------------------------------------------------------------------------

def list_dynamic_tables() -> list[dict[str, Any]]:
    """All currently-registered dynamic tables with metadata.

    WARNING: registry-sourced and therefore GLOBAL across tenants. Do NOT use
    this for any per-request/user-facing listing — it leaks other tenants'
    table names, columns, source filenames, and row counts. The HTTP route
    GET /tables/dynamic uses :func:`list_dynamic_tables_for_tenant` instead,
    which introspects only the current tenant's schema. This sync helper is
    kept for internal callers (prompt schema summary, validator whitelist)
    that run outside a request/tenant context.
    """
    reg = load_registry()
    return [
        {"table": name, **meta}
        for name, meta in sorted(reg.items())
    ]


# Metadata columns every dynamic table carries (see _build_create_sql*).
# Excluded from the user-facing column list so the introspected shape matches
# what the registry historically reported.
_DYNAMIC_META_COLUMNS: frozenset[str] = frozenset(
    {"_id", "_batch_id", "_source_sheet", "_inserted_at"}
)


async def list_dynamic_tables_for_tenant() -> list[dict[str, Any]]:
    """Tenant-scoped dynamic-table listing — ZERO cross-tenant leakage.

    Unlike :func:`list_dynamic_tables` (which reads the global registry file),
    this introspects the CURRENT tenant's schema only, so it sees exactly the
    ``u_*`` tables that tenant owns. The caller is responsible for binding the
    query tenant (e.g. ``set_query_tenant``) BEFORE awaiting this, so that
    ``current_schema()`` / ``search_path`` already point at the tenant's schema.

    Each entry mirrors the legacy registry shape as closely as the per-tenant
    schema allows::

        {
          "table": "u_sales",
          "columns": [{"name": "...", "type": "..."}, ...],   # user cols only
          "row_count": <int>,
          "source_sheet": "...",          # from the table's _source_sheet
          "uploaded_at": "...",           # MAX(_inserted_at), ISO string
        }

    ``source_file`` is intentionally OMITTED: it is not tracked inside the
    tenant's own schema (only in the shared ``public.uploads`` table, keyed by
    batch_id), so sourcing it would either reintroduce cross-tenant exposure or
    require an unreliable join. Consumers must treat it as optional.
    """
    from app.db_engine import is_postgres

    out: list[dict[str, Any]] = []
    async with get_connection() as db:
        # 1. List this tenant's u_* tables (own schema only).
        if is_postgres():
            cur = await db.execute(
                "SELECT table_name AS name FROM information_schema.tables "
                "WHERE table_schema=current_schema() "
                "AND table_name LIKE 'u\\_%' ESCAPE '\\' "
                "ORDER BY table_name"
            )
        else:
            cur = await db.execute(
                "SELECT name FROM sqlite_master "
                "WHERE type='table' AND name LIKE 'u\\_%' ESCAPE '\\' "
                "ORDER BY name"
            )
        table_rows = await cur.fetchall()
        await cur.close()
        names = [n for n in (dict(r).get("name") for r in table_rows) if n]

        for name in names:
            # 2. Columns (name + type), excluding the internal _* metadata cols.
            if is_postgres():
                cur2 = await db.execute(
                    "SELECT column_name AS name, data_type AS type "
                    "FROM information_schema.columns "
                    "WHERE table_schema=current_schema() AND table_name=? "
                    "ORDER BY ordinal_position",
                    (name,),
                )
            else:
                cur2 = await db.execute(f"PRAGMA table_info({quoted(name)})")
            col_rows = await cur2.fetchall()
            await cur2.close()
            columns = [
                {"name": dict(c)["name"], "type": (dict(c).get("type") or "TEXT")}
                for c in col_rows
                if dict(c).get("name") not in _DYNAMIC_META_COLUMNS
            ]

            entry: dict[str, Any] = {"table": name, "columns": columns}

            # 3. row_count + source_sheet + uploaded_at, sourced from the
            #    tenant's OWN table (the _source_sheet / _inserted_at metadata
            #    columns written at ingest time). Best-effort: a freshly
            #    created/locked table shouldn't 500 the whole listing.
            try:
                cur3 = await db.execute(
                    f'SELECT COUNT(*) AS n, '
                    f'MIN(_source_sheet) AS source_sheet, '
                    f'MAX(_inserted_at) AS uploaded_at '
                    f'FROM {quoted(name)}'
                )
                agg = await cur3.fetchall()
                await cur3.close()
                if agg:
                    a = dict(agg[0])
                    entry["row_count"] = int(a.get("n") or 0)
                    if a.get("source_sheet") is not None:
                        entry["source_sheet"] = a["source_sheet"]
                    if a.get("uploaded_at") is not None:
                        # asyncpg may return a datetime; normalise to str so
                        # the JSON envelope matches the registry's ISO strings.
                        entry["uploaded_at"] = str(a["uploaded_at"])
            except Exception:
                log.warning(
                    "list_dynamic_tables_for_tenant: metadata aggregate failed "
                    "for %r; reporting row_count=0",
                    name,
                    exc_info=True,
                )
                entry.setdefault("row_count", 0)

            out.append(entry)
    return out


def dynamic_table_columns(table: str) -> list[dict[str, str]] | None:
    reg = load_registry()
    entry = reg.get(table)
    if entry is None:
        return None
    return entry.get("columns") or []


def dynamic_schema_summary() -> str:
    """Compact human-readable schema injected into the LLM system prompt.

    Returns "" when no dynamic tables exist so we don't pollute the prompt.
    """
    # TODO(tenant-leak): switch to list_dynamic_tables_for_tenant once caller is async.
    # This function is sync; its callers (LLM prompt builder in core_system) are also
    # sync-call sites, so converting would require a non-trivial async refactor.
    tables = list_dynamic_tables()
    if not tables:
        return ""
    lines = ["", "User-uploaded tables (raw, queryable via raw SQL):"]
    for t in tables:
        cols = ", ".join(
            f'"{c["name"]}" ({c["type"]})'
            for c in t.get("columns", [])
        )
        lines.append(
            f'  - "{t["table"]}" (from sheet "{t["source_sheet"]}", '
            f'{t["row_count"]} rows): {cols}'
        )
    return "\n".join(lines)


def known_dynamic_identifiers() -> set[str]:
    """All table + column names from the registry — for SQL validator
    whitelist."""
    # TODO(tenant-leak): switch to list_dynamic_tables_for_tenant once caller is async.
    # This function is sync; the SQL validator whitelist is built synchronously in the
    # coordinator tool, so converting would require a non-trivial async refactor.
    out: set[str] = set()
    for t in list_dynamic_tables():
        out.add(t["table"])
        for c in t.get("columns", []):
            out.add(c["name"])
        # Built-in metadata columns
        out.update({"_id", "_batch_id", "_source_sheet", "_inserted_at"})
    return out


async def drop_all_dynamic_tables() -> int:
    """Delete every registered dynamic table from the DB and clear the
    registry. Returns the count dropped. Used by /upload/disconnect_all."""
    reg = load_registry()
    count = 0
    async with get_connection() as conn:
        from app.db_engine import is_postgres
        cascade = " CASCADE" if is_postgres() else ""
        for name in list(reg.keys()):
            try:
                await conn.execute(f'DROP TABLE IF EXISTS "{name}"{cascade}')
                count += 1
            except Exception:
                log.warning("drop_all_dynamic_tables: failed on %r", name, exc_info=True)
        await conn.commit()
    with _REGISTRY_LOCK:
        _save_registry({})
    return count
