"""Storage + ingestion data layer — single-user local-first MVP.

Sections:
    1.  Settings          — pydantic-settings singleton loaded from .env
    2.  Errors            — `envelope` JSON helper + SAFE_MESSAGE constant
    3.  Schema            — SCHEMA_SPEC / ALLOWED_TABLES / HEADER_ALIASES
    4.  Connection        — async + sync SQLite helpers, DDL bootstrap
    5.  Upload registry   — uploads-table CRUD
    6.  Header detection  — alias index + header-row picker
    7.  Upload parsers    — CSV / XLSX streaming with auto header detect
    8.  Response cache    — `data/response_store.json` atomic JSON store
    9.  Memory / synonyms — entity-resolution backing store
"""
from __future__ import annotations

import csv
import hashlib
import io
import json
import logging
import os
import sqlite3
from contextlib import asynccontextmanager
from datetime import datetime, timezone
from functools import lru_cache
from pathlib import Path
from threading import Lock
from typing import Any, AsyncIterator, Iterable, Iterator

import aiosqlite
from openpyxl import load_workbook
from pydantic import Field
from pydantic_settings import BaseSettings, SettingsConfigDict

log = logging.getLogger("agentic_ai.database")


# ===========================================================================
# 1. SETTINGS
# ===========================================================================

PROJECT_ROOT = Path(__file__).resolve().parents[2]


def _abs(p: str) -> str:
    if not p:
        return p
    return p if os.path.isabs(p) else str(PROJECT_ROOT / p)


class Settings(BaseSettings):
    model_config = SettingsConfigDict(
        env_file=str(PROJECT_ROOT / "backend" / ".env"),
        extra="ignore",
        case_sensitive=False,
    )

    # --- Local LLM (Ollama + Qwen 3) -----------------------------------
    # The Coordinator talks to a LOCAL model through Ollama's OpenAI-
    # compatible endpoint. There is intentionally no cloud fallback - if
    # Ollama is unreachable the backend fails loudly at startup.
    llm_base_url: str = Field(
        default="http://localhost:11434/v1", alias="LLM_BASE_URL"
    )
    llm_api_key: str = Field(default="ollama", alias="LLM_API_KEY")
    llm_model: str = Field(default="qwen3:8b", alias="LLM_MODEL")
    llm_temperature: float = Field(default=0.0, alias="LLM_TEMPERATURE")
    llm_max_tokens: int = Field(default=900, alias="LLM_MAX_TOKENS")

    # --- LLM fallback ----------------------------------------------------
    # Auto-failover provider. When the primary LLM is rate-limited, out of
    # quota or unreachable, the SAME request is retried once against this
    # provider. Leave the key blank to disable failover.
    llm_fallback_base_url: str = Field(default="", alias="LLM_FALLBACK_BASE_URL")
    llm_fallback_api_key: str = Field(default="", alias="LLM_FALLBACK_API_KEY")
    llm_fallback_model: str = Field(default="", alias="LLM_FALLBACK_MODEL")

    # --- Cost / safety budgets ------------------------------------------
    max_loop_iterations: int = Field(default=8, alias="MAX_LOOP_ITERATIONS")
    cost_limit_usd: float = Field(default=1.0, alias="COST_LIMIT_USD")
    sql_max_bytes_scanned: int = Field(
        default=10 * 1024 * 1024 * 1024, alias="SQL_MAX_BYTES_SCANNED"
    )

    # --- Database (Supabase / Postgres) ---------------------------------
    # When set to a postgres://... URL, the backend uses asyncpg against
    # that database instead of the local SQLite file. See app/db_engine.py.
    database_url: str = Field(default="", alias="DATABASE_URL")

    # --- Storage paths (absolute) ---------------------------------------
    financial_db_path: str = Field(
        default=str(PROJECT_ROOT / "data" / "financial_records.db"),
        alias="FINANCIAL_DB_PATH",
    )
    response_store_path: str = Field(
        default=str(PROJECT_ROOT / "data" / "response_store.json"),
        alias="RESPONSE_STORE_PATH",
    )
    synonyms_path: str = Field(
        default=str(PROJECT_ROOT / "backend" / "memory" / "synonyms.json"),
        alias="SYNONYMS_PATH",
    )

    # --- Upload limits --------------------------------------------------
    max_upload_bytes: int = Field(default=50 * 1024 * 1024, alias="MAX_UPLOAD_BYTES")
    upload_chunk_bytes: int = Field(default=1024 * 1024, alias="UPLOAD_CHUNK_BYTES")

    # --- Rate limiting ---------------------------------------------------
    rate_limit_per_minute: int = Field(default=30, alias="RATE_LIMIT_PER_MINUTE")

    # --- Vector embeddings ----------------------------------------------
    # Dimension of the in-memory entity/synonym embedding store. Small by
    # default — enough for entity-space discrimination without burning RAM.
    vector_dim: int = Field(default=128, alias="VECTOR_DIM")

    # --- Auth (Phase 3) -------------------------------------------------
    # When False (default), every route is public — matches the historical
    # single-user-MVP behaviour and avoids bricking a deploy that hasn't
    # set credentials. Flip to True on Render once AUTH_TOKEN_SECRET is
    # configured; user accounts are self-serve via POST /auth/signup.
    auth_enabled:        bool = Field(default=False, alias="AUTH_ENABLED")
    # Sentinel value used as the in-source default. If auth is enabled and
    # AUTH_TOKEN_SECRET is still this value (i.e. the deployer forgot to
    # set a real one), the app refuses to start — see `_validate_auth()`
    # called below. Keeps the previous dev-mode UX (auth off → no config
    # required) without ever shipping a guessable token signer in prod.
    auth_token_secret:   str  = Field(default="dev-auth-secret-CHANGE-ME",
                                      alias="AUTH_TOKEN_SECRET")
    auth_token_ttl_hours: int = Field(default=168,   alias="AUTH_TOKEN_TTL_HOURS")

    # --- Admin portal ---------------------------------------------------
    # Shared secret for the owner-only admin portal (a SEPARATE static site).
    # Admin API routes (/admin/*) require this exact value in the
    # `X-Admin-Token` header. Empty (default) = admin API disabled: every
    # /admin call is rejected, so a deploy that forgets to set it fails
    # closed rather than exposing customer access-control to the world.
    admin_portal_token:  str  = Field(default="", alias="ADMIN_PORTAL_TOKEN")

    # --- Access enforcement (Phase 2) -----------------------------------
    # When True, the auth middleware blocks any customer whose access_status
    # is not 'allowed' (i.e. pending/denied) from the data routes. Default
    # False so deploying the code changes NOTHING until it is deliberately
    # flipped on — a safe rollout switch.
    access_enforcement: bool = Field(default=False, alias="ACCESS_ENFORCEMENT")

    # --- Server ---------------------------------------------------------
    host: str = Field(default="0.0.0.0", alias="HOST")
    port: int = Field(default=8000, alias="PORT")
    reload: bool = Field(default=False, alias="RELOAD")

    # --- CORS -----------------------------------------------------------
    # Comma-separated list of allowed origins. "*" means allow all (dev only).
    # Example: https://myapp.vercel.app,https://myapp.com
    allowed_origins: str = Field(default="*", alias="ALLOWED_ORIGINS")

    # --- Observability (Sentry — optional, no-op when DSN empty) -------
    sentry_dsn:                    str = Field(default="", alias="SENTRY_DSN")
    sentry_environment:            str = Field(default="development", alias="SENTRY_ENVIRONMENT")
    sentry_traces_sample_rate:     float = Field(default=0.1, alias="SENTRY_TRACES_SAMPLE_RATE")
    sentry_profiles_sample_rate:   float = Field(default=0.0, alias="SENTRY_PROFILES_SAMPLE_RATE")
    sentry_send_default_pii:       bool = Field(default=False, alias="SENTRY_SEND_DEFAULT_PII")

    # --- Google Drive OAuth (Drive routes/tool stay inert until set) ----
    google_client_id:     str = Field(default="", alias="GOOGLE_CLIENT_ID")
    google_client_secret: str = Field(default="", alias="GOOGLE_CLIENT_SECRET")
    google_redirect_uri:  str = Field(
        default="http://localhost:8000/auth/google/callback",
        alias="GOOGLE_REDIRECT_URI",
    )
    google_token_path: str = Field(
        default=str(PROJECT_ROOT / "data" / "google_token.json"),
        alias="GOOGLE_TOKEN_PATH",
    )
    # Where the OAuth callback bounces the browser back to after success.
    frontend_url: str = Field(default="http://localhost:5173", alias="FRONTEND_URL")

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.financial_db_path = _abs(self.financial_db_path)
        self.response_store_path = _abs(self.response_store_path)
        self.synonyms_path = _abs(self.synonyms_path)
        self.google_token_path = _abs(self.google_token_path)
        self._validate_auth()

    # Default token secret value — must be overridden when auth is on.
    _AUTH_DEFAULT_SECRET = "dev-auth-secret-CHANGE-ME"
    # Minimum acceptable AUTH_TOKEN_SECRET length (chars) when auth is on. A
    # short signer is brute-forceable; 32 chars is the conventional floor for a
    # token-signing secret (≈256 bits if random).
    _AUTH_SECRET_MIN_LEN = 32

    def _on_postgres(self) -> bool:
        """True when this deployment is backed by Postgres — the multi-tenant /
        production substrate — as signalled by a ``DATABASE_URL`` *in the
        process environment*.

        This is the same notion as ``app.db_engine.is_postgres()`` and uses the
        same prefix test, but with two deliberate differences that make it both
        SAFE and CORRECT for a constructor-time security guard:

        * It does NOT call ``db_engine.is_postgres()``. That function's fallback
          calls ``get_settings()`` — the very ``@lru_cache``d call still
          executing this constructor (cache entry not yet populated) — so
          invoking it from ``__init__`` recurses. We import it below purely to
          anchor intent for readers; the actual check is inlined.
        * It reads ``os.environ['DATABASE_URL']`` ONLY and intentionally skips
          the ``.env``-file fallback that ``db_engine.database_url()`` adds.
          Rationale: production (Render) injects ``DATABASE_URL`` as a real
          environment variable (render.yaml: ``sync: false`` dashboard secret),
          so the env var is the authoritative "this is prod/multi-tenant"
          signal. The unit suite forces SQLite by clearing
          ``os.environ['DATABASE_URL']`` at import; keying off the env var alone
          guarantees the guard can NEVER trip the SQLite test suite (or a local
          dev box) merely because ``backend/.env`` happens to hold a Postgres
          URL for convenience. ``self.database_url`` is therefore not consulted
          here.
        """
        from app.db_engine import is_postgres  # noqa: F401  (intent anchor; see docstring)
        url = os.environ.get("DATABASE_URL", "").strip()
        return url.startswith("postgres://") or url.startswith("postgresql://")

    def _validate_auth(self) -> None:
        """Refuse to start with insecure auth configuration.

        Two classes of check:

        1. STARTUP GUARD (Postgres + auth off) — Postgres is the multi-tenant /
           prod substrate, so running it with ``AUTH_ENABLED=false`` leaves the
           entire data plane public. We refuse to boot in that combination. This
           fires ONLY when ``_on_postgres()`` is true, so local SQLite single-
           user dev and the (SQLite-backed) unit suite are unaffected.

        2. AUTH-ON config (only when ``auth_enabled=True``) — AUTH_TOKEN_SECRET
           must be non-empty, must differ from the in-source sentinel, and must
           be at least ``_AUTH_SECRET_MIN_LEN`` chars. This catches the
           "AUTH_ENABLED=true but I forgot to set a real/strong secret" deploy
           bug that would otherwise sign tokens with a guessable string. User
           accounts are self-serve via ``POST /auth/signup`` (no separate admin
           login), so ADMIN_USERNAME/ADMIN_PASSWORD are not required.
        """
        # --- Startup guard: Postgres (prod/multi-tenant) demands auth ON -----
        # Gated on Postgres ONLY. SQLite (local dev / unit tests) is never
        # touched: they keep the historical no-auth single-user behaviour.
        if self._on_postgres() and not self.auth_enabled:
            raise RuntimeError(
                "Refusing to start: DATABASE_URL points at Postgres (the "
                "multi-tenant / production data plane) but AUTH_ENABLED is "
                "false, which would leave every tenant's data publicly "
                "accessible. Set AUTH_ENABLED=true (and a strong "
                "AUTH_TOKEN_SECRET) on this deployment, or point DATABASE_URL "
                "at a local SQLite database for single-user development."
            )

        if not self.auth_enabled:
            return
        problems: list[str] = []
        if not self.auth_token_secret or self.auth_token_secret == self._AUTH_DEFAULT_SECRET:
            problems.append(
                "AUTH_TOKEN_SECRET is unset or still the in-source default — "
                "set it to a strong random string."
            )
        elif len(self.auth_token_secret) < self._AUTH_SECRET_MIN_LEN:
            problems.append(
                f"AUTH_TOKEN_SECRET is too short "
                f"({len(self.auth_token_secret)} chars) — use at least "
                f"{self._AUTH_SECRET_MIN_LEN} characters of strong random "
                "entropy (e.g. `python -c \"import secrets; "
                "print(secrets.token_urlsafe(48))\"`)."
            )
        if problems:
            raise RuntimeError(
                "AUTH_ENABLED=true but auth config is insecure:\n  - "
                + "\n  - ".join(problems)
                + "\nFix these env vars (Render dashboard) or set "
                "AUTH_ENABLED=false to fall back to no-auth mode."
            )


@lru_cache
def get_settings() -> Settings:
    return Settings()


settings = get_settings()


# ===========================================================================
# 2. ERRORS — envelope helper used by every JSON route
# ===========================================================================

SAFE_MESSAGE = "Something went wrong (safely handled)"

ErrorKind = str  # "validation" | "auth" | "upload" | "internal" | ...


def envelope(
    error: str,
    *,
    detail: str | None = None,
    kind: ErrorKind = "internal",
    message: str = SAFE_MESSAGE,
    extra: dict[str, Any] | None = None,
) -> dict[str, Any]:
    body: dict[str, Any] = {
        "error": error,
        "detail": detail or "",
        "kind": kind,
        "message": message,
    }
    if extra:
        body.update(extra)
    return body


# ===========================================================================
# 3. SCHEMA — single source of truth for the financial_records.db tables.
# ===========================================================================

SCHEMA_SPEC: list[tuple[str, str, bool]] = [
    ("Date",                 "TEXT", True),
    ("Order No",             "TEXT", False),
    ("Invoice No",           "TEXT", False),
    ("Party Name",           "TEXT", False),
    ("Product Name",         "TEXT", False),
    ("GSTIN",                "TEXT", False),
    ("Party Phone No.",      "TEXT", False),
    ("Transaction Type",     "TEXT", False),
    ("Total Amount",         "REAL", True),
    ("Loyalty Redeemed",     "REAL", False),
    ("Payment Type",         "TEXT", False),
    ("Received/Paid Amount", "REAL", False),
    ("Balance Due",          "REAL", False),
    ("Payment Status",       "TEXT", False),
    ("Description",          "TEXT", False),
    ("Cash",                 "REAL", False),
    ("BHARAT PAY",           "REAL", False),
    ("Credit",               "REAL", False),
    ("Debit Cards",          "REAL", False),
    ("HDFC BANK",            "REAL", False),
]

SCHEMA_COLUMNS: list[str] = [name for name, _t, _r in SCHEMA_SPEC]
COLUMN_TYPES: dict[str, str] = {name: t for name, t, _r in SCHEMA_SPEC}
REQUIRED_COLUMNS: list[str] = [name for name, _t, r in SCHEMA_SPEC if r]
OPTIONAL_COLUMNS: list[str] = [name for name, _t, r in SCHEMA_SPEC if not r]

ALLOWED_TABLES: tuple[str, ...] = ("sales", "purchase")

HEADER_ALIASES: dict[str, list[str]] = {
    "Date": [
        "date", "dt", "sale date", "sales date", "purchase date",
        "transaction date", "txn date", "trans date",
        "order date", "bill date", "invoice date", "voucher date",
    ],
    "Order No": [
        "order no", "order number", "order #", "ord no", "ordno",
        "sl no", "sl. no", "s. no", "serial no", "sr no", "sno",
    ],
    "Invoice No": [
        "invoice no", "invoice number", "invoice #", "inv no",
        "bill no", "bill no.", "bill number", "voucher no", "voucher number",
    ],
    "Party Name": [
        "party name", "party", "name", "customer", "customer name",
        "client", "client name", "buyer",
        "supplier", "supplier name", "vendor", "vendor name",
    ],
    "Product Name": [
        "product name", "product", "item", "item name",
        "sku", "sku name", "brand", "brand name", "model",
        "article", "article name", "variant",
    ],
    "GSTIN": ["gstin", "gst no", "gst number", "gst", "gst id"],
    "Party Phone No.": [
        "party phone no", "party phone", "phone", "phone no",
        "phone number", "mobile", "mobile no", "mobile number",
        "contact", "contact no", "contact number",
    ],
    "Transaction Type": ["transaction type", "type", "txn type", "trans type"],
    "Total Amount": [
        "total amount", "total", "amount", "amt", "total amt",
        "grand total", "net amount", "bill amount", "invoice amount",
        "value", "final amount", "sale amount",
    ],
    "Loyalty Redeemed": [
        "loyalty redeemed", "loyalty", "loyalty points",
        "points redeemed", "loyalty pts", "reward points",
    ],
    "Payment Type": [
        "payment type", "payment method", "mode", "pay mode",
        "payment mode", "method",
    ],
    "Received/Paid Amount": [
        "received/paid amount", "received paid amount",
        "received amount", "paid amount", "received", "paid",
        "amount paid", "amount received",
    ],
    "Balance Due": [
        "balance due", "balance", "due", "outstanding",
        "due amount", "amount due", "remaining",
    ],
    "Payment Status": ["payment status", "status", "pay status"],
    "Description": [
        "description", "desc", "note", "notes", "remarks",
        "narration", "particulars",
    ],
    "Cash":        ["cash"],
    "BHARAT PAY":  ["bharat pay", "bharatpay", "bhim", "upi", "bhim upi"],
    "Credit":      ["credit"],
    "Debit Cards": ["debit cards", "debit card", "debit"],
    "HDFC BANK":   ["hdfc bank", "hdfc"],
}


def quoted(name: str) -> str:
    """SQL-quote an identifier that may contain spaces / dots / slashes."""
    return '"' + str(name).replace('"', '""') + '"'


def schema_dict() -> dict:
    tables: dict[str, dict] = {}
    for table in ALLOWED_TABLES:
        tables[table] = {
            "description": (
                "Customer sales transactions"
                if table == "sales"
                else "Supplier purchase transactions"
            ),
            "columns": [
                {"name": name, "type": t, "required": r, "nullable": not r}
                for name, t, r in SCHEMA_SPEC
            ],
            "indexes": ["Date", "Party Name", "batch_id"],
        }
    return {
        "dialect": "sqlite",
        "database_file": "data/financial_records.db",
        "tables": tables,
        "notes": [
            "Date column is always ISO YYYY-MM-DD (normalized at ingest).",
            "Optional columns are SQL NULL when missing — never the string '0' or 0.0.",
            'Always double-quote columns that contain spaces or dots: "Total Amount", "Party Name", "Party Phone No.".',
        ],
    }


# ===========================================================================
# 4. CONNECTION — DDL bootstrap + async/sync SQLite helpers
# ===========================================================================

def db_path() -> Path:
    return Path(settings.financial_db_path)


def _build_create(table: str) -> str:
    cols = []
    for name, sql_type, required in SCHEMA_SPEC:
        null_clause = "NOT NULL" if required else "NULL"
        cols.append(f'  {quoted(name)} {sql_type} {null_clause}')
    cols_sql = ",\n".join(cols)
    return (
        f"CREATE TABLE IF NOT EXISTS {quoted(table)} (\n"
        f"  id INTEGER PRIMARY KEY AUTOINCREMENT,\n"
        f"  batch_id TEXT NOT NULL,\n"
        f"  source TEXT NOT NULL DEFAULT 'upload',\n"
        f"  file_name TEXT,\n"
        f"  row_hash TEXT,\n"
        f"  inserted_at TEXT NOT NULL DEFAULT (datetime('now')),\n"
        f"{cols_sql}\n"
        f")"
    )


def _build_indexes(table: str) -> list[str]:
    return [
        f'CREATE INDEX IF NOT EXISTS "idx_{table}_date"      ON {quoted(table)}("Date")',
        f'CREATE INDEX IF NOT EXISTS "idx_{table}_party"     ON {quoted(table)}("Party Name")',
        f'CREATE INDEX IF NOT EXISTS "idx_{table}_batch"     ON {quoted(table)}(batch_id)',
        f'CREATE INDEX IF NOT EXISTS "idx_{table}_row_hash"  ON {quoted(table)}(row_hash)',
    ]


# Idempotent column additions for already-deployed sales/purchase tables.
_TABLE_HASH_ALTERS: tuple[str, ...] = ("row_hash",)


# is_mock_named: 1 when a row's "Product Name" was filled by the mock-
# backfill engine because the original upload row had a blank product
# name. Real product-name rows always have is_mock_named = 0.
def _table_mock_alter(table: str) -> str:
    return f'ALTER TABLE {quoted(table)} ADD COLUMN is_mock_named INTEGER NOT NULL DEFAULT 0'


# Quantity column — synthesized when a row lands without it, so unit-
# velocity / profit-per-product analytics work even for legacy uploads.
def _table_quantity_alter(table: str) -> str:
    return f'ALTER TABLE {quoted(table)} ADD COLUMN "Quantity" REAL'


# is_mock_quantity: 1 when Quantity was backfilled (vs supplied at upload).
def _table_qty_flag_alter(table: str) -> str:
    return f'ALTER TABLE {quoted(table)} ADD COLUMN is_mock_quantity INTEGER NOT NULL DEFAULT 0'


_UPLOADS_DDL = """
CREATE TABLE IF NOT EXISTS uploads (
    batch_id      TEXT PRIMARY KEY,
    filename      TEXT NOT NULL,
    target        TEXT NOT NULL,
    rows_inserted INTEGER NOT NULL DEFAULT 0,
    rows_failed   INTEGER NOT NULL DEFAULT 0,
    source        TEXT NOT NULL DEFAULT 'upload',
    status        TEXT NOT NULL DEFAULT 'active',  -- active | error | removed
    min_date      TEXT,
    max_date      TEXT,
    error_message TEXT,
    file_path     TEXT,                            -- absolute path to persisted source file
    uploaded_at   TEXT NOT NULL DEFAULT (datetime('now'))
)
"""

_UPLOADS_ALTERS: list[str] = [
    "ALTER TABLE uploads ADD COLUMN status        TEXT NOT NULL DEFAULT 'active'",
    "ALTER TABLE uploads ADD COLUMN min_date      TEXT",
    "ALTER TABLE uploads ADD COLUMN max_date      TEXT",
    "ALTER TABLE uploads ADD COLUMN error_message TEXT",
    "ALTER TABLE uploads ADD COLUMN file_path     TEXT",
    "ALTER TABLE uploads ADD COLUMN file_hash     TEXT",
    "ALTER TABLE uploads ADD COLUMN file_bytes    INTEGER",
    "ALTER TABLE uploads ADD COLUMN dedup_mode    TEXT",
    "ALTER TABLE uploads ADD COLUMN rows_skipped_duplicate INTEGER NOT NULL DEFAULT 0",
    "ALTER TABLE uploads ADD COLUMN rows_replaced INTEGER NOT NULL DEFAULT 0",
]

_UPLOADS_INDEXES: tuple[str, ...] = (
    'CREATE INDEX IF NOT EXISTS "idx_uploads_file_hash" ON uploads(file_hash, status)',
)


# ---------------------------------------------------------------------------
# Chat sessions (read-only history). Timestamps are stored as TEXT ISO
# strings in BOTH engines (passed from Python) to avoid datetime('now') vs
# NOW() / TIMESTAMPTZ divergence. The schema is all TEXT/INTEGER, so the
# SAME CREATE TABLE / CREATE INDEX strings are applied in the SQLite branch
# AND in _init_database_postgres().
# ---------------------------------------------------------------------------

_CONVERSATIONS_DDL = """
CREATE TABLE IF NOT EXISTS conversations (
    id            TEXT PRIMARY KEY,
    title         TEXT NOT NULL DEFAULT '',
    tenant_id     TEXT,                              -- nullable; multi-user readiness, unused now
    created_at    TEXT NOT NULL,
    updated_at    TEXT NOT NULL,
    message_count INTEGER NOT NULL DEFAULT 0
)
"""

_CONVERSATION_MESSAGES_DDL = """
CREATE TABLE IF NOT EXISTS conversation_messages (
    id              TEXT PRIMARY KEY,
    conversation_id TEXT NOT NULL,
    role            TEXT NOT NULL,                   -- 'user' | 'assistant'
    content         TEXT NOT NULL DEFAULT '',
    chart_json      TEXT,                            -- json.dumps(chart) or NULL
    created_at      TEXT NOT NULL
)
"""

_CONVERSATIONS_INDEXES: tuple[str, ...] = (
    'CREATE INDEX IF NOT EXISTS idx_convo_messages_cid '
    'ON conversation_messages(conversation_id, created_at)',
    'CREATE INDEX IF NOT EXISTS idx_conversations_updated '
    'ON conversations(updated_at)',
)


# ---------------------------------------------------------------------------
# Identity / Auth foundation (multi-tenant slice 1). Per-user accounts plus a
# tenant per owner. Additive and OFF by default (the routes that consume these
# only enforce when AUTH_ENABLED=true). All-TEXT schema with Python-supplied
# timestamps, so the SAME CREATE TABLE / CREATE INDEX strings apply in the
# SQLite branch AND in _init_database_postgres() — exactly like the
# conversations tables above.
# ---------------------------------------------------------------------------

_USERS_DDL = """
CREATE TABLE IF NOT EXISTS users (
    id            TEXT PRIMARY KEY,
    email         TEXT NOT NULL UNIQUE,
    password_hash TEXT NOT NULL,
    created_at    TEXT NOT NULL,
    access_status TEXT NOT NULL DEFAULT 'allowed',
    admin_notes   TEXT,
    trial_ends_at TEXT
)
"""

# Additive columns for databases created before access-control existed (live
# prod). Adding `access_status` with DEFAULT 'allowed' backfills every EXISTING
# account to 'allowed' so nobody is locked out when enforcement later ships;
# new signups explicitly set 'pending'. (col_name, column_declaration).
_USERS_MIGRATION_COLUMNS: tuple[tuple[str, str], ...] = (
    ("access_status", "TEXT NOT NULL DEFAULT 'allowed'"),
    ("admin_notes", "TEXT"),
    # 7-day free trial (Phase 3). Set at signup for access_status='trial'
    # accounts; NULL for grandfathered/allowed/denied accounts.
    ("trial_ends_at", "TEXT"),
)


async def _apply_users_migrations(db) -> None:
    """Idempotently add the access-control columns to an existing users table.

    Engine-aware so it never bricks startup: on Postgres uses
    ``ADD COLUMN IF NOT EXISTS`` (no error, so it can't poison the init
    transaction); on SQLite uses a plain ``ADD COLUMN`` and swallows the
    "duplicate column" error that occurs when it's already applied.
    """
    from app.db_engine import is_postgres

    pg = is_postgres()
    for name, decl in _USERS_MIGRATION_COLUMNS:
        stmt = (
            f"ALTER TABLE users ADD COLUMN IF NOT EXISTS {name} {decl}"
            if pg
            else f"ALTER TABLE users ADD COLUMN {name} {decl}"
        )
        try:
            await db.execute(stmt)
        except Exception:  # noqa: BLE001 - already-applied / migrations must not block startup
            pass

_TENANTS_DDL = """
CREATE TABLE IF NOT EXISTS tenants (
    id             TEXT PRIMARY KEY,
    owner_user_id  TEXT NOT NULL,
    name           TEXT,
    created_at     TEXT NOT NULL
)
"""

_IDENTITY_INDEXES: tuple[str, ...] = (
    'CREATE INDEX IF NOT EXISTS idx_users_email ON users(email)',
)


# ---------------------------------------------------------------------------
# Per-tenant schema mapping (LLM/user-driven concept → column map). First-party
# table in `public`, one row per tenant, isolated by the tenant_id PRIMARY KEY
# (NOT by search_path) exactly like uploads/conversations. The active mapping
# is stored as JSON text; the schema is all-TEXT with a Python-supplied
# timestamp, so the SAME CREATE TABLE string applies in the SQLite branch AND
# in _init_database_postgres(). See app/schema_mapping/store.py.
# ---------------------------------------------------------------------------

_TENANT_SCHEMA_MAPS_DDL = """
CREATE TABLE IF NOT EXISTS tenant_schema_maps (
    tenant_id     TEXT PRIMARY KEY,
    mapping_json  TEXT,
    updated_at    TEXT
)
"""


def uploads_dir() -> Path:
    """Directory for source CSV/XLSX files.

    On Postgres (DATABASE_URL set) the data lives in PG; /data may not be
    writable on Render without a mounted disk, so fall back to /tmp which is
    always writable. Files under /tmp are ephemeral across restarts, but that
    is acceptable — ingest re-reads the file then stores rows in Postgres.

    On SQLite the file is the source of truth, so we keep it on the persistent
    path under financial_db_path.
    """
    import os
    if os.environ.get("DATABASE_URL"):
        p = Path("/tmp") / "agentic_uploads"
    else:
        p = Path(settings.financial_db_path).parent / "uploads"
    p.mkdir(parents=True, exist_ok=True)
    return p


# ---------------------------------------------------------------------------
# Product & Location hierarchy tables — populated from uploaded data only.
# Trees are stored adjacency-list style (parent_id pointing to same table).
# ---------------------------------------------------------------------------

_PRODUCT_HIERARCHY_DDL = """
CREATE TABLE IF NOT EXISTS product_hierarchy (
    id          TEXT PRIMARY KEY,
    level       TEXT NOT NULL,           -- 'business' | 'category' | 'subcategory' | 'product_type' | 'brand'
    parent_id   TEXT,                    -- NULL for root
    name        TEXT NOT NULL,
    slug        TEXT NOT NULL,
    created_at  TEXT NOT NULL DEFAULT (datetime('now')),
    UNIQUE(parent_id, slug)
)
"""

_PRODUCT_MASTER_DDL = """
CREATE TABLE IF NOT EXISTS product_master (
    id             TEXT PRIMARY KEY,
    product_name   TEXT NOT NULL UNIQUE,         -- canonical name as it appears in uploads
    business_id    TEXT,
    category_id    TEXT,
    subcategory_id TEXT,
    product_type_id TEXT,
    brand_id       TEXT,
    created_at     TEXT NOT NULL DEFAULT (datetime('now')),
    updated_at     TEXT NOT NULL DEFAULT (datetime('now'))
)
"""

_LOCATION_HIERARCHY_DDL = """
CREATE TABLE IF NOT EXISTS location_hierarchy (
    id          TEXT PRIMARY KEY,
    level       TEXT NOT NULL,           -- 'business' | 'city' | 'branch' | 'counter'
    parent_id   TEXT,
    name        TEXT NOT NULL,
    slug        TEXT NOT NULL,
    created_at  TEXT NOT NULL DEFAULT (datetime('now')),
    UNIQUE(parent_id, slug)
)
"""

_BRANCH_MASTER_DDL = """
CREATE TABLE IF NOT EXISTS branch_master (
    id          TEXT PRIMARY KEY,
    branch_name TEXT NOT NULL UNIQUE,
    city        TEXT,
    address     TEXT,
    is_default  INTEGER NOT NULL DEFAULT 0,      -- exactly one row should have is_default=1
    location_id TEXT,                            -- → location_hierarchy.id (the branch node)
    enabled     INTEGER NOT NULL DEFAULT 1,
    created_at  TEXT NOT NULL DEFAULT (datetime('now')),
    updated_at  TEXT NOT NULL DEFAULT (datetime('now'))
)
"""

_HIERARCHY_INDEXES: tuple[str, ...] = (
    'CREATE INDEX IF NOT EXISTS "idx_product_hierarchy_parent" ON product_hierarchy(parent_id)',
    'CREATE INDEX IF NOT EXISTS "idx_product_hierarchy_level"  ON product_hierarchy(level)',
    'CREATE INDEX IF NOT EXISTS "idx_product_master_name"      ON product_master(product_name)',
    'CREATE INDEX IF NOT EXISTS "idx_product_master_category"  ON product_master(category_id)',
    'CREATE INDEX IF NOT EXISTS "idx_location_hierarchy_parent" ON location_hierarchy(parent_id)',
    'CREATE INDEX IF NOT EXISTS "idx_branch_master_enabled"    ON branch_master(enabled)',
)


# ---------------------------------------------------------------------------
# Product Hierarchy v2 — 6-level synthetic enterprise hierarchy.
#
# Coexists with the original 5-level product_hierarchy table; all existing
# KPIs continue to use the original. v2 is a strictly additive enrichment
# layer for enterprise-style drilldown queries.
#
#   Need → Family → Class → Line → Type → Item (SKU)
# ---------------------------------------------------------------------------

_PRODUCT_HIERARCHY_V2_DDL = """
CREATE TABLE IF NOT EXISTS product_hierarchy_v2 (
    id          TEXT PRIMARY KEY,
    level       TEXT NOT NULL,       -- 'need' | 'family' | 'class' | 'line' | 'type'
    parent_id   TEXT,
    name        TEXT NOT NULL,
    slug        TEXT NOT NULL,
    code        TEXT,                -- short code like 'FW', 'SHOE', 'ATH'
    created_at  TEXT NOT NULL DEFAULT (datetime('now')),
    UNIQUE(parent_id, slug)
)
"""

# product_sku_master: maps a canonical Product Name (from sales/purchase
# uploads) to its full 6-level path + a stable SKU code. Joinable from
# any v2 KPI exactly like product_master joins for the v1 KPIs.
_PRODUCT_SKU_MASTER_DDL = """
CREATE TABLE IF NOT EXISTS product_sku_master (
    id              TEXT PRIMARY KEY,            -- 'sku-<hex12>'
    sku_code        TEXT NOT NULL UNIQUE,        -- 'SKU-FW-001' style
    product_name    TEXT NOT NULL UNIQUE,        -- matches sales."Product Name"
    need_id         TEXT,                        -- → product_hierarchy_v2.id (level=need)
    family_id       TEXT,
    class_id        TEXT,
    line_id         TEXT,
    type_id         TEXT,
    created_at      TEXT NOT NULL DEFAULT (datetime('now')),
    updated_at      TEXT NOT NULL DEFAULT (datetime('now'))
)
"""

_HIERARCHY_V2_INDEXES: tuple[str, ...] = (
    'CREATE INDEX IF NOT EXISTS "idx_product_hierarchy_v2_parent" ON product_hierarchy_v2(parent_id)',
    'CREATE INDEX IF NOT EXISTS "idx_product_hierarchy_v2_level"  ON product_hierarchy_v2(level)',
    'CREATE INDEX IF NOT EXISTS "idx_product_sku_master_name"     ON product_sku_master(product_name)',
    'CREATE INDEX IF NOT EXISTS "idx_product_sku_master_class"    ON product_sku_master(class_id)',
    'CREATE INDEX IF NOT EXISTS "idx_product_sku_master_family"   ON product_sku_master(family_id)',
)


# ---------------------------------------------------------------------------
# Enrichment layers — Inventory + Forecast.
#
# These tables are DERIVED from the real sales/purchase data:
#   • sku_inventory   — current stock + velocity + status per SKU
#   • sku_forecast    — 14-day forward projection per SKU
#
# Every row is regenerated by app/enrichment/* on startup and after every
# upload. The real transactional tables (sales, purchase) are never written
# or modified by enrichment — they are read-only inputs.
# ---------------------------------------------------------------------------

_SKU_INVENTORY_DDL = """
CREATE TABLE IF NOT EXISTS sku_inventory (
    id                  INTEGER PRIMARY KEY AUTOINCREMENT,
    sku_code            TEXT NOT NULL UNIQUE,
    product_name        TEXT NOT NULL,
    on_hand_qty         INTEGER NOT NULL DEFAULT 0,
    reorder_level       INTEGER NOT NULL DEFAULT 0,
    avg_daily_sales     REAL NOT NULL DEFAULT 0,
    avg_daily_revenue   REAL NOT NULL DEFAULT 0,
    days_of_cover       REAL,
    status              TEXT NOT NULL DEFAULT 'unknown',  -- ok | low | overstocked | dead | unknown
    last_refreshed_at   TEXT NOT NULL DEFAULT (datetime('now')),
    source              TEXT NOT NULL DEFAULT 'synthetic'
)
"""

_SKU_FORECAST_DDL = """
CREATE TABLE IF NOT EXISTS sku_forecast (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    sku_code        TEXT NOT NULL,
    forecast_date   TEXT NOT NULL,
    forecast_qty    REAL NOT NULL DEFAULT 0,
    forecast_revenue REAL NOT NULL DEFAULT 0,
    method          TEXT NOT NULL DEFAULT 'trailing_30',
    confidence      REAL,
    generated_at    TEXT NOT NULL DEFAULT (datetime('now')),
    UNIQUE(sku_code, forecast_date)
)
"""

_ENRICHMENT_INDEXES: tuple[str, ...] = (
    'CREATE INDEX IF NOT EXISTS "idx_sku_inventory_status" ON sku_inventory(status)',
    'CREATE INDEX IF NOT EXISTS "idx_sku_inventory_sku"    ON sku_inventory(sku_code)',
    'CREATE INDEX IF NOT EXISTS "idx_sku_forecast_sku"     ON sku_forecast(sku_code)',
    'CREATE INDEX IF NOT EXISTS "idx_sku_forecast_date"    ON sku_forecast(forecast_date)',
)


# Per-product unit cost — required for profit / margin / loss KPIs that
# operate at SKU granularity. Costs are deterministically generated from
# the average sale price * a class-aware margin band (e.g. ~55-65% of
# avg sale price for footwear). Source is tagged 'synthetic' or 'manual'
# so the user knows what's mock vs. user-supplied.
_PRODUCT_COST_MASTER_DDL = """
CREATE TABLE IF NOT EXISTS product_cost_master (
    id                 INTEGER PRIMARY KEY AUTOINCREMENT,
    product_name       TEXT NOT NULL UNIQUE,
    unit_cost          REAL NOT NULL,
    avg_sale_price     REAL NOT NULL,
    margin_pct         REAL NOT NULL,
    source             TEXT NOT NULL DEFAULT 'synthetic',
    last_refreshed_at  TEXT NOT NULL DEFAULT (datetime('now'))
)
"""

_PRODUCT_COST_INDEXES: tuple[str, ...] = (
    'CREATE INDEX IF NOT EXISTS "idx_product_cost_master_name" ON product_cost_master(product_name)',
)


# ---------------------------------------------------------------------------
# Central error log table.
#
# Every uncaught exception, validation failure, upload error, AI pipeline
# crash, and explicitly-reported frontend error lands here. The /errors API
# reads from this table; analytics rolls up counts by module + severity.
# ---------------------------------------------------------------------------

_ERROR_LOG_DDL = """
CREATE TABLE IF NOT EXISTS error_log (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    error_id        TEXT NOT NULL UNIQUE,
    tenant_id       TEXT NOT NULL DEFAULT 'public',  -- owning tenant; reads are scoped to it
    occurred_at     TEXT NOT NULL DEFAULT (datetime('now')),
    severity        TEXT NOT NULL,        -- critical | high | medium | low
    module          TEXT NOT NULL,        -- sales | upload | ai | kpi | dashboard | auth | database | system | frontend | ...
    error_type      TEXT NOT NULL,        -- ExceptionType or 'frontend' / 'validation' / etc.
    message         TEXT NOT NULL,
    endpoint        TEXT,                  -- HTTP path or null
    method          TEXT,                  -- HTTP verb or null
    user_facing     INTEGER NOT NULL DEFAULT 0,
    suggested_fix   TEXT,
    source          TEXT,                  -- file:line or logical source identifier
    stack_trace     TEXT,                  -- full traceback for developers
    request_payload TEXT,                  -- JSON snippet of request body / args
    context         TEXT,                  -- JSON of extra fields
    resolved        INTEGER NOT NULL DEFAULT 0,
    resolved_at     TEXT,
    resolved_note   TEXT
)
"""

_ERROR_LOG_INDEXES: tuple[str, ...] = (
    'CREATE INDEX IF NOT EXISTS "idx_error_log_occurred"   ON error_log(occurred_at DESC)',
    'CREATE INDEX IF NOT EXISTS "idx_error_log_module"     ON error_log(module, severity)',
    'CREATE INDEX IF NOT EXISTS "idx_error_log_resolved"   ON error_log(resolved, occurred_at DESC)',
    'CREATE INDEX IF NOT EXISTS "idx_error_log_type"       ON error_log(error_type)',
    'CREATE INDEX IF NOT EXISTS "idx_error_log_tenant"     ON error_log(tenant_id, occurred_at DESC)',
)


_UPLOADS_DDL_PG = """
CREATE TABLE IF NOT EXISTS uploads (
    batch_id      TEXT PRIMARY KEY,
    filename      TEXT NOT NULL,
    target        TEXT NOT NULL,
    rows_inserted INTEGER NOT NULL DEFAULT 0,
    rows_failed   INTEGER NOT NULL DEFAULT 0,
    source        TEXT NOT NULL DEFAULT 'upload',
    status        TEXT NOT NULL DEFAULT 'active',
    min_date      TEXT,
    max_date      TEXT,
    error_message TEXT,
    file_path     TEXT,
    file_hash     TEXT,
    file_bytes    BIGINT,
    dedup_mode    TEXT,
    rows_skipped_duplicate INTEGER NOT NULL DEFAULT 0,
    rows_replaced INTEGER NOT NULL DEFAULT 0,
    uploaded_at   TIMESTAMPTZ NOT NULL DEFAULT NOW()
)
"""

_COLUMN_PROFILE_DDL_PG = """
CREATE TABLE IF NOT EXISTS _column_profile (
    table_name      TEXT NOT NULL,
    column_name     TEXT NOT NULL,
    pct_non_null    DOUBLE PRECISION NOT NULL DEFAULT 0,
    pct_numeric     DOUBLE PRECISION NOT NULL DEFAULT 0,
    pct_date        DOUBLE PRECISION NOT NULL DEFAULT 0,
    distinct_count  BIGINT NOT NULL DEFAULT 0,
    min_val         TEXT,
    max_val         TEXT,
    profiled_at     TIMESTAMPTZ NOT NULL DEFAULT NOW(),
    PRIMARY KEY (table_name, column_name)
)
"""

_RELATIONSHIPS_DDL_PG = """
CREATE TABLE IF NOT EXISTS _relationships (
    from_table     TEXT NOT NULL,
    from_column    TEXT NOT NULL,
    to_table       TEXT NOT NULL,
    to_column      TEXT NOT NULL,
    overlap_pct    DOUBLE PRECISION NOT NULL DEFAULT 0,
    from_distinct  BIGINT NOT NULL DEFAULT 0,
    to_distinct    BIGINT NOT NULL DEFAULT 0,
    detected_at    TIMESTAMPTZ NOT NULL DEFAULT NOW(),
    PRIMARY KEY (from_table, from_column, to_table, to_column)
)
"""


_ERROR_LOG_DDL_PG = """
CREATE TABLE IF NOT EXISTS error_log (
    id              BIGSERIAL PRIMARY KEY,
    error_id        TEXT NOT NULL UNIQUE,
    tenant_id       TEXT NOT NULL DEFAULT 'public',
    occurred_at     TIMESTAMPTZ NOT NULL DEFAULT NOW(),
    severity        TEXT NOT NULL,
    module          TEXT NOT NULL,
    error_type      TEXT NOT NULL,
    message         TEXT NOT NULL,
    endpoint        TEXT,
    method          TEXT,
    user_facing     INTEGER NOT NULL DEFAULT 0,
    suggested_fix   TEXT,
    source          TEXT,
    stack_trace     TEXT,
    request_payload TEXT,
    context         TEXT,
    resolved        INTEGER NOT NULL DEFAULT 0,
    resolved_at     TIMESTAMPTZ,
    resolved_note   TEXT
)
"""


async def _init_database_postgres() -> None:
    """Postgres init — minimum schema only. Static sales/purchase, hierarchy,
    enrichment, and KPI tables are SQLite-only legacy paths that we don't
    port in the 5-hour migration window. The dynamic u_* tables are created
    on demand by dynamic_ingest."""
    from app.db_engine import pg_connection
    async with pg_connection() as db:
        await db.execute(_UPLOADS_DDL_PG)
        await db.execute(_ERROR_LOG_DDL_PG)
        await db.execute(_COLUMN_PROFILE_DDL_PG)
        await db.execute(_RELATIONSHIPS_DDL_PG)
        await db.execute(
            'CREATE INDEX IF NOT EXISTS idx_uploads_file_hash '
            'ON uploads(file_hash, status)'
        )
        await db.execute(
            'CREATE INDEX IF NOT EXISTS idx_error_log_occurred '
            'ON error_log(occurred_at DESC)'
        )
        # Chat sessions (read-only history) — identical TEXT/INTEGER schema
        # as SQLite, so the same DDL/index strings apply here.
        await db.execute(_CONVERSATIONS_DDL)
        await db.execute(_CONVERSATION_MESSAGES_DDL)
        for stmt in _CONVERSATIONS_INDEXES:
            await db.execute(stmt)
        # Identity / auth foundation — additive, all-TEXT, same DDL as SQLite.
        await db.execute(_USERS_DDL)
        await _apply_users_migrations(db)
        await db.execute(_TENANTS_DDL)
        for stmt in _IDENTITY_INDEXES:
            await db.execute(stmt)
        # Per-tenant schema mapping — first-party, all-TEXT, same DDL as SQLite.
        await db.execute(_TENANT_SCHEMA_MAPS_DDL)
        # Multi-tenant slice 2a — additive tenant_id scaffolding (PG-native,
        # idempotent). The 'public' sentinel is tenant_context.DEFAULT_TENANT_ID.
        # No RLS and no read-scoping in this step; existing rows backfilled so
        # nothing disappears from any list. SQLite branch is intentionally
        # untouched (it goes dormant once Postgres is required).
        await db.execute(
            "ALTER TABLE uploads "
            "ADD COLUMN IF NOT EXISTS tenant_id TEXT NOT NULL DEFAULT 'public'"
        )
        await db.execute(
            "ALTER TABLE conversation_messages "
            "ADD COLUMN IF NOT EXISTS tenant_id TEXT NOT NULL DEFAULT 'public'"
        )
        # conversations already has a nullable tenant_id from Slice 1:
        # backfill existing NULLs, then set the column DEFAULT.
        await db.execute(
            "UPDATE conversations SET tenant_id = 'public' "
            "WHERE tenant_id IS NULL"
        )
        await db.execute(
            "ALTER TABLE conversations ALTER COLUMN tenant_id SET DEFAULT 'public'"
        )
        await db.execute(
            "CREATE INDEX IF NOT EXISTS idx_conversations_tenant "
            "ON conversations(tenant_id, updated_at)"
        )
        # error_log tenant scoping — additive, idempotent. Existing rows
        # backfill to 'public' so the operator/global view still finds them,
        # while every per-tenant read filters WHERE tenant_id = <caller>.
        await db.execute(
            "ALTER TABLE error_log "
            "ADD COLUMN IF NOT EXISTS tenant_id TEXT NOT NULL DEFAULT 'public'"
        )
        await db.execute(
            "CREATE INDEX IF NOT EXISTS idx_error_log_tenant "
            "ON error_log(tenant_id, occurred_at DESC)"
        )
    log.info(
        "postgres DB initialized (minimum schema: uploads + error_log "
        "+ _column_profile + _relationships + conversations + users/tenants)"
    )


async def init_database() -> None:
    """Create / verify the financial DB. Idempotent."""
    from app.db_engine import is_postgres
    if is_postgres():
        await _init_database_postgres()
        return
    p = db_path()
    p.parent.mkdir(parents=True, exist_ok=True)
    async with aiosqlite.connect(p) as db:
        await db.execute("PRAGMA journal_mode=WAL")
        await db.execute("PRAGMA synchronous=NORMAL")
        for table in ALLOWED_TABLES:
            await db.execute(_build_create(table))
            # Forward-migration: add hash + SCHEMA_SPEC columns missing
            # from older DBs (must run BEFORE indexes that reference them).
            for col in _TABLE_HASH_ALTERS:
                try:
                    await db.execute(
                        f'ALTER TABLE {quoted(table)} ADD COLUMN {col} TEXT'
                    )
                except aiosqlite.OperationalError as e:
                    if "duplicate column" not in str(e).lower():
                        raise
            # is_mock_named flag for the product-name backfill engine.
            try:
                await db.execute(_table_mock_alter(table))
            except aiosqlite.OperationalError as e:
                if "duplicate column" not in str(e).lower():
                    raise
            # Quantity + is_mock_quantity flag for unit-velocity analytics.
            for alter in (_table_quantity_alter(table), _table_qty_flag_alter(table)):
                try:
                    await db.execute(alter)
                except aiosqlite.OperationalError as e:
                    if "duplicate column" not in str(e).lower():
                        raise
            for stmt in _build_indexes(table):
                await db.execute(stmt)
            for col_name, sql_type, required in SCHEMA_SPEC:
                null_clause = "NOT NULL" if required else "NULL"
                try:
                    await db.execute(
                        f'ALTER TABLE {quoted(table)} '
                        f'ADD COLUMN {quoted(col_name)} {sql_type} {null_clause}'
                    )
                except aiosqlite.OperationalError as e:
                    if "duplicate column" not in str(e).lower():
                        raise
            # Archive table — identical shape to the live table. Rows are
            # physically moved here when a dataset is archived, so live
            # analytics SELECTs (which target sales/purchase) never see
            # archived data without any KPI-template changes.
            archive_table = f"{table}_archive"
            await db.execute(_build_create(archive_table))
            # Mirror the same ALTERs onto the archive table so a future
            # archive_upload() can INSERT ... SELECT * without column-
            # count mismatches.
            for col in _TABLE_HASH_ALTERS:
                try:
                    await db.execute(
                        f'ALTER TABLE {quoted(archive_table)} ADD COLUMN {col} TEXT'
                    )
                except aiosqlite.OperationalError as e:
                    if "duplicate column" not in str(e).lower():
                        raise
            try:
                await db.execute(_table_mock_alter(archive_table))
            except aiosqlite.OperationalError as e:
                if "duplicate column" not in str(e).lower():
                    raise
            for alter in (_table_quantity_alter(archive_table),
                          _table_qty_flag_alter(archive_table)):
                try:
                    await db.execute(alter)
                except aiosqlite.OperationalError as e:
                    if "duplicate column" not in str(e).lower():
                        raise
        await db.execute(_UPLOADS_DDL)
        for stmt in _UPLOADS_ALTERS:
            try:
                await db.execute(stmt)
            except aiosqlite.OperationalError as e:
                if "duplicate column" not in str(e).lower():
                    raise
        for stmt in _UPLOADS_INDEXES:
            await db.execute(stmt)
        # Hierarchy tables — idempotent.
        await db.execute(_PRODUCT_HIERARCHY_DDL)
        await db.execute(_PRODUCT_MASTER_DDL)
        await db.execute(_LOCATION_HIERARCHY_DDL)
        await db.execute(_BRANCH_MASTER_DDL)
        for stmt in _HIERARCHY_INDEXES:
            await db.execute(stmt)
        # v2 enterprise hierarchy — additive, alongside the original.
        await db.execute(_PRODUCT_HIERARCHY_V2_DDL)
        await db.execute(_PRODUCT_SKU_MASTER_DDL)
        for stmt in _HIERARCHY_V2_INDEXES:
            await db.execute(stmt)
        # Enrichment layers — inventory + forecast, derived from real sales.
        await db.execute(_SKU_INVENTORY_DDL)
        await db.execute(_SKU_FORECAST_DDL)
        for stmt in _ENRICHMENT_INDEXES:
            await db.execute(stmt)
        # Product cost master (per-product unit cost, for profit/margin KPIs).
        await db.execute(_PRODUCT_COST_MASTER_DDL)
        for stmt in _PRODUCT_COST_INDEXES:
            await db.execute(stmt)
        # Error-log table.
        await db.execute(_ERROR_LOG_DDL)
        for stmt in _ERROR_LOG_INDEXES:
            await db.execute(stmt)
        # Chat sessions (read-only history) — idempotent.
        await db.execute(_CONVERSATIONS_DDL)
        await db.execute(_CONVERSATION_MESSAGES_DDL)
        for stmt in _CONVERSATIONS_INDEXES:
            await db.execute(stmt)
        # Identity / auth foundation (multi-tenant slice 1) — idempotent.
        await db.execute(_USERS_DDL)
        await _apply_users_migrations(db)
        await db.execute(_TENANTS_DDL)
        for stmt in _IDENTITY_INDEXES:
            await db.execute(stmt)
        # Per-tenant schema mapping (LLM/user concept → column map) — idempotent.
        await db.execute(_TENANT_SCHEMA_MAPS_DDL)
        await db.commit()
    log.info("financial DB initialized at %s", p)


@asynccontextmanager
async def get_connection() -> AsyncIterator[Any]:
    """Yields an aiosqlite.Connection, OR (when DATABASE_URL is set to a
    postgres URL) an aiosqlite-compatible shim around asyncpg.

    The shim exposes the same `execute / fetchall / commit` surface so
    every existing call site keeps working without rewrites. See
    app/db_engine.py for the SQL translation that bridges the dialects."""
    # Import here so a missing asyncpg dep doesn't break SQLite-only deploys.
    from app.db_engine import is_postgres, pg_connection
    if is_postgres():
        async with pg_connection() as db:
            yield db
        return
    async with aiosqlite.connect(db_path()) as db:
        db.row_factory = aiosqlite.Row
        yield db


async def fetch_all(sql: str, params: Iterable[Any] = ()) -> list[dict[str, Any]]:
    async with get_connection() as db:
        cur = await db.execute(sql, tuple(params))
        rows = await cur.fetchall()
        await cur.close()
        return [dict(r) for r in rows]


async def fetch_one(sql: str, params: Iterable[Any] = ()) -> dict[str, Any] | None:
    rows = await fetch_all(sql, params)
    return rows[0] if rows else None


async def count_rows(table: str) -> int:
    # Accept the live tables AND their archive counterparts. Everything
    # else returns 0 — keeps count_rows safe from arbitrary table reads.
    allowed = set(ALLOWED_TABLES) | {f"{t}_archive" for t in ALLOWED_TABLES}
    if table not in allowed:
        return 0
    # Static sales/purchase tables don't exist on Postgres (SQLite-only
    # legacy path). Return 0 cleanly instead of raising UndefinedTableError.
    from app.db_engine import is_postgres
    if is_postgres():
        return 0
    row = await fetch_one(f"SELECT COUNT(*) AS n FROM {quoted(table)}")
    return int(row["n"]) if row else 0


def insert_rows(
    table: str,
    rows: list[dict[str, Any]],
    *,
    batch_id: str,
    source: str = "upload",
    file_name: str | None = None,
    row_hashes: list[str] | None = None,
) -> int:
    """Synchronous bulk insert wrapped in one transaction. Returns rows inserted.

    Run this inside `asyncio.to_thread` from async callers — it uses the
    plain sqlite3 driver because executemany is fastest there.

    `row_hashes` parallels `rows`. When None, row_hash is NULL — older
    callers keep working but lose duplicate-detection benefits on those rows.

    On Postgres (DATABASE_URL set), this becomes a no-op: the static
    sales/purchase tables are SQLite-only legacy paths that weren't ported
    in the 5-hour Supabase migration window. XLSX uploads still work via
    dynamic_ingest's u_* tables.
    """
    from app.db_engine import is_postgres
    if is_postgres():
        log.info(
            "insert_rows skipped on Postgres (table=%s, rows=%d) — "
            "use XLSX upload for u_* dynamic tables instead",
            table, len(rows),
        )
        return 0
    if table not in ALLOWED_TABLES:
        raise ValueError(f"unknown table: {table!r}")
    if not rows:
        return 0
    if row_hashes is not None and len(row_hashes) != len(rows):
        raise ValueError(
            f"row_hashes length ({len(row_hashes)}) != rows length ({len(rows)})"
        )

    cols = ["batch_id", "source", "file_name", "row_hash", *SCHEMA_COLUMNS]
    placeholders = ",".join(["?"] * len(cols))
    col_sql = ",".join(quoted(c) for c in cols)
    sql = f"INSERT INTO {quoted(table)} ({col_sql}) VALUES ({placeholders})"

    payload: list[tuple[Any, ...]] = []
    for i, r in enumerate(rows):
        rh = row_hashes[i] if row_hashes is not None else None
        payload.append((
            batch_id,
            source,
            file_name,
            rh,
            *(r.get(c) for c in SCHEMA_COLUMNS),
        ))

    conn = sqlite3.connect(str(db_path()), isolation_level=None)
    try:
        conn.execute("PRAGMA journal_mode=WAL")
        conn.execute("PRAGMA synchronous=NORMAL")
        conn.execute("BEGIN")
        conn.executemany(sql, payload)
        conn.execute("COMMIT")
        return len(payload)
    except Exception:
        log.exception("insert_rows failed")
        try:
            conn.execute("ROLLBACK")
        except Exception:
            pass
        return 0
    finally:
        conn.close()


# ===========================================================================
# 5. UPLOAD REGISTRY — uploads-table CRUD
# ===========================================================================

_UPLOAD_META_COLS = (
    "batch_id, filename, target, rows_inserted, rows_failed, "
    "source, status, min_date, max_date, error_message, file_path, "
    "file_hash, file_bytes, dedup_mode, rows_skipped_duplicate, rows_replaced, uploaded_at"
)


async def record_upload_meta(
    batch_id: str,
    filename: str,
    target: str,
    rows_inserted: int,
    rows_failed: int,
    *,
    tenant_id: str = "public",
    source: str = "upload",
    status: str = "active",
    min_date: str | None = None,
    max_date: str | None = None,
    error_message: str | None = None,
    file_path: str | None = None,
    file_hash: str | None = None,
    file_bytes: int | None = None,
    dedup_mode: str | None = None,
    rows_skipped_duplicate: int = 0,
    rows_replaced: int = 0,
) -> None:
    if status not in ("active", "archived", "error", "removed"):
        raise ValueError(f"unknown upload status: {status!r}")
    # ON CONFLICT syntax works on SQLite >=3.24 and Postgres natively.
    # Replaces the prior `INSERT OR REPLACE` which was SQLite-only.
    async with get_connection() as db:
        await db.execute(
            """INSERT INTO public.uploads
               (batch_id, tenant_id, filename, target, rows_inserted, rows_failed,
                source, status, min_date, max_date, error_message, file_path,
                file_hash, file_bytes, dedup_mode,
                rows_skipped_duplicate, rows_replaced)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
               ON CONFLICT (batch_id) DO UPDATE SET
                 tenant_id = EXCLUDED.tenant_id,
                 filename = EXCLUDED.filename,
                 target = EXCLUDED.target,
                 rows_inserted = EXCLUDED.rows_inserted,
                 rows_failed = EXCLUDED.rows_failed,
                 source = EXCLUDED.source,
                 status = EXCLUDED.status,
                 min_date = EXCLUDED.min_date,
                 max_date = EXCLUDED.max_date,
                 error_message = EXCLUDED.error_message,
                 file_path = EXCLUDED.file_path,
                 file_hash = EXCLUDED.file_hash,
                 file_bytes = EXCLUDED.file_bytes,
                 dedup_mode = EXCLUDED.dedup_mode,
                 rows_skipped_duplicate = EXCLUDED.rows_skipped_duplicate,
                 rows_replaced = EXCLUDED.rows_replaced""",
            (batch_id, tenant_id, filename, target, rows_inserted, rows_failed,
             source, status, min_date, max_date, error_message, file_path,
             file_hash, file_bytes, dedup_mode,
             rows_skipped_duplicate, rows_replaced),
        )
        await db.commit()


async def list_uploads_meta(
    limit: int = 200, *, tenant_id: str = "public"
) -> list[dict[str, Any]]:
    """List uploads for one tenant (default 'public'). The uploads table
    lives in `public` and isolates via `WHERE tenant_id = ?` (slice 2a
    stamps tenant_id on every write)."""
    return await fetch_all(
        f"SELECT {_UPLOAD_META_COLS} FROM public.uploads "
        "WHERE tenant_id = ? "
        "ORDER BY uploaded_at DESC LIMIT ?",
        (tenant_id, limit),
    )


async def get_upload_meta(batch_id: str, tenant_id: str | None = None) -> dict[str, Any] | None:
    if tenant_id is not None:
        return await fetch_one(
            f"SELECT {_UPLOAD_META_COLS} FROM public.uploads WHERE batch_id = ? AND tenant_id = ?",
            (batch_id, tenant_id),
        )
    return await fetch_one(
        f"SELECT {_UPLOAD_META_COLS} FROM public.uploads WHERE batch_id = ?",
        (batch_id,),
    )


async def find_active_upload_by_file_hash(file_hash: str, tenant_id: str | None = None) -> dict[str, Any] | None:
    """Return the metadata of an existing active upload whose source file
    hashes to the same SHA256. None if no match."""
    if not file_hash:
        return None
    if tenant_id is not None:
        return await fetch_one(
            f"SELECT {_UPLOAD_META_COLS} FROM public.uploads "
            "WHERE file_hash = ? AND status = 'active' AND tenant_id = ? "
            "ORDER BY uploaded_at DESC LIMIT 1",
            (file_hash, tenant_id),
        )
    return await fetch_one(
        f"SELECT {_UPLOAD_META_COLS} FROM public.uploads "
        "WHERE file_hash = ? AND status = 'active' "
        "ORDER BY uploaded_at DESC LIMIT 1",
        (file_hash,),
    )


async def archive_upload(batch_id: str, tenant_id: str | None = None) -> dict[str, Any]:
    """Soft-deactivate a dataset: move its rows from the live table to the
    archive table and flip status='archived'. The source file is kept.
    Analytics SELECTs target the live table only, so the AI immediately
    stops using archived data — no KPI template changes required.

    Idempotent: archiving an already-archived batch is a no-op success.
    """
    meta = await get_upload_meta(batch_id, tenant_id=tenant_id)
    if meta is None:
        raise ValueError(f"unknown batch_id: {batch_id}")
    target = meta["target"]
    if target not in ALLOWED_TABLES:
        raise ValueError(f"upload metadata has invalid target: {target!r}")
    status = meta.get("status")
    if status == "archived":
        return {
            "batch_id": batch_id, "rows_moved": 0,
            "table": target, "status": "archived", "already_archived": True,
        }
    if status == "removed":
        raise ValueError(f"cannot archive a removed batch: {batch_id}")
    if status != "active":
        raise ValueError(f"cannot archive batch in status {status!r}")

    live_t = quoted(target)
    arch_t = quoted(target + "_archive")
    rows_moved = 0
    async with get_connection() as db:
        # Move rows: INSERT then DELETE in one transaction. SQLite supports
        # INSERT ... SELECT with column lists; we use SELECT * since the
        # archive schema is identical (built from the same _build_create).
        # We list columns explicitly so a future schema drift produces a
        # clear error rather than a silent mismatch.
        cur = await db.execute(
            f'SELECT COUNT(*) AS n FROM {live_t} WHERE batch_id = ?', (batch_id,)
        )
        row = await cur.fetchone()
        await cur.close()
        rows_moved = int(dict(row).get("n") or 0) if row else 0

        if rows_moved > 0:
            await db.execute(
                f'INSERT INTO {arch_t} SELECT * FROM {live_t} WHERE batch_id = ?',
                (batch_id,),
            )
            await db.execute(
                f'DELETE FROM {live_t} WHERE batch_id = ?', (batch_id,),
            )
        await db.execute(
            "UPDATE public.uploads SET status='archived' WHERE batch_id = ?",
            (batch_id,),
        )
        await db.commit()
    log.info("archive_upload: batch=%s table=%s rows_moved=%d", batch_id, target, rows_moved)
    return {
        "batch_id": batch_id, "rows_moved": rows_moved,
        "table": target, "status": "archived", "already_archived": False,
    }


async def unarchive_upload(batch_id: str, tenant_id: str | None = None) -> dict[str, Any]:
    """Reverse of archive_upload: move rows back from the archive table to
    the live table and flip status='active'."""
    meta = await get_upload_meta(batch_id, tenant_id=tenant_id)
    if meta is None:
        raise ValueError(f"unknown batch_id: {batch_id}")
    target = meta["target"]
    if target not in ALLOWED_TABLES:
        raise ValueError(f"upload metadata has invalid target: {target!r}")
    status = meta.get("status")
    if status == "active":
        return {
            "batch_id": batch_id, "rows_moved": 0,
            "table": target, "status": "active", "already_active": True,
        }
    if status != "archived":
        raise ValueError(f"cannot unarchive batch in status {status!r}")

    live_t = quoted(target)
    arch_t = quoted(target + "_archive")
    rows_moved = 0
    async with get_connection() as db:
        cur = await db.execute(
            f'SELECT COUNT(*) AS n FROM {arch_t} WHERE batch_id = ?', (batch_id,)
        )
        row = await cur.fetchone()
        await cur.close()
        rows_moved = int(dict(row).get("n") or 0) if row else 0
        if rows_moved > 0:
            await db.execute(
                f'INSERT INTO {live_t} SELECT * FROM {arch_t} WHERE batch_id = ?',
                (batch_id,),
            )
            await db.execute(
                f'DELETE FROM {arch_t} WHERE batch_id = ?', (batch_id,),
            )
        await db.execute(
            "UPDATE public.uploads SET status='active' WHERE batch_id = ?",
            (batch_id,),
        )
        await db.commit()
    log.info("unarchive_upload: batch=%s table=%s rows_moved=%d", batch_id, target, rows_moved)
    return {
        "batch_id": batch_id, "rows_moved": rows_moved,
        "table": target, "status": "active", "already_active": False,
    }


async def _disconnect_workbook(batch_id: str, meta: dict[str, Any]) -> dict[str, Any]:
    """Disconnect a workbook dataset. Its rows live in dynamic u_* tables
    tagged with a _batch_id column (not in sales/purchase). Delete every
    row for this batch so the Coordinator no longer sees the data, then
    mark the upload 'removed'."""
    from app.db_engine import is_postgres as _is_pg
    rows_removed = 0
    async with get_connection() as db:
        if _is_pg():
            cur = await db.execute(
                "SELECT table_name AS name FROM information_schema.tables "
                "WHERE table_schema = current_schema() AND table_name LIKE 'u\\_%' ESCAPE '\\'"
            )
        else:
            cur = await db.execute(
                "SELECT name FROM sqlite_master WHERE type='table' "
                "AND name LIKE 'u\\_%' ESCAPE '\\'"
            )
        tables = [dict(r).get("name") or "" for r in await cur.fetchall()]
        await cur.close()
        for t in tables:
            if not t:
                continue
            if _is_pg():
                cur = await db.execute(
                    "SELECT column_name AS name FROM information_schema.columns "
                    "WHERE table_schema = current_schema() AND table_name=?",
                    (t,),
                )
            else:
                cur = await db.execute(f"PRAGMA table_info({quoted(t)})")
            cols = [dict(c).get("name") for c in await cur.fetchall()]
            await cur.close()
            if "_batch_id" not in cols:
                continue
            cur = await db.execute(
                f'DELETE FROM {quoted(t)} WHERE _batch_id = ?', (batch_id,)
            )
            rows_removed += getattr(cur, "rowcount", 0) or 0
            await cur.close()
        await db.execute(
            "UPDATE public.uploads SET status='removed' WHERE batch_id = ?",
            (batch_id,),
        )
        await db.commit()
    file_removed = False
    fp = meta.get("file_path")
    if fp:
        try:
            p = Path(fp)
            if p.exists() and p.is_file():
                p.unlink()
                file_removed = True
        except Exception:
            log.warning("failed to delete upload file: %s", fp, exc_info=True)
    log.info(
        "disconnect_upload (workbook): batch=%s rows_removed=%d file_removed=%s",
        batch_id, rows_removed, file_removed,
    )
    return {
        "batch_id": batch_id,
        "rows_removed": int(rows_removed),
        "file_removed": file_removed,
        "table": meta.get("target") or "(workbook)",
        "already_removed": False,
        "status": "removed",
    }


async def disconnect_upload(batch_id: str, tenant_id: str | None = None) -> dict[str, Any]:
    """Remove a dataset permanently. Deletes:
      • All rows in sales/purchase AND sales_archive/purchase_archive
        tagged with this batch_id (handles archived datasets correctly),
        OR — for workbook datasets — every u_* row tagged with the batch.
      • The persisted source file on disk (if any).
    Marks the uploads row 'removed' (kept for audit history).

    The user must explicitly call this — nothing else removes a dataset.
    """
    meta = await get_upload_meta(batch_id, tenant_id=tenant_id)
    if meta is None:
        raise ValueError(f"unknown batch_id: {batch_id}")
    target = meta["target"]
    if meta["status"] == "removed":
        return {
            "batch_id": batch_id,
            "rows_removed": 0,
            "file_removed": False,
            "table": target,
            "already_removed": True,
            "status": "removed",
        }
    # Workbook datasets live in dynamic u_* tables, not sales/purchase.
    if target not in ALLOWED_TABLES:
        return await _disconnect_workbook(batch_id, meta)
    async with get_connection() as db:
        cur = await db.execute(
            f'DELETE FROM {quoted(target)} WHERE batch_id = ?',
            (batch_id,),
        )
        rows_removed = cur.rowcount or 0
        await cur.close()
        # Also remove from the archive table — handles datasets that were
        # archived before being deleted.
        cur = await db.execute(
            f'DELETE FROM {quoted(target + "_archive")} WHERE batch_id = ?',
            (batch_id,),
        )
        rows_removed += (cur.rowcount or 0)
        await cur.close()
        await db.execute(
            "UPDATE public.uploads SET status='removed' WHERE batch_id = ?",
            (batch_id,),
        )
        await db.commit()
    # Also delete the persisted source file, if it exists. Failure to delete
    # is logged but doesn't break the request — the database state is what
    # matters for analytics correctness.
    file_removed = False
    fp = meta.get("file_path")
    if fp:
        try:
            p = Path(fp)
            if p.exists() and p.is_file():
                p.unlink()
                file_removed = True
        except Exception:
            log.warning("failed to delete persisted upload file: %s", fp, exc_info=True)
    log.info(
        "disconnect_upload: batch=%s table=%s rows_removed=%d file_removed=%s",
        batch_id, target, rows_removed, file_removed,
    )
    return {
        "batch_id": batch_id,
        "rows_removed": int(rows_removed),
        "file_removed": file_removed,
        "table": target,
        "already_removed": False,
        "status": "removed",
    }


# ===========================================================================
# 6. HEADER DETECTION — alias index + best-row picker
# ===========================================================================

def normalize_key(s: Any) -> str:
    """Lowercase, replace any non-alphanumeric char with space, collapse whitespace."""
    raw = str(s if s is not None else "").lower()
    cleaned_chars = [
        ch if ch.isalnum() or ch.isspace() else " "
        for ch in raw
    ]
    return " ".join("".join(cleaned_chars).split())


_ALIAS_INDEX: dict[str, str] = {}
for _canonical, _aliases in HEADER_ALIASES.items():
    _ALIAS_INDEX[normalize_key(_canonical)] = _canonical
    for _alias in _aliases:
        _ALIAS_INDEX[normalize_key(_alias)] = _canonical


def alias_lookup(raw: Any) -> str | None:
    if raw is None:
        return None
    return _ALIAS_INDEX.get(normalize_key(raw))


def score_row_as_header(
    cells: list[Any],
) -> tuple[int, dict[str, str], list[str]]:
    seen_canonical: set[str] = set()
    header_index: dict[str, str] = {}
    for raw in cells:
        if raw is None:
            continue
        s = str(raw).strip()
        if not s:
            continue
        canonical = _ALIAS_INDEX.get(normalize_key(s))
        if canonical is None:
            continue
        if canonical in seen_canonical:
            continue
        header_index[s] = canonical
        seen_canonical.add(canonical)

    required_matched = sum(1 for c in REQUIRED_COLUMNS if c in seen_canonical)
    optional_matched = len(seen_canonical) - required_matched
    score = required_matched * 100 + optional_matched
    missing = [c for c in REQUIRED_COLUMNS if c not in seen_canonical]
    return score, header_index, missing


def find_header_row(
    rows_buffer: list[list[Any]],
) -> tuple[int, list[Any], dict[str, str]]:
    candidates: list[tuple[int, int, list[Any], dict[str, str]]] = []
    for idx, row in enumerate(rows_buffer):
        if row is None:
            continue
        score, header_index, missing = score_row_as_header(list(row))
        if missing:
            continue
        candidates.append((score, idx, list(row), header_index))

    if not candidates:
        sample = []
        for i, row in enumerate(rows_buffer[:5]):
            preview = [str(c)[:30] for c in (row or [])][:8]
            sample.append(f"row {i + 1}: {preview}")
        raise ValueError(
            f"No row in the first {len(rows_buffer)} contained all required "
            f"columns ({REQUIRED_COLUMNS}). Sample: {' | '.join(sample)}"
        )

    candidates.sort(key=lambda x: (-x[0], x[1]))
    score, idx, header_cells, header_index = candidates[0]
    return idx, header_cells, header_index


def map_headers_strict(
    header: list[str],
) -> tuple[dict[str, str], list[str], list[str]]:
    seen_canonical: set[str] = set()
    header_index: dict[str, str] = {}
    unmatched: list[str] = []
    for raw in header:
        if not raw:
            continue
        canonical = _ALIAS_INDEX.get(normalize_key(raw))
        if canonical is None:
            unmatched.append(raw)
            continue
        if canonical in seen_canonical:
            unmatched.append(raw)
            continue
        header_index[raw] = canonical
        seen_canonical.add(canonical)
    missing = [c for c in REQUIRED_COLUMNS if c not in seen_canonical]
    return header_index, missing, unmatched


# ===========================================================================
# 7. UPLOAD PARSERS — CSV + XLSX with auto header detection
# ===========================================================================

MAX_HEADER_SCAN = 15


class UploadError(ValueError):
    """File-level failure: empty / unsupported / unreadable / no header."""


def _row_to_record(keys: list[str], row: Any) -> dict[str, Any]:
    if row is None:
        return {}
    out: dict[str, Any] = {}
    for i, k in enumerate(keys):
        if not k:
            continue
        v = row[i] if i < len(row) else None
        out[k] = v
    return out


def parse_csv_bytes(content: bytes) -> tuple[list[str], dict[str, str], list[dict[str, Any]]]:
    text = content.decode("utf-8-sig", errors="replace")
    rows = list(csv.reader(io.StringIO(text)))
    if not rows:
        raise UploadError("CSV is empty")
    return _materialize(rows)


def parse_xlsx_bytes(content: bytes) -> tuple[list[str], dict[str, str], list[dict[str, Any]]]:
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        raise UploadError(f"Could not open xlsx: {e}")
    try:
        ws = _select_xlsx_sheet(wb)[1]
        rows = [list(r) if r is not None else [] for r in ws.iter_rows(values_only=True)]
        if not rows:
            raise UploadError("Selected sheet is empty")
        return _materialize(rows)
    finally:
        try:
            wb.close()
        except Exception:
            pass


def _materialize(rows: list[list[Any]]) -> tuple[list[str], dict[str, str], list[dict[str, Any]]]:
    buffer = rows[:MAX_HEADER_SCAN]
    try:
        header_idx, header_cells, header_index = find_header_row(buffer)
    except ValueError as e:
        raise UploadError(str(e)) from e
    header = [str(c).strip() if c is not None else "" for c in header_cells]
    data = []
    for row in rows[header_idx + 1:]:
        rec = _row_to_record(header, row)
        if rec:
            data.append(rec)
    return header, header_index, data


def stream_parse_csv_with_detection(
    path: Path,
) -> tuple[list[str], dict[str, str], Iterator[dict[str, Any]]]:
    f = open(path, "r", encoding="utf-8-sig", errors="replace", newline="")
    try:
        reader = csv.reader(f)
        buffer: list[list[Any]] = []
        for _ in range(MAX_HEADER_SCAN):
            try:
                buffer.append(next(reader))
            except StopIteration:
                break
        if not buffer:
            f.close()
            raise UploadError("CSV is empty")
        try:
            header_idx, header_cells, header_index = find_header_row(buffer)
        except ValueError as e:
            f.close()
            raise UploadError(str(e)) from e
        header = [str(c).strip() if c is not None else "" for c in header_cells]
        data_buffer_tail = list(buffer[header_idx + 1:])
    except Exception:
        f.close()
        raise

    def _gen() -> Iterator[dict[str, Any]]:
        try:
            for row in data_buffer_tail:
                rec = _row_to_record(header, row)
                if rec:
                    yield rec
            for row in reader:
                rec = _row_to_record(header, row)
                if rec:
                    yield rec
        finally:
            try:
                f.close()
            except Exception:
                pass

    return header, header_index, _gen()


def _select_xlsx_sheet(wb) -> tuple[str, Any, int, list[Any], dict[str, str]]:
    def _scan_sheet(name: str) -> tuple[int, int, list[Any], dict[str, str]] | None:
        ws_local = wb[name]
        buffer: list[list[Any]] = []
        for i, row in enumerate(ws_local.iter_rows(values_only=True)):
            if i >= MAX_HEADER_SCAN:
                break
            buffer.append(list(row) if row is not None else [])
        if not buffer:
            return None
        try:
            idx, cells, hi = find_header_row(buffer)
        except ValueError:
            return None
        score = len(hi)
        return score, idx, cells, hi

    active_name = wb.active.title if wb.active is not None else None
    if active_name:
        active_result = _scan_sheet(active_name)
        if active_result is not None:
            score, idx, cells, hi = active_result
            log.info("xlsx: using active sheet %r (score=%d, header_row=%d)",
                     active_name, score, idx + 1)
            return active_name, wb[active_name], idx, cells, hi

    candidates: list[tuple[int, str, int, list[Any], dict[str, str]]] = []
    for name in wb.sheetnames:
        if name == active_name:
            continue
        result = _scan_sheet(name)
        if result is None:
            continue
        score, idx, cells, hi = result
        candidates.append((score, name, idx, cells, hi))

    if not candidates:
        raise UploadError(
            f"No sheet contained a valid header row in first {MAX_HEADER_SCAN} rows. "
            f"Sheets tried: {wb.sheetnames}"
        )
    candidates.sort(key=lambda x: -x[0])
    score, name, idx, cells, hi = candidates[0]
    log.info("xlsx: using sheet %r (score=%d, header_row=%d)", name, score, idx + 1)
    return name, wb[name], idx, cells, hi


def stream_parse_xlsx_with_detection(
    path: Path,
) -> tuple[list[str], dict[str, str], Iterator[dict[str, Any]], str]:
    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    try:
        sheet_name, ws, header_idx, header_cells, header_index = _select_xlsx_sheet(wb)
        header = [str(c).strip() if c is not None else "" for c in header_cells]
    except Exception:
        try:
            wb.close()
        except Exception:
            pass
        raise

    def _gen() -> Iterator[dict[str, Any]]:
        try:
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i <= header_idx:
                    continue
                row_list = list(row) if row is not None else []
                rec = _row_to_record(header, row_list)
                if rec:
                    yield rec
        finally:
            try:
                wb.close()
            except Exception:
                pass

    return header, header_index, _gen(), sheet_name


def parse_file(filename: str, content: bytes):
    name = (filename or "").lower()
    if name.endswith(".csv"):
        return parse_csv_bytes(content)
    if name.endswith(".xlsx"):
        return parse_xlsx_bytes(content)
    raise UploadError(f"Unsupported file type: {filename!r} (use .csv or .xlsx)")


# ===========================================================================
# 8. RESPONSE CACHE — local JSON file, self-invalidating via data_version
# ===========================================================================

_CACHE_LOCK = Lock()
_cache_log = logging.getLogger("agentic_ai.cache")

_DATA_VERSION_LOCK = Lock()


def _data_version_path() -> Path:
    # On Postgres (DATABASE_URL set) the version sidecar must NOT live next to
    # financial_db_path: Render points FINANCIAL_DB_PATH at /data, which has no
    # writable disk on the free plan, so bump_data_version() would raise
    # PermissionError('/data') and 500 an otherwise-successful upload. Mirror
    # uploads_dir() / _cache_path() and fall back to /tmp (always writable;
    # ephemeral is fine — data_version is just a monotonic cache-busting
    # counter). SQLite / local keeps the persistent path beside the DB file.
    import os
    if os.environ.get("DATABASE_URL"):
        return Path("/tmp") / "agentic_data.version"
    return Path(settings.financial_db_path).with_suffix(".version")


def get_data_version() -> int:
    p = _data_version_path()
    if not p.exists():
        return 0
    try:
        return int(p.read_text(encoding="utf-8").strip() or "0")
    except (ValueError, OSError):
        return 0


def bump_data_version() -> int:
    with _DATA_VERSION_LOCK:
        v = get_data_version() + 1
        p = _data_version_path()
        try:
            p.parent.mkdir(parents=True, exist_ok=True)
            tmp = p.with_suffix(".version.tmp")
            tmp.write_text(str(v), encoding="utf-8")
            tmp.replace(p)
        except OSError:
            # A non-writable path must never turn a successful ingest into a
            # 500 (the bug this guards: FINANCIAL_DB_PATH=/data on Render's
            # diskless free plan). get_data_version() already tolerates a
            # missing/unreadable file (returns 0), so degrade gracefully:
            # log and return the in-memory increment without persisting.
            _cache_log.warning(
                "data_version write failed at %s (continuing un-persisted)",
                p, exc_info=True,
            )
            return v
    _cache_log.info("data_version bumped to %d", v)
    return v


def _cache_path() -> Path:
    import os
    if os.environ.get("DATABASE_URL"):
        return Path("/tmp") / "agentic_response_store.json"
    return Path(settings.response_store_path)


def _cache_load() -> dict[str, Any]:
    p = _cache_path()
    if not p.exists():
        return {}
    try:
        with p.open("r", encoding="utf-8") as f:
            data = json.load(f)
        return data if isinstance(data, dict) else {}
    except Exception:
        _cache_log.warning("response_store load failed; treating as empty", exc_info=True)
        return {}


def _cache_save(data: dict[str, Any]) -> None:
    p = _cache_path()
    p.parent.mkdir(parents=True, exist_ok=True)
    tmp = p.with_suffix(".json.tmp")
    with tmp.open("w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False, default=str)
    tmp.replace(p)


_PROMPT_VERSION = "v4-roman-hindi-2026-06-16"


def cache_key_for(
    question: str,
    *,
    conversation_id: str | None = None,
    tenant_id: str = "public",
    language: str = "en",
) -> str:
    """Cache key embeds (tenant_id, prompt_version, data_version, conversation_id, language, question).
    Bumping ``_PROMPT_VERSION`` invalidates every cached answer in one shot —
    used whenever the system prompt or formatter prompt changes meaningfully,
    so users immediately see the new behaviour instead of stale answers.

    ``tenant_id`` is placed FIRST so the key is unambiguously tenant-scoped:
    two tenants asking the identical question never collide on the same global
    cache file (which would leak one tenant's financial answers to another).
    It defaults to ``"public"`` so AUTH-off / single-tenant keys are byte-for-byte
    unchanged from before tenant scoping was added."""
    tenant = (tenant_id or "public").strip()
    convo = (conversation_id or "default").strip()
    # The answer language is part of the identity of a cached answer: the same
    # question in English vs Hindi vs Gujarati produces different prose, so they
    # must never collide on one cache entry (which would serve the wrong-language
    # answer back). "en" keeps the historical default.
    lang = (language or "en").strip().lower() or "en"
    version = get_data_version()
    norm_q = (question or "").strip().lower()
    payload = f"{tenant}|{_PROMPT_VERSION}|{version}|{convo}|{lang}|{norm_q}"
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def get_cached(key: str) -> dict[str, Any] | None:
    with _CACHE_LOCK:
        data = _cache_load()
        entry = data.get(key)
    return entry if isinstance(entry, dict) else None


# Hard cap on cached entries. The store was previously unbounded — every
# unique question grew the JSON file forever, and the whole file is
# rewritten on every put. 500 covers a useful working set without making
# the cache a multi-MB serialization tax. Override via env if needed.
_MAX_CACHE_ENTRIES = int(os.environ.get("RESPONSE_CACHE_MAX_ENTRIES") or 500)


def put_cached(key: str, record: dict[str, Any]) -> None:
    payload = dict(record)
    payload.setdefault(
        "stored_at", datetime.now(timezone.utc).isoformat(timespec="seconds"),
    )
    with _CACHE_LOCK:
        data = _cache_load()
        data[key] = payload
        # LRU-by-stored_at eviction. We can't track real reads cheaply in a
        # flat-file cache, so newest-stored wins — keeps recently-asked
        # questions and drops the long tail of one-offs.
        if len(data) > _MAX_CACHE_ENTRIES:
            ordered = sorted(
                data.items(),
                key=lambda kv: (kv[1] or {}).get("stored_at") or "",
            )
            drop_n = len(data) - _MAX_CACHE_ENTRIES
            for old_key, _ in ordered[:drop_n]:
                data.pop(old_key, None)
            _cache_log.info(
                "response cache trimmed: %d entries evicted (cap=%d)",
                drop_n, _MAX_CACHE_ENTRIES,
            )
        _cache_save(data)


def invalidate_all() -> int:
    with _CACHE_LOCK:
        data = _cache_load()
        n = len(data)
        _cache_save({})
    _cache_log.info("cache invalidated (%d entries removed)", n)
    return n


def cache_size() -> int:
    with _CACHE_LOCK:
        return len(_cache_load())


# ===========================================================================
# 9. MEMORY — entity-resolution synonyms (backing store for EntityResolver)
# ===========================================================================

_synonyms_log = logging.getLogger("agentic_ai.memory.synonyms")

_SYNONYM_DEFAULTS: dict[str, list[str]] = {
    "swiggy":      ["swiggy ltd", "bundl technologies", "swiggy app"],
    "zomato":      ["zomato ltd", "zomato app"],
    "groceries":   ["grocery", "kirana", "fmcg"],
    "electronics": ["consumer electronics", "appliances"],
    "fashion":     ["apparel", "clothing", "lifestyle"],
}


def _synonyms_path() -> Path:
    return Path(settings.synonyms_path)


def load_synonyms() -> dict[str, list[str]]:
    p = _synonyms_path()
    if not p.exists():
        try:
            p.parent.mkdir(parents=True, exist_ok=True)
            p.write_text(json.dumps(_SYNONYM_DEFAULTS, indent=2), encoding="utf-8")
        except Exception:
            _synonyms_log.warning("could not write default synonyms file", exc_info=True)
        return dict(_SYNONYM_DEFAULTS)
    try:
        return json.loads(p.read_text(encoding="utf-8"))
    except Exception:
        _synonyms_log.warning("synonyms.json unreadable; using defaults", exc_info=True)
        return dict(_SYNONYM_DEFAULTS)


def resolve_entities(question: str) -> list[dict]:
    if not question:
        return []
    q = question.lower()
    syns = load_synonyms()
    out: list[dict] = []
    for canonical, aliases in syns.items():
        hits = [a for a in [canonical, *aliases] if a.lower() in q]
        if hits:
            out.append({"canonical": canonical, "matched_aliases": hits})
    return out
