"""File parsers — CSV / XLSX with automatic header-row detection.

Real-world ERP exports often have:
  * Title / generated-on / company-name rows above the table.
  * One or more blank rows between the title and the header.
  * Multiple sheets where only one carries the data.

These parsers buffer the first MAX_HEADER_SCAN rows of each sheet (or the
CSV file), score them via `header_detection.find_header_row`, and only then
start streaming actual data rows.

For XLSX:
  1. wb.active is tried first (default).
  2. If wb.active doesn't yield a valid header row, every other sheet is
     scored and the highest-scoring sheet is used.
  3. If no sheet has a valid header → UploadError.
"""
from __future__ import annotations

import csv
import io
import logging
from pathlib import Path
from typing import Any, Iterator

from openpyxl import load_workbook

from app.upload.header_detection import find_header_row

log = logging.getLogger("agentic_ai.upload.parser")

MAX_HEADER_SCAN = 15


class UploadError(ValueError):
    """File-level failure: empty / unsupported / unreadable / no header."""


def _row_to_record(keys: list[str], row: Any) -> dict[str, Any]:
    if row is None:
        return {}
    out: dict[str, Any] = {}
    for i, k in enumerate(keys):
        if not k:
            continue
        v = row[i] if i < len(row) else None
        out[k] = v
    return out


# ---------- in-memory parsers (kept for callers that already hold bytes) ----

def parse_csv_bytes(content: bytes) -> tuple[list[str], dict[str, str], list[dict[str, Any]]]:
    text = content.decode("utf-8-sig", errors="replace")
    rows = list(csv.reader(io.StringIO(text)))
    if not rows:
        raise UploadError("CSV is empty")
    return _materialize(rows)


def parse_xlsx_bytes(content: bytes) -> tuple[list[str], dict[str, str], list[dict[str, Any]]]:
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        raise UploadError(f"Could not open xlsx: {e}")
    try:
        ws = _select_xlsx_sheet(wb)[1]
        rows = [list(r) if r is not None else [] for r in ws.iter_rows(values_only=True)]
        if not rows:
            raise UploadError("Selected sheet is empty")
        return _materialize(rows)
    finally:
        try:
            wb.close()
        except Exception:
            pass


def _materialize(rows: list[list[Any]]) -> tuple[list[str], dict[str, str], list[dict[str, Any]]]:
    """Detect header in `rows[:MAX_HEADER_SCAN]`, then build header + data list."""
    buffer = rows[:MAX_HEADER_SCAN]
    try:
        header_idx, header_cells, header_index = find_header_row(buffer)
    except ValueError as e:
        raise UploadError(str(e)) from e
    header = [str(c).strip() if c is not None else "" for c in header_cells]
    data = []
    for row in rows[header_idx + 1 :]:
        rec = _row_to_record(header, row)
        if rec:
            data.append(rec)
    return header, header_index, data


# ---------- streaming with auto-detection (used by /upload tempfile path) ---

def stream_parse_csv_with_detection(
    path: Path,
) -> tuple[list[str], dict[str, str], Iterator[dict[str, Any]]]:
    """Open `path`, detect header row in first MAX_HEADER_SCAN rows, return
    (header, header_index, generator).

    The generator owns the file handle: it closes the file in its `finally`
    block when exhausted (or when its `.close()` is called).
    """
    f = open(path, "r", encoding="utf-8-sig", errors="replace", newline="")
    try:
        reader = csv.reader(f)
        buffer: list[list[Any]] = []
        for _ in range(MAX_HEADER_SCAN):
            try:
                buffer.append(next(reader))
            except StopIteration:
                break
        if not buffer:
            f.close()
            raise UploadError("CSV is empty")
        try:
            header_idx, header_cells, header_index = find_header_row(buffer)
        except ValueError as e:
            f.close()
            raise UploadError(str(e)) from e
        header = [str(c).strip() if c is not None else "" for c in header_cells]
        # Anything in the buffer between the header row and end-of-buffer is
        # the first stretch of data; the rest comes from `reader`.
        data_buffer_tail = list(buffer[header_idx + 1 :])
    except Exception:
        f.close()
        raise

    def _gen() -> Iterator[dict[str, Any]]:
        try:
            for row in data_buffer_tail:
                rec = _row_to_record(header, row)
                if rec:
                    yield rec
            for row in reader:
                rec = _row_to_record(header, row)
                if rec:
                    yield rec
        finally:
            try:
                f.close()
            except Exception:
                pass

    return header, header_index, _gen()


def _select_xlsx_sheet(wb) -> tuple[str, Any, int, list[Any], dict[str, str]]:
    """Pick the sheet to read from. Returns (sheet_name, worksheet, header_idx,
    header_cells, header_index).

    Algorithm:
      1. Try wb.active first. If first MAX_HEADER_SCAN rows yield a valid
         header, use it (per spec: "default to first sheet").
      2. Otherwise, scan every sheet's first MAX_HEADER_SCAN rows; pick the
         highest-scoring valid sheet.
      3. If still none found, raise UploadError.
    """
    def _scan_sheet(name: str) -> tuple[int, int, list[Any], dict[str, str]] | None:
        ws_local = wb[name]
        buffer: list[list[Any]] = []
        for i, row in enumerate(ws_local.iter_rows(values_only=True)):
            if i >= MAX_HEADER_SCAN:
                break
            buffer.append(list(row) if row is not None else [])
        if not buffer:
            return None
        try:
            idx, cells, hi = find_header_row(buffer)
        except ValueError:
            return None
        score = len(hi)
        return score, idx, cells, hi

    active_name = wb.active.title if wb.active is not None else None
    if active_name:
        active_result = _scan_sheet(active_name)
        if active_result is not None:
            score, idx, cells, hi = active_result
            log.info("xlsx: using active sheet %r (score=%d, header_row=%d)",
                     active_name, score, idx + 1)
            return active_name, wb[active_name], idx, cells, hi
        log.info("xlsx: active sheet %r has no valid header — scanning others", active_name)

    candidates: list[tuple[int, str, int, list[Any], dict[str, str]]] = []
    for name in wb.sheetnames:
        if name == active_name:
            continue  # already tried
        result = _scan_sheet(name)
        if result is None:
            continue
        score, idx, cells, hi = result
        candidates.append((score, name, idx, cells, hi))

    if not candidates:
        raise UploadError(
            f"No sheet contained a valid header row in first {MAX_HEADER_SCAN} rows. "
            f"Sheets tried: {wb.sheetnames}"
        )
    candidates.sort(key=lambda x: -x[0])
    score, name, idx, cells, hi = candidates[0]
    log.info("xlsx: using sheet %r (score=%d, header_row=%d)", name, score, idx + 1)
    return name, wb[name], idx, cells, hi


def stream_parse_xlsx_with_detection(
    path: Path,
) -> tuple[list[str], dict[str, str], Iterator[dict[str, Any]], str]:
    """Open the XLSX, pick the best sheet, detect header, return
    (header, header_index, generator, sheet_name).
    """
    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    try:
        sheet_name, ws, header_idx, header_cells, header_index = _select_xlsx_sheet(wb)
        header = [str(c).strip() if c is not None else "" for c in header_cells]
    except Exception:
        try:
            wb.close()
        except Exception:
            pass
        raise

    def _gen() -> Iterator[dict[str, Any]]:
        try:
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i <= header_idx:
                    continue
                row_list = list(row) if row is not None else []
                rec = _row_to_record(header, row_list)
                if rec:
                    yield rec
        finally:
            try:
                wb.close()
            except Exception:
                pass

    return header, header_index, _gen(), sheet_name


# Backwards-compatible name used elsewhere if it shows up in tests.
def parse_file(filename: str, content: bytes):
    name = (filename or "").lower()
    if name.endswith(".csv"):
        return parse_csv_bytes(content)
    if name.endswith(".xlsx"):
        return parse_xlsx_bytes(content)
    raise UploadError(f"Unsupported file type: {filename!r} (use .csv or .xlsx)")
