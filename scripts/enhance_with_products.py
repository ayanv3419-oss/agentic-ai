"""Enhance the live sales dataset with a realistic Product Name column.

What it does:
  1. Reads every row from `data/financial_records.db.sales`.
  2. Assigns each row a real-world shoe-brand product. The distribution is
     weighted (a few bestsellers dominate, a long tail of niche brands
     appears occasionally) AND time-aware (some brands trend up, others
     decline, one has a mid-period spike) AND price-aware (premium brands
     never appear on a tiny ticket).
  3. UPDATEs each row's "Product Name" in the live DB.
  4. Exports the fully enhanced sales export to
     `data/sales_with_products.xlsx` for download / re-upload demos.

Run:
    python scripts/enhance_with_products.py

The script is idempotent — running it again re-assigns products using the
same fixed seed, so distributions stay stable across runs.
"""
from __future__ import annotations

import random
import sqlite3
import sys
from collections import Counter
from datetime import date, datetime
from pathlib import Path

# Allow importing app.* whether the script is launched from project root
# or from inside the scripts/ folder.
ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "backend"))

from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Alignment, Font, PatternFill  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

from app.database import (  # noqa: E402
    SCHEMA_COLUMNS,
    db_path,
    quoted,
    settings,
)

OUT_XLSX = ROOT / "data" / "sales_with_products.xlsx"

# Reproducible distribution.
random.seed(42)


# ---------------------------------------------------------------------------
# Catalog — REAL shoe / sportswear brands. Weights are RELATIVE; the script
# samples proportionally after applying time + price modifiers.
# ---------------------------------------------------------------------------

BASE_WEIGHTS: dict[str, int] = {
    # Tier 1 — bestsellers (~50% of volume)
    "Nike Air Max":            18,
    "Adidas Ultraboost":       14,
    "Puma RS-X":               10,
    "Skechers GoWalk":          8,
    # Tier 2 — core mid-range (~35%)
    "Reebok Classic":           7,
    "Converse Chuck Taylor":    6,
    "Vans Old Skool":           6,
    "Fila Disruptor":           5,
    "Jordan Retro 1":           5,
    "Asics Gel-Kayano":         4,
    # Tier 3 — niche (~12%)
    "New Balance 574":          3,
    "Under Armour HOVR":        3,
    "Jack & Jones Sneakers":    2,
    "Nike Revolution":          2,
    "Adidas Stan Smith":        2,
    # Tier 4 — premium / aspirational (~3%)
    "Yeezy Boost 350":          1,
    "Balenciaga Triple S":      1,
    "Gucci Ace":                1,
}

# Per-month multiplier — index 0=May, 1=June, 2=July (the data window).
# Anything outside the window uses the closest edge.
TIME_MULTIPLIERS: dict[str, tuple[float, float, float]] = {
    "Nike Air Max":          (0.9, 1.1, 1.5),  # rising hard in July
    "Adidas Ultraboost":     (1.2, 1.0, 0.9),  # gently cooling
    "Vans Old Skool":        (1.6, 1.0, 0.5),  # declining
    "Yeezy Boost 350":       (0.4, 2.5, 0.7),  # mid-period sneaker drop
    "Skechers GoWalk":       (1.0, 1.1, 1.2),  # gentle climb
    "Jordan Retro 1":        (0.8, 1.2, 1.5),  # building hype
    "Fila Disruptor":        (1.4, 1.0, 0.7),  # fading trend
    "Balenciaga Triple S":   (0.5, 0.5, 2.0),  # July premium push
    "Reebok Classic":        (1.3, 1.0, 0.8),  # spring favorite cooling
    "Adidas Stan Smith":     (1.4, 1.1, 0.8),  # post-summer dip
}

# Premium brands need a minimum ticket to feel realistic.
PRICE_FLOOR: dict[str, float] = {
    "Yeezy Boost 350":     2000,
    "Balenciaga Triple S": 2500,
    "Gucci Ace":           2500,
    "Jordan Retro 1":       900,
    "Adidas Ultraboost":    400,
    "Nike Air Max":         350,
    "Asics Gel-Kayano":     500,
    "Under Armour HOVR":    400,
}


def _month_idx(d: date) -> int:
    """Map a real date onto one of the three monthly buckets the catalog
    cares about (May/June/July 2025). Outside the window clamps to the
    nearest edge so older data still gets sensible weights."""
    if d.year < 2025:
        return 0
    if d.year > 2025:
        return 2
    if d.month <= 5:
        return 0
    if d.month == 6:
        return 1
    return 2  # July+


def pick_product(d: date, total_amount: float) -> str:
    """Choose a Product Name for this row using the weighted, time-aware,
    price-aware distribution above."""
    products: list[str] = []
    weights: list[float] = []
    midx = _month_idx(d)
    for product, base in BASE_WEIGHTS.items():
        floor = PRICE_FLOOR.get(product, 0.0)
        if total_amount < floor:
            continue
        mul = 1.0
        time_mul = TIME_MULTIPLIERS.get(product)
        if time_mul is not None:
            mul = time_mul[midx]
        weights.append(base * mul)
        products.append(product)
    if not products:
        # Every product was floor-gated out (very tiny ticket). Fall back
        # to the cheapest mainstream brand.
        return "Skechers GoWalk"
    return random.choices(products, weights=weights, k=1)[0]


# ---------------------------------------------------------------------------
# Main — UPDATE the DB + export the Excel companion file.
# ---------------------------------------------------------------------------

def main() -> None:
    db_file = Path(settings.financial_db_path)
    if not db_file.exists():
        raise SystemExit(f"financial_records.db not found at {db_file}")

    conn = sqlite3.connect(str(db_file))
    conn.row_factory = sqlite3.Row
    try:
        rows = conn.execute(
            f'SELECT id, "Date", "Total Amount" FROM {quoted("sales")} '
            f'ORDER BY "Date" ASC, id ASC'
        ).fetchall()
        if not rows:
            print("No rows in sales table — nothing to enhance.")
            return

        counter: Counter[str] = Counter()
        per_month: dict[int, Counter[str]] = {0: Counter(), 1: Counter(), 2: Counter()}
        updates: list[tuple[str, int]] = []

        for r in rows:
            try:
                d = datetime.strptime(str(r["Date"])[:10], "%Y-%m-%d").date()
            except (TypeError, ValueError):
                d = date(2025, 7, 1)
            try:
                amt = float(r["Total Amount"] or 0)
            except (TypeError, ValueError):
                amt = 0.0
            product = pick_product(d, amt)
            updates.append((product, int(r["id"])))
            counter[product] += 1
            per_month[_month_idx(d)][product] += 1

        cur = conn.cursor()
        cur.executemany(
            f'UPDATE {quoted("sales")} SET "Product Name" = ? WHERE id = ?',
            updates,
        )
        conn.commit()
        updated = cur.rowcount
        print(f"Updated {updated} rows in {db_file.name} with Product Name.")

        # ---- Distribution report --------------------------------------
        total = sum(counter.values())
        print("\nGlobal distribution:")
        for product, count in counter.most_common():
            pct = count / total * 100
            print(f"  {product:25s} {count:4d}  ({pct:5.1f}%)")

        print("\nMonthly distribution (May / Jun / Jul):")
        sorted_brands = [b for b, _ in counter.most_common()]
        print(f"  {'Brand':25s}  May  Jun  Jul")
        for brand in sorted_brands:
            m = per_month[0].get(brand, 0)
            j = per_month[1].get(brand, 0)
            jl = per_month[2].get(brand, 0)
            print(f"  {brand:25s}  {m:3d}  {j:3d}  {jl:3d}")

        # ---- Export Excel companion -----------------------------------
        export_excel(conn)
        print(f"\nWrote {OUT_XLSX}")
    finally:
        conn.close()


def export_excel(conn: sqlite3.Connection) -> None:
    """Pull every row (now with Product Name) and write a clean .xlsx."""
    cols = SCHEMA_COLUMNS  # canonical column order
    select = ", ".join(quoted(c) for c in cols)
    rows = conn.execute(
        f'SELECT {select} FROM {quoted("sales")} '
        f'ORDER BY "Date" ASC, "Invoice No" ASC'
    ).fetchall()

    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales"

    # Header row — bold + slight fill so it reads cleanly.
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="E5E7EB")
    for col_idx, col_name in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")

    # Data rows.
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, col_name in enumerate(cols, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row[col_name])

    # Reasonable column widths — auto-size each column up to 32 chars.
    for col_idx, col_name in enumerate(cols, start=1):
        max_len = len(str(col_name))
        for row in rows[:200]:  # sample first 200 rows for width hint
            v = row[col_name]
            if v is None:
                continue
            max_len = max(max_len, len(str(v)[:60]))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 32)

    # Freeze header row.
    ws.freeze_panes = "A2"
    wb.save(OUT_XLSX)


if __name__ == "__main__":
    main()
